"""Create the two Ichor workbooks in the reference report structure."""

import os

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from .constants import DETAIL_HEADERS, SUMMARY_COLUMNS

MONEY_FORMAT = '"$"#,##0.00'
THIN_SIDE = Side(style="thin", color="000000")
GRID_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

PALETTE = {
    "identity": "CCCCFF",
    "non_tax_non_cpf": "DBEEF3",
    "non_tax_cpf": "E2F0D9",
    "taxable_cpf": "FFF2CC",
    "total": "DDEBF7",
    "grand": "00B0F0",
}


def _value_or_none(value):
    return None if pd.isna(value) else value


def _excel_text(value):
    value = _value_or_none(value)
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _set_calc_on_open(workbook):
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True


def _style_detail_row(sheet, row, bold=False):
    for column in range(1, 10):
        cell = sheet.cell(row, column)
        cell.border = GRID_BORDER
        cell.font = Font(name="Calibri", size=11, bold=bold)
    sheet.cell(row, 1).number_format = "@"
    sheet.cell(row, 5).number_format = "dd/mm/yyyy"
    sheet.cell(row, 7).number_format = MONEY_FORMAT
    sheet.cell(row, 8).number_format = MONEY_FORMAT


def _write_detail_claim(sheet, row, claim):
    values = (
        claim["StaffID6"],
        _excel_text(claim["EmployeeName"]),
        _excel_text(claim["ReferenceNo"]),
        _excel_text(claim["Claim Type"]),
        claim["IncurredDate"].to_pydatetime(),
        _excel_text(claim["Service Provider"]),
        float(claim["IncurredAmount"]),
        float(claim["PaymentAmount"]),
        _excel_text(claim["Admin Remark"]),
    )
    for column, value in enumerate(values, start=1):
        sheet.cell(row, column, value)
    _style_detail_row(sheet, row)


def _configure_details_sheet(sheet):
    widths = (10, 39.54, 15, 96.36, 17.54, 41.54, 14, 16.18, 24)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.sheet_view.zoomScale = 70
    sheet.sheet_view.zoomScaleNormal = 70
    sheet.sheet_format.defaultRowHeight = 14.5
    sheet.row_dimensions[1].height = 15
    sheet.row_dimensions[3].height = 26
    sheet.page_setup.orientation = "landscape"
    sheet.page_margins = PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0)
    sheet.auto_filter.ref = "A3:I3"


def write_details(claims, pay_month, outdir):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Claim Report"
    _set_calc_on_open(workbook)
    _configure_details_sheet(sheet)

    month_label = pay_month.strftime("%B %Y")
    sheet["A1"] = f"Claims Report for {month_label} Reimbursement"
    sheet["A1"].font = Font(name="Arial", size=10, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="top")
    for column, label in enumerate(DETAIL_HEADERS, start=1):
        cell = sheet.cell(3, column, label)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.fill = PatternFill("solid", fgColor=PALETTE["identity"])
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = GRID_BORDER
        if column in (7, 8):
            cell.number_format = MONEY_FORMAT

    ordered = claims.assign(_sort_name=claims["EmployeeName"].str.casefold()).sort_values(
        ["_sort_name", "IncurredDate", "ReferenceNo"], kind="stable"
    )
    row = 4
    for (_, employee_name), employee_claims in ordered.groupby(
        ["StaffID6", "EmployeeName"], sort=False
    ):
        first_claim_row = row
        for _, claim in employee_claims.iterrows():
            _write_detail_claim(sheet, row, claim)
            row += 1
        sheet.cell(row, 2, f"{employee_name} Total")
        sheet.cell(row, 7, f"=SUBTOTAL(9,G{first_claim_row}:G{row - 1})")
        sheet.cell(row, 8, f"=SUBTOTAL(9,H{first_claim_row}:H{row - 1})")
        _style_detail_row(sheet, row, bold=True)
        row += 1

    sheet.cell(row, 2, "Grand Total")
    sheet.cell(row, 7, f"=SUBTOTAL(9,G4:G{row - 1})")
    sheet.cell(row, 8, f"=SUBTOTAL(9,H4:H{row - 1})")
    _style_detail_row(sheet, row, bold=True)

    path = os.path.join(outdir, f"Claims Details - {month_label}.xlsx")
    workbook.save(path)
    return path


def _merge_summary_headers(sheet):
    for reference in (
        "A2:A3",
        "B2:B3",
        "A4:B4",
        "C2:K2",
        "L2:O2",
        "P2:S2",
        "F3:G3",
        "H3:I3",
        "J3:K3",
        "L3:M3",
        "N3:O3",
        "P3:Q3",
        "R3:S3",
        "T2:T4",
        "U2:U4",
        "V2:V4",
        "W2:W4",
    ):
        sheet.merge_cells(reference)


def _write_summary_headers(sheet):
    _merge_summary_headers(sheet)
    labels = {
        "A2": "Employee ID",
        "B2": "Employee Name",
        "C2": "Non-Taxable \n& Non CPF Payable",
        "L2": "Non-Taxable \n& CPF Payable",
        "P2": "Taxable \n& CPF Payable ",
        "T2": "Grand Total",
        "U2": "Total for \nNon-Taxable \n& Non CPF Payable",
        "V2": "Total for \nNon-Taxable \n& CPF Payable",
        "W2": "Total for \nTaxable \n& CPF Payable",
        "A4": "Relation",
    }
    for cell, value in labels.items():
        sheet[cell] = value

    category_starts = {3, 4, 5, 6, 8, 10, 12, 14, 16, 18}
    for offset, (claim_type, relation) in enumerate(SUMMARY_COLUMNS, start=3):
        if offset in category_starts:
            sheet.cell(3, offset, claim_type)
        sheet.cell(4, offset, relation)


def _style_summary_headers(sheet):
    for row in range(2, 5):
        for column in range(1, 24):
            cell = sheet.cell(row, column)
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = GRID_BORDER
            cell.number_format = MONEY_FORMAT if column >= 3 else "General"
    fills = (
        (1, 2, "identity"),
        (3, 11, "non_tax_non_cpf"),
        (12, 15, "non_tax_cpf"),
        (16, 19, "taxable_cpf"),
        (20, 23, "total"),
    )
    for start, end, color_key in fills:
        fill = PatternFill("solid", fgColor=PALETTE[color_key])
        for row in range(2, 5):
            for column in range(start, end + 1):
                sheet.cell(row, column).fill = fill


def _configure_summary_sheet(sheet):
    widths = {
        "A": 11.91,
        "B": 37.18,
        "C": 9.82,
        "D": 9.82,
        "E": 17.36,
        "F": 10.82,
        "G": 11,
        "H": 10.82,
        "I": 11,
        "J": 10.82,
        "K": 11,
        "L": 9.82,
        "M": 11,
        "N": 9.82,
        "O": 11,
        "P": 13,
        "Q": 14.36,
        "R": 10.82,
        "S": 11,
        "T": 11.91,
        "U": 17.63,
        "V": 13.45,
        "W": 13.45,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[2].height = 54
    sheet.row_dimensions[3].height = 72
    sheet.row_dimensions[4].height = 30
    sheet.page_setup.orientation = "landscape"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0


def _employee_summary_rows(claims):
    rows = []
    for (staff_id, employee_name), employee_claims in claims.groupby(
        ["StaffID6", "EmployeeName"], sort=True
    ):
        amounts = employee_claims.groupby(["SummaryType", "SummaryRelation"])[
            "PaymentAmount"
        ].sum()
        categories = [round(float(amounts.get(key, 0)), 2) for key in SUMMARY_COLUMNS]
        non_tax_non_cpf = round(sum(categories[:9]), 2)
        non_tax_cpf = round(sum(categories[9:13]), 2)
        taxable_cpf = round(sum(categories[13:]), 2)
        rows.append(
            [staff_id, _excel_text(employee_name)]
            + [value if value else None for value in categories]
            + [round(non_tax_non_cpf + non_tax_cpf + taxable_cpf, 2)]
            + [non_tax_non_cpf, non_tax_cpf, taxable_cpf]
        )
    return rows


def _style_summary_data(sheet, start_row, end_row, total_row):
    for row in range(start_row, total_row + 1):
        for column in range(1, 24):
            cell = sheet.cell(row, column)
            cell.border = GRID_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="Calibri", size=11, bold=(row == total_row))
            if column >= 3:
                cell.number_format = MONEY_FORMAT
    total_fill = PatternFill("solid", fgColor=PALETTE["non_tax_non_cpf"])
    for column in range(1, 24):
        sheet.cell(total_row, column).fill = total_fill
    sheet.cell(total_row, 20).fill = PatternFill("solid", fgColor=PALETTE["grand"])
    for row in range(start_row, end_row + 1):
        sheet.cell(row, 1).number_format = "@"


def write_summary(claims, pay_month, outdir):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet"
    _set_calc_on_open(workbook)
    _configure_summary_sheet(sheet)
    _write_summary_headers(sheet)
    _style_summary_headers(sheet)

    rows = _employee_summary_rows(claims)
    start_row = 5
    for row_number, values in enumerate(rows, start=start_row):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, value)

    total_row = start_row + len(rows)
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    sheet.cell(total_row, 1, "Grand Total")
    for column in range(3, 24):
        total = round(sum(float(row[column - 1] or 0) for row in rows), 2)
        sheet.cell(total_row, column, total)
    _style_summary_data(sheet, start_row, total_row - 1, total_row)

    month_label = pay_month.strftime("%B %Y")
    path = os.path.join(outdir, f"Claims Summary - {month_label}.xlsx")
    workbook.save(path)
    return path
