"""STM (STMicroelectronics) Flex Report adapter.

Transformation logic verified against the May 2026 manual outputs (1,154 payroll
rows, totals to the cent). Generates the monthly medical reimbursement pack:
Summary Reimbursement Report, IT15 payroll upload, refreshed Mapping Master,
per-country SEA reports and — when checks fire — a Validation Exceptions Report.
"""
import os
import shutil
import warnings
from copy import copy
from datetime import datetime
from itertools import combinations

import openpyxl
import pandas as pd
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from flex_services.errors import FlexInputError

# openpyxl warns about unsupported extensions in client-authored workbooks; those
# are expected here. Scoped to openpyxl so other modules keep their warnings.
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

PAYROLL_TEMPLATE_SHEET = 'SG1xPaymtTemplate'

# Saved IT15 payroll template. Drop a blank template here (header block, formulas and
# row-20 formatting, no employee data) and the monthly run only needs the three data
# files. Any employee data present in the template is cleared before writing, so a
# prior-month IT15 file works here too.
TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
DEFAULT_IT15_TEMPLATE = os.path.join(TEMPLATE_DIR, 'stm_it15_template.xlsx')


def saved_template():
    """Path of the saved IT15 template, or None when none has been placed on the server."""
    return DEFAULT_IT15_TEMPLATE if os.path.exists(DEFAULT_IT15_TEMPLATE) else None


COMPANY = {
    "id": "stm",
    "name": "STMicroelectronics (STM)",
    "status": "active",
    "files": [
        {"key": "claims",   "label": "Employee claims export", "required": True},
        {"key": "leavers",  "label": "STM Leavers file", "required": True},
        {"key": "listing",  "label": "Employee listing report", "required": True},
    ],
    "notes": (
        "Generates: Summary Reimbursement Report, IT15 payroll upload, refreshed Mapping Master, "
        "SEA reports (TH/ID/VN) when applicable, plus a Validation Exceptions Report when issues are found. "
        "The saved IT15 payroll template is applied automatically."
    ),
}

# Leavers file column layout (header row is the second row of the sheet).
LEAVERS_COLUMNS = ['EmpID', 'Name', 'LastDay', 'CostCentre', 'Company', 'HealthS', 'NMedi',
                   'Optical', 'Total', 'LastClaimMonth', 'Clawback', 'EmailDate']

# Exhaustive subset matching is combinatorial; above this many claims in one
# category the search is skipped and the category is flagged for manual check.
MAX_COMBINATION_ITEMS = 15

# Errors that invalidate every claim of the employee, not just the offending row.
EMPLOYEE_LEVEL_ERRORS = ('MISSING COST CENTRE', 'INVALID STAFF ID')

# Hard errors that physically hold claim rows OUT of both output files (they need a source
# fix and a re-run before those amounts can be submitted anywhere).
HELD_ERROR_CHECKS = frozenset({
    'INVALID STAFF ID', 'MISSING COST CENTRE', 'UNKNOWN ENTITY', 'UNKNOWN CLAIM TYPE',
    'NEGATIVE AMOUNT', 'DUPLICATE CLAIM REF',
})

# One-line operator guidance per check, used to drive the include/exclude decision UI.
HELD_GUIDANCE = {
    'INVALID STAFF ID': 'Staff ID is not a usable 6-digit EEID. Correct it in the claims export and re-run.',
    'MISSING COST CENTRE': 'No Cost Centre anywhere for this employee. Add it to the listing/master and re-run.',
    'UNKNOWN ENTITY': 'Entity has no Legal Entity Code mapping. Add the mapping and re-run.',
    'UNKNOWN CLAIM TYPE': 'Claim type is unmapped to a wage code. Confirm the type and re-run.',
    'NEGATIVE AMOUNT': 'Negative amount (possible clawback). Handle manually — not placed on payroll.',
    'DUPLICATE CLAIM REF': 'Same claim reference appears more than once. De-duplicate the export and re-run.',
}
WARN_GUIDANCE = {
    'COST CENTRE FROM FALLBACK': 'Cost Centre came from a historical mapping — verify it is still current. Row is on the payroll.',
    'ZERO/BLANK AMOUNT': 'Zero/blank amount — excluded from outputs automatically. No action needed.',
    'NOT APPROVED': 'Not approved — excluded from outputs automatically. No action needed.',
    'COST CENTRE / ENTITY MISMATCH': 'Cost Centre prefix and claim entity disagree. Verify the entity with HR. Row is on the payroll.',
    'LEAVER MISSING LAST DAY': 'Leaver has claims but no last employment day. Fill it in the leavers file. Row is on the payroll.',
    'LEAVER CATEGORY MISMATCH': 'Amount matched across categories and EXCLUDED from payroll. Ask HR to fix the category in the leavers file.',
}

# Payroll-user helper columns in the IT15 sheet (P = EE ID, Q = Cost Center, R = WT
# lookup). Owned by payroll; only extended when new data runs past the template's rows.
HELPER_COLUMNS = (16, 17, 18)


def _require_columns(df, columns, label):
    missing = [c for c in columns if c not in df.columns]
    if missing:
        raise FlexInputError(
            f"{label}: missing required column(s): {', '.join(missing)}. "
            f"Found: {', '.join(str(c) for c in df.columns[:20])}"
        )


def _normalise_eeid(series):
    """Six-digit EEID from an ID column, whatever dtype pandas inferred.

    A single blank cell (a trailing empty row is enough) makes pandas type the whole
    column as float64, so a plain ``astype(str).zfill(6)`` yields '1001.0' for every
    row and nothing matches. Numeric values are formatted from the number; anything
    non-numeric is kept as text so the INVALID STAFF ID check can report it.
    """
    numeric = pd.to_numeric(series, errors='coerce')
    text = series.astype(str).str.strip().replace({'nan': '', 'None': '', 'NaT': '', '<NA>': ''})
    formatted = numeric.map(lambda v: f'{int(v):06d}' if pd.notna(v) else None)
    return formatted.fillna(text.str.zfill(6)).astype(str)


def _lec_label(value):
    """Format a Legal Entity Code as 4 digits; blank when unmapped (never crash)."""
    if pd.isna(value):
        return ''
    return f'{int(value):04d}'


def run(files, pay_month, outdir):
    CLAIMS_F, LEAVERS_F, LISTING_F = files['claims'], files['leavers'], files['listing']
    MASTER_F = files.get('master')
    TEMPLATE_F = files.get('template') or saved_template()
    if not TEMPLATE_F:
        raise FlexInputError(
            "No IT15 payroll template available. Upload the IT15 file with this run, or save a blank "
            f"template (header block, formulas and row-{20} formatting, no employee data) on the server at "
            f"{DEFAULT_IT15_TEMPLATE} so future runs only need the three data files."
        )
    USING_SAVED_TEMPLATE = not files.get('template')
    PAY_MONTH = datetime.fromisoformat(pay_month)
    OUTDIR = outdir
    MON = PAY_MONTH.strftime('%b')
    MMYYYY = PAY_MONTH.strftime('%m%Y')
    LOG = []

    def echo(*a):
        LOG.append(' '.join(str(x) for x in a))

    # Rule 2: Entity -> Legal Entity Code
    ENTITY_LEC = {
        'STMICROELECTRONICS ASIA PACIFIC PTE LTD - G0005086': 2800,
        'STMICROELECTRONICS PTE LTD AMK - G0005088': 800,
        'STMICROELECTRONICS PTE LTD TPY - G0005089': 800,
    }

    # Rule 4: Claim Type -> Wage Code
    def map_code(ct):
        ct_l = str(ct).lower()
        if 'optical' in ct_l:                                   return 'Optical(T)'
        if 'childcare' in ct_l or 'health screening' in ct_l or ct_l.startswith('healths'): return 'HealthS/ChildC(NT)'
        return 'N-Medi/Dental(NT)'   # Dental / Medical-related / Outpatient GP / Polyclinic / other medical

    # SEA locations -> legal entity code label
    SEA_MAP = {'Bangkok': ('Thailand', '7900TH'), 'Hanoi': ('Vietnam', '0800VN'), 'JAKARTA': ('Indonesia', '2800INDON')}

    # ---------------- LOAD ----------------
    claims = pd.read_excel(CLAIMS_F)
    _require_columns(claims, ['Staff ID', 'Employee Name', 'Reference No.', 'Entity', 'Claim Type',
                              'Payment Amt', 'Status', 'Incurred Date'], 'Employee claims export')
    claims['EEID6'] = _normalise_eeid(claims['Staff ID'])

    listing = pd.read_excel(LISTING_F)
    _require_columns(listing, ['User ID', 'Cost Centre', 'Location Description', 'Last Day of Service'],
                     'Employee listing report')
    listing['EEID6'] = _normalise_eeid(listing['User ID'])
    listing = listing.drop_duplicates('EEID6')

    lv = pd.read_excel(LEAVERS_F, header=1)
    if len(lv.columns) < len(LEAVERS_COLUMNS):
        raise FlexInputError(
            f"STM Leavers file: expected at least {len(LEAVERS_COLUMNS)} columns "
            f"({', '.join(LEAVERS_COLUMNS)}) with the header on row 2, found {len(lv.columns)}"
        )
    lv.columns = LEAVERS_COLUMNS + list(lv.columns[len(LEAVERS_COLUMNS):])
    lv['EEID6'] = _normalise_eeid(lv['EmpID'])
    lv = lv[lv['EEID6'].str.match(r'\d{6}')].drop_duplicates('EEID6', keep='last')

    # ---- Cumulative mapping master (fallback source) ----
    if MASTER_F:
        master = pd.read_excel(MASTER_F)
        _require_columns(master, ['EEID', 'Cost Centre'], 'STM Mapping Master')
        master = master.dropna(subset=['EEID'])
        master['EEID6'] = _normalise_eeid(master['EEID'])
        master = master.drop_duplicates('EEID6', keep='last')
    else:
        master = pd.DataFrame(columns=['EEID6', 'Name', 'Cost Centre', 'Entity', 'Legal Entity Code', 'Source', 'Last Updated'])
    master_cc = master.dropna(subset=['Cost Centre']).set_index('EEID6')['Cost Centre'].to_dict()
    lv_cc = lv[lv['CostCentre'].astype(str).str.match(r'[A-Z]{2}\d{4}', na=False)].set_index('EEID6')['CostCentre'].to_dict()

    # ---------------- VALIDATE ----------------
    def is_known_type(ct):
        s = str(ct).strip().lower()
        keywords = ['optical', 'childcare', 'health screening', 'dental', 'medical-related',
                    'outpatient gp', 'polyclinic']
        return any(k in s for k in keywords)

    # ---------------- VALIDATION ENGINE ----------------
    def run_validations(claims, listing, lv, ENTITY_LEC, FALLBACK_CC):
        """Returns (errors_df, warnings_df). Errors block affected rows; warnings don't.

        Every entry records the source row index so rows are held out by identity, not by
        Claim Reference - a claim with a blank reference must still be held.
        """
        E, W = [], []
        # Employees present in the listing but with a blank Cost Centre cell cannot be
        # grouped into the payroll file, so they are treated as missing, not as present.
        listing_cc = listing.set_index('EEID6')['Cost Centre']
        lset = set(listing_cc[listing_cc.notna() & (listing_cc.astype(str).str.strip() != '')].index)

        def add(bucket, check, ref, eeid, name, detail, row=None):
            bucket.append({'Check': check, 'Claim Reference': ref, 'EEID': eeid, 'Name': name,
                           'Detail': detail, '_row': row})

        for idx, r in claims.iterrows():
            ref, e6, nm = r.get('Reference No.'), r['EEID6'], r.get('Employee Name')
            # 0. Staff ID must be a usable 6-digit employee number
            if not str(e6).isdigit():
                add(E, 'INVALID STAFF ID', ref, e6, nm,
                    f"Staff ID '{r.get('Staff ID')}' is not numeric - cannot be used as an EEID", idx)
                continue
            # 1. Missing cost centre (not in listing, or in the listing with a blank cell)
            if e6 not in lset:
                if e6 in FALLBACK_CC:
                    add(W, 'COST CENTRE FROM FALLBACK', ref, e6, nm,
                        f"No usable Cost Centre in the current employee listing; used historical mapping: "
                        f"{FALLBACK_CC[e6]} - verify still current", idx)
                else:
                    add(E, 'MISSING COST CENTRE', ref, e6, nm,
                        'No Cost Centre in the employee listing, master mapping, or leavers file - '
                        'the claim cannot be placed on the payroll file', idx)
            # 2. Unknown entity
            if r['Entity'] not in ENTITY_LEC:
                add(E, 'UNKNOWN ENTITY', ref, e6, nm, f"No Legal Entity Code mapping for: {r['Entity']}", idx)
            # 3. Unknown claim type (would otherwise silently default)
            if not is_known_type(r['Claim Type']):
                add(E, 'UNKNOWN CLAIM TYPE', ref, e6, nm,
                    f"Unmapped claim type: {r['Claim Type']} - confirm wage code", idx)
            # 4. Amount problems
            amt = r.get('Payment Amt')
            if pd.isna(amt) or amt == 0:
                add(W, 'ZERO/BLANK AMOUNT', ref, e6, nm, f'Payment Amt = {amt}', idx)
            elif amt < 0:
                add(E, 'NEGATIVE AMOUNT', ref, e6, nm, f'Payment Amt = {amt} - clawback? handle manually', idx)
            # 5. Status
            if r.get('Status') != 'Approved':
                add(W, 'NOT APPROVED', ref, e6, nm, f"Status = {r.get('Status')} - excluded from outputs", idx)

        # 6. Duplicate claim references (blank references are not duplicates of each other)
        refs = claims['Reference No.']
        dups = claims[refs.notna() & refs.duplicated(keep=False)]
        for idx, r in dups.iterrows():
            add(E, 'DUPLICATE CLAIM REF', r['Reference No.'], r['EEID6'], r.get('Employee Name'),
                'Same reference appears more than once', idx)

        # 7. Cost-centre prefix vs entity cross-check
        PFX = {'AM': 800, 'AP': 800, 'GO': 800, 'TP': 800, 'SH': 2800, 'LY': 2800}
        mrg = claims.merge(listing[['EEID6', 'Cost Centre']], on='EEID6', how='left')
        mrg['exp'] = mrg['Cost Centre'].astype(str).str[:2].map(PFX)
        mrg['lec'] = mrg['Entity'].map(ENTITY_LEC)
        bad = mrg[mrg['exp'].notna() & mrg['lec'].notna() & (mrg['exp'] != mrg['lec'])].drop_duplicates('EEID6')
        for _, r in bad.iterrows():
            add(W, 'COST CENTRE / ENTITY MISMATCH', None, r['EEID6'], r.get('Employee Name'),
                f"Cost centre {r['Cost Centre']} implies LEC {int(r['exp']):04d} but claims entity implies "
                f"{int(r['lec']):04d} - verify with HR")

        # 8. Leaver data quality: leaver with claims but no last-day date
        lv_claims = lv[lv['EEID6'].isin(set(claims['EEID6']))]
        for _, r in lv_claims[lv_claims['LastDay'].isna()].iterrows():
            add(W, 'LEAVER MISSING LAST DAY', None, r['EEID6'], r.get('Name'),
                'In leavers file but Last employment day blank')

        return pd.DataFrame(E), pd.DataFrame(W)

    FALLBACK_CC = {**master_cc, **lv_cc}
    errors_df, warnings_df = run_validations(claims, listing, lv, ENTITY_LEC, FALLBACK_CC)

    # exclude non-approved from processing (warned above)
    claims = claims[claims['Status'] == 'Approved']
    # exclude hard-error rows from outputs (they go to the exception report instead).
    # Rows are held by source-row identity, so a claim with a blank Reference No. is
    # still held; EMPLOYEE_LEVEL_ERRORS additionally hold every claim of that employee.
    if len(errors_df):
        err_rows = set(errors_df['_row'].dropna())
        err_eeids = set(errors_df.loc[errors_df['Check'].isin(EMPLOYEE_LEVEL_ERRORS), 'EEID'])
    else:
        err_rows, err_eeids = set(), set()
    held_mask = claims.index.isin(err_rows) | claims['EEID6'].isin(err_eeids)
    held = claims[held_mask]
    claims = claims[~held_mask]

    # ---------------- ENRICH ----------------
    df = claims.merge(listing[['EEID6', 'Cost Centre', 'Location Description', 'Last Day of Service']],
                      on='EEID6', how='left')
    df['Cost Centre'] = df['Cost Centre'].fillna(df['EEID6'].map(FALLBACK_CC))
    df = df.merge(lv[['EEID6', 'LastDay', 'LastClaimMonth', 'EmailDate']], on='EEID6', how='left')

    df['Legal Entity Code'] = df['Entity'].map(ENTITY_LEC)
    df['Code'] = df['Claim Type'].map(map_code)
    df['Amount'] = df['Payment Amt'].round(2)
    df['Termination Date'] = df['LastDay'].fillna(df['Last Day of Service'])
    df['Active/Inactive'] = df['Termination Date'].notna().map({True: 'Inactive', False: 'Active'})

    # Rule (claim-level, prevents double pay): the leavers file records per-category amounts
    # already emailed to STM HR for final-pay processing. Identify WHICH claims those amounts
    # cover (earliest-incurred first, exact subset match). Covered claims: shown in the
    # breakdown with the email date, EXCLUDED from the payroll IT15. Uncovered claims of the
    # same leaver flow to payroll with a blank email-date column.
    def _cat(code):
        return {'Optical(T)': 'Optical', 'HealthS/ChildC(NT)': 'HealthS'}.get(code, 'NMedi')

    def _find_subset(items, target):
        t = round(float(target), 2)
        if t == 0:
            return []
        run_total = 0.0
        pref = []
        for idx, a in items:
            run_total = round(run_total + a, 2)
            pref.append(idx)
            if abs(run_total - t) < 0.005:
                return pref
        if len(items) > MAX_COMBINATION_ITEMS:
            return None
        for k in range(1, len(items) + 1):
            for comb in combinations(items, k):
                if abs(round(sum(a for _, a in comb), 2) - t) < 0.005:
                    return [i for i, _ in comb]
        return None

    df['LeaverCat'] = df['Code'].map(_cat)
    df['CoveredByLeaverEmail'] = False
    df['LeaverUnreconciled'] = False
    leaver_mismatches = []
    leaver_warns = []
    lv_idx = lv.set_index('EEID6')
    for e6 in df.loc[df['EEID6'].isin(lv_idx.index), 'EEID6'].unique():
        lrow = lv_idx.loc[e6]
        for cat in ('HealthS', 'NMedi', 'Optical'):
            target = lrow[cat] if pd.notna(lrow[cat]) else 0
            if float(target) == 0:
                continue

            def _cross_category_match():
                """Leavers file sometimes records an amount under the wrong category (seen in
                May: an Optical claim logged in the NMedi column). Try to match the amount
                against ALL of this employee's not-yet-covered claims, any category."""
                allsub = df[(df['EEID6'] == e6) & (~df['CoveredByLeaverEmail'])].sort_values(
                    ['Incurred Date', 'Reference No.'])
                if allsub.empty:
                    return False
                xmatch = _find_subset(list(zip(allsub.index, allsub['Payment Amt'].round(2))), target)
                if xmatch is None:
                    return False
                df.loc[xmatch, 'CoveredByLeaverEmail'] = True
                cats_found = sorted(df.loc[xmatch, 'LeaverCat'].unique())
                leaver_warns.append((e6, allsub['Employee Name'].iloc[0],
                    f"Leaver amount {float(target):.2f} recorded under {cat} but matched {'/'.join(cats_found)} "
                    f"claim(s) - excluded from payroll; ask HR to correct the leavers file category"))
                return True

            sub = df[(df['EEID6'] == e6) & (df['LeaverCat'] == cat)].sort_values(['Incurred Date', 'Reference No.'])
            if sub.empty:
                # An amount was already emailed to HR for final pay but no claim of that
                # category is in this export. It may have been logged under the wrong
                # category, so try the cross-category match before giving up - otherwise
                # the claim it actually covers would be paid a second time via payroll.
                if _cross_category_match():
                    continue
                other = df[(df['EEID6'] == e6) & (~df['CoveredByLeaverEmail'])]
                df.loc[other.index, 'LeaverUnreconciled'] = True
                leaver_mismatches.append((e6, other, lrow.get('Name'), cat, target,
                    f"Leavers file shows {float(target):.2f} emailed under {cat} but this export has no {cat} "
                    f"claims - cannot tell which claim it covers, so this leaver's remaining claims are blocked "
                    f"from payroll pending manual check (double-pay risk)"))
                continue

            match = _find_subset(list(zip(sub.index, sub['Payment Amt'].round(2))), target)
            if match is None:
                if _cross_category_match():
                    continue
                df.loc[sub.index, 'LeaverUnreconciled'] = True
                leaver_mismatches.append((e6, sub, sub['Employee Name'].iloc[0], cat, target,
                    f"Leaver amount {float(target):.2f} does not match any combination of current {cat} claims "
                    f"(claims total {sub['Payment Amt'].round(2).sum():.2f}) - double-pay risk, category excluded "
                    f"from payroll pending manual check"))
            else:
                df.loc[match, 'CoveredByLeaverEmail'] = True
    # email date only shown on rows actually covered by the leavers email
    df.loc[~df['CoveredByLeaverEmail'], 'EmailDate'] = pd.NaT
    df['FinalPayThisCycle'] = df['CoveredByLeaverEmail'] | df['LeaverUnreconciled']

    # SEA split
    df['SEA'] = df['Location Description'].map(lambda x: SEA_MAP.get(x, (None, None))[0])
    sea_df = df[df['SEA'].notna()].copy()
    sg_df = df[df['SEA'].isna()].copy()

    # ---------------- BREAKDOWN (SG) ----------------
    def build_breakdown(d):
        out = pd.DataFrame({
            'EEID': d['EEID6'].astype(int),
            'Name': d['Employee Name'],
            'Cost Centre': d['Cost Centre'],
            'Entity': d['Entity'],
            'Legal Entity Code': d['Legal Entity Code'].map(_lec_label),
            'Claim Reference': d['Reference No.'],
            'Code': d['Code'],
            'Month of Payment': PAY_MONTH,
            'Amount': d['Amount'],
            'Currency': 'SGD',
            'Termination Date': pd.to_datetime(d['Termination Date']),
            'Active/Inactive': d['Active/Inactive'],
            'Date Email sent to STM': pd.to_datetime(d['EmailDate']),
            '_final_pay': d['FinalPayThisCycle'],
            '_covered': d['CoveredByLeaverEmail'],
            '_unrec': d['LeaverUnreconciled'],
        })
        ent_order = {'STMICROELECTRONICS ASIA PACIFIC PTE LTD - G0005086': 0,
                     'STMICROELECTRONICS PTE LTD AMK - G0005088': 1,
                     'STMICROELECTRONICS PTE LTD TPY - G0005089': 2}
        out['_a'] = (out['Active/Inactive'] == 'Active').astype(int)   # Inactive first
        out['_e'] = out['Entity'].map(ent_order)
        out['_n'] = out['Name'].astype(str).str.lower()
        out = out.sort_values(['_a', '_e', '_n', 'Claim Reference']).drop(columns=['_a', '_e', '_n']).reset_index(drop=True)
        return out

    bd = build_breakdown(sg_df)

    # register leaver mismatches as validation errors (rows remain in breakdown, blocked from payroll)
    if leaver_warns:
        lw_rows = [{'Check': 'LEAVER CATEGORY MISMATCH', 'Claim Reference': None, 'EEID': e6, 'Name': nm, 'Detail': det}
                   for (e6, nm, det) in leaver_warns]
        warnings_df = pd.concat([warnings_df, pd.DataFrame(lw_rows)], ignore_index=True)
    if leaver_mismatches:
        mm_rows = [{'Check': 'LEAVER AMOUNT MISMATCH', 'Claim Reference': None, 'EEID': e6, 'Name': nm, 'Detail': det}
                   for (e6, _sub, nm, cat, tgt, det) in leaver_mismatches]
        errors_df = pd.concat([errors_df, pd.DataFrame(mm_rows)], ignore_index=True)

    # ---------------- REFRESH MAPPING MASTER ----------------
    cur = bd.drop_duplicates('EEID', keep='first')[['EEID', 'Name', 'Cost Centre', 'Entity', 'Legal Entity Code']].copy()
    cur['EEID6'] = cur['EEID'].astype(str).str.zfill(6)
    cur['Source'] = f'{PAY_MONTH:%b %Y} breakdown report'
    cur['Last Updated'] = pd.Timestamp.now().normalize()
    keep_old = master[~master['EEID6'].isin(set(cur['EEID6']))]
    new_master = pd.concat([cur, keep_old], ignore_index=True)
    new_master = new_master[['EEID6', 'Name', 'Cost Centre', 'Entity', 'Legal Entity Code', 'Source', 'Last Updated']]
    new_master = new_master.rename(columns={'EEID6': 'EEID'}).sort_values('EEID')
    master_path = os.path.join(OUTDIR, 'STM_Mapping_Master.xlsx')
    new_master.to_excel(master_path, index=False)

    # ---------------- WRITE SUMMARY REPORT ----------------
    def write_summary_report(bd, path, currency='SGD'):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Summary'
        ws2 = wb.create_sheet(f'{currency} - Breakdown report')

        cols = ['EEID', 'Name', 'Cost Centre', 'Entity', 'Legal Entity Code', 'Claim Reference', 'Code',
                'Month of Payment', 'Amount', 'Currency', 'Termination Date', 'Active/Inactive',
                'Date Email sent to STM']
        thin = Border(*[Side(style='thin')] * 4)
        hdr_font = Font(name='Arial', size=10, bold=True)
        body_font = Font(name='Arial', size=10)
        fill = PatternFill('solid', start_color='D9E1F2')
        for j, cn in enumerate(cols, start=1):
            cell = ws2.cell(1, j, cn)
            cell.font = hdr_font
            cell.fill = fill
            cell.border = thin
        for i, row in bd.iterrows():
            r = i + 2
            for j, cn in enumerate(cols, start=1):
                v = row[cn]
                if pd.isna(v):
                    v = None
                if isinstance(v, pd.Timestamp):
                    v = v.to_pydatetime()
                cell = ws2.cell(r, j, v)
                cell.font = body_font
                cell.border = thin
                if cn in ('Month of Payment', 'Termination Date', 'Date Email sent to STM'):
                    cell.number_format = 'DD/MM/YYYY'
                if cn == 'Amount':
                    cell.number_format = '#,##0.00'
        widths = [8, 30, 11, 46, 15, 15, 17, 15, 10, 9, 15, 14, 20]
        for j, w in enumerate(widths, start=1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

        n = max(len(bd), 1)
        sh = f"'{ws2.title}'"
        ws.cell(3, 1, 'Legal Entity Code')
        ws.cell(3, 2, 'Active')
        ws.cell(3, 3, 'Inactive\n - Includes Leavers ')
        ws.cell(3, 4, 'Grand Total')
        codes = sorted(bd['Legal Entity Code'].astype(str).unique())
        r = 4
        for lec in codes:
            ws.cell(r, 1, lec)
            ws.cell(r, 2, f'=SUMIFS({sh}!I2:I{n+1},{sh}!E2:E{n+1},"{lec}",{sh}!L2:L{n+1},"Active")')
            ws.cell(r, 3, f'=SUMIFS({sh}!I2:I{n+1},{sh}!E2:E{n+1},"{lec}",{sh}!L2:L{n+1},"Inactive")')
            ws.cell(r, 4, f'=B{r}+C{r}')
            r += 1
        ws.cell(r, 1, 'Grand Total')
        ws.cell(r, 2, f'=SUM(B4:B{r-1})')
        ws.cell(r, 3, f'=SUM(C4:C{r-1})')
        ws.cell(r, 4, f'=SUM(D4:D{r-1})')
        for row in ws.iter_rows(min_row=3, max_row=r, min_col=1, max_col=4):
            for cell in row:
                cell.font = Font(name='Arial', size=10, bold=(cell.row in (3, r)))
                cell.border = thin
                if cell.row == 3:
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                elif cell.column > 1:
                    cell.number_format = '#,##0.00'
        for col, w in zip('ABCD', [16, 13, 22, 13]):
            ws.column_dimensions[col].width = w
        wb.save(path)

    sum_path = os.path.join(OUTDIR, f'STM - {MON} Summary Reimbursment Report.xlsx')
    write_summary_report(bd.drop(columns=[c0 for c0 in ('_final_pay', '_covered', '_unrec') if c0 in bd]), sum_path)

    # ---------------- PAYROLL IT15 ----------------
    pay = bd[~bd['_final_pay']].copy()
    excluded = bd[bd['_covered']]
    blocked = bd[bd['_unrec']]
    # dropna=False: a blank grouping key (e.g. an unmapped Cost Centre) must never make a
    # claim disappear from the payroll file while still counting in the summary report.
    agg = pay.groupby(['EEID', 'Name', 'Cost Centre', 'Legal Entity Code', 'Code'],
                      as_index=False, dropna=False)['Amount'].sum()
    code_order = {'HealthS/ChildC(NT)': 0, 'N-Medi/Dental(NT)': 1, 'Optical(T)': 2}
    agg['_c'] = agg['Code'].map(code_order)
    agg = agg.sort_values(['_c', 'EEID']).reset_index(drop=True)   # matches May file: grouped by wage item, then EEID

    template_ext = os.path.splitext(TEMPLATE_F)[1].lower()
    if template_ext not in ('.xlsx', '.xlsm'):
        template_ext = '.xlsx'
    pay_path = os.path.join(
        OUTDIR,
        f'{MMYYYY}_STM SING Payroll_Additional (IT15)_Payment for Medical reimbursement{template_ext}'
    )
    shutil.copy(TEMPLATE_F, pay_path)
    wb = openpyxl.load_workbook(pay_path, keep_vba=(template_ext == '.xlsm'))
    if PAYROLL_TEMPLATE_SHEET not in wb.sheetnames:
        raise FlexInputError(
            f"Prior month IT15 payroll file: sheet '{PAYROLL_TEMPLATE_SHEET}' not found. "
            f"Sheets present: {', '.join(wb.sheetnames)}"
        )
    ws = wb[PAYROLL_TEMPLATE_SHEET]

    # month of request cell L3
    ws.cell(3, 12).value = PAY_MONTH

    START = 20
    # capture style templates from first data row
    style_cols = list(range(1, 13))
    styles = {c: ws.cell(START, c) for c in style_cols}

    def clone_style(src, dst):
        dst._style = copy(src._style)

    # Map the data area. The whole sheet is scanned rather than stopping at the first
    # blank cell (a gap would leave last month's rows in the file sent to payroll), but
    # only rows whose EEID cell is numeric count as data - the template ends with a
    # 'Grand Total' row carrying a SUM that must survive untouched.
    old_end = START - 1
    footer_row = None
    for r in range(START, ws.max_row + 1):
        value = ws.cell(r, 2).value
        if value is None:
            continue
        if str(value).strip().isdigit():
            old_end = r
        elif footer_row is None:
            footer_row = r

    if footer_row is not None and START + len(agg) - 1 >= footer_row:
        raise FlexInputError(
            f"IT15 template has room for {footer_row - START} payroll rows before the "
            f"'{ws.cell(footer_row, 2).value}' row at row {footer_row}, but this run produced "
            f"{len(agg)}. Extend the template's data area (and its total formula) and re-run."
        )

    # clear old data (cols A-L only; leave payroll-user cols M/P formulas intact)
    for r in range(START, old_end + 1):
        for c in range(1, 13):
            ws.cell(r, c).value = None

    # write new rows
    for i, row in agg.iterrows():
        r = START + i
        vals = {1: i + 1, 2: str(row['EEID']).zfill(6), 3: row['Name'], 4: row['Cost Centre'],
                5: str(row['Legal Entity Code']), 6: '1-Time', 7: '3 Earnings',
                8: row['Code'], 9: PAY_MONTH, 10: round(row['Amount'], 2), 11: 'SGD', 12: None}
        for c, v in vals.items():
            cell = ws.cell(r, c)
            cell.value = v
            clone_style(styles[c], cell)
        # Extend the payroll-user helper formulas (P/Q/R) when the new data runs past the
        # rows the template pre-fills. Existing cells are left alone - these columns belong
        # to payroll, not to this generator.
        for helper_col in HELPER_COLUMNS:
            src = ws.cell(START, helper_col)
            dst = ws.cell(r, helper_col)
            if src.value is None or dst.value is not None:
                continue
            dst.value = (
                Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
                if isinstance(src.value, str) and src.value.startswith('=')
                else src.value
            )
            clone_style(src, dst)

    wb.save(pay_path)

    # ---------------- SEA REPORTS ----------------
    sea_paths = []
    if len(sea_df):
        for country, grp in sea_df.groupby('SEA'):
            bd_c = build_breakdown(grp)
            p = os.path.join(OUTDIR, f'STM - {MON} SEA {country} Summary Reimbursment Report.xlsx')
            write_summary_report(bd_c, p)
            sea_paths.append(p)

    # ---------------- LOG ----------------
    echo("=== GENERATION LOG ===")
    echo(f"Pay month: {PAY_MONTH:%b %Y} | claims rows: {len(claims)} | employees: {claims['EEID6'].nunique()}")
    echo(f"IT15 template: {'saved server template' if USING_SAVED_TEMPLATE else 'uploaded file'} "
         f"({os.path.basename(TEMPLATE_F)}) | prior data rows cleared: {max(old_end - START + 1, 0)}")
    echo(f"Breakdown rows: {len(bd)} | total: {bd['Amount'].sum():,.2f}")
    echo(bd.groupby(['Legal Entity Code', 'Active/Inactive'])['Amount'].sum().to_string())
    echo(f"Payroll rows: {len(agg)} | total: {agg['Amount'].sum():,.2f}")
    echo(f"Excluded from payroll - already emailed to STM HR (final pay): "
         f"{excluded['EEID'].nunique()} leavers / {len(excluded)} claims / {excluded['Amount'].sum():,.2f}")
    echo(f"Blocked from payroll - leaver amount mismatch, manual check: "
         f"{blocked['EEID'].nunique()} employees / {blocked['Amount'].sum():,.2f}")
    echo(f"SEA claims rows: {len(sea_df)} -> reports: {[os.path.basename(p) for p in sea_paths] or 'none'}")
    echo(f"Mapping master refreshed: {len(new_master)} employees ({len(cur)} updated this run)")
    echo(f"Inactive employees flagged: {bd[bd['Active/Inactive'] == 'Inactive']['EEID'].nunique()}")

    # ---------------- VALIDATION REPORT ----------------
    # _row is internal bookkeeping for holding rows out; it never reaches the operator.
    errors_df = errors_df.drop(columns=['_row'], errors='ignore')
    warnings_df = warnings_df.drop(columns=['_row'], errors='ignore')
    n_err = len(errors_df)
    n_warn = len(warnings_df)
    vpath = None
    echo(f"VALIDATION: {n_err} error(s), {n_warn} warning(s)")
    if n_err or n_warn:
        vpath = os.path.join(OUTDIR, f'{MMYYYY}_Validation Exceptions Report.xlsx')
        with pd.ExcelWriter(vpath, engine='openpyxl') as xw:
            (errors_df if n_err else pd.DataFrame([{'Check': 'NONE', 'Detail': 'No errors'}])).to_excel(
                xw, sheet_name='ERRORS - excluded', index=False)
            (warnings_df if n_warn else pd.DataFrame([{'Check': 'NONE', 'Detail': 'No warnings'}])).to_excel(
                xw, sheet_name='WARNINGS - included', index=False)
            if len(held):
                held.to_excel(xw, sheet_name='Held claim rows', index=False)
        echo(f"Exception report written: {os.path.basename(vpath)}")
        if n_err and len(held):
            echo(f"!! {held['Reference No.'].nunique()} claim row(s) totalling {held['Payment Amt'].sum():,.2f} "
                 f"HELD OUT of both output files - resolve and re-run, or process manually")
        for _, r in pd.concat([errors_df.assign(Sev='ERROR'), warnings_df.assign(Sev='WARN')]).iterrows():
            echo(f"  [{r['Sev']}] {r['Check']}: EEID {r['EEID']} {r['Name'] or ''} - {r['Detail']}")
    else:
        echo("All checks passed - no exception report needed.")

    # ---------------- ACTIONABLE DISPOSITIONS ----------------
    # Attach, per exception, the SGD at stake and a plain include/exclude recommendation so
    # the frontend can tell the operator exactly what belongs in the payroll submission.
    # All amounts come from structured data, never parsed from the Detail strings.
    def _k6(v):
        try:
            return f'{int(v):06d}'
        except (TypeError, ValueError):
            return str(v)

    # All three lookups are keyed per EEID and hold the employee-level total, so an exception
    # amount is attributed once per (disposition, employee) below - never once per claim row.
    blocked_amt = {_k6(k): float(v) for k, v in bd[bd['_unrec']].groupby('EEID')['Amount'].sum().items()}
    held_amt = ({_k6(k): float(v) for k, v in held.groupby('EEID6')['Payment Amt'].sum().items()}
                if len(held) else {})
    # Emailed final-pay total per leaver, summed across mismatched categories. Compared against
    # the employee's blocked-claims total (blocked_amt) - both employee-level, so neither the
    # figure nor the recommendation double-counts when a leaver mismatches in >1 category.
    emailed_by_eeid = {}
    for (e6, _sub, _nm, _cat, tgt, _det) in leaver_mismatches:
        k = _k6(e6)
        emailed_by_eeid[k] = round(emailed_by_eeid.get(k, 0.0) + float(tgt), 2)

    def _disposition(check, eeid):
        e = _k6(eeid) if not pd.isna(eeid) else ''
        if check == 'LEAVER AMOUNT MISMATCH':
            emailed, here = emailed_by_eeid.get(e), blocked_amt.get(e)
            if emailed is not None and here is not None:
                if emailed >= here:
                    action = 'Exclude from payroll'
                    guidance = (f"Final-pay email {emailed:.2f} >= blocked claims {here:.2f} - the claim(s) are "
                                "most likely already settled via final pay. Keep them EXCLUDED from the payroll "
                                "file; confirm the claim references with HR.")
                else:
                    action = 'Split / include'
                    guidance = (f"Final-pay email {emailed:.2f} < blocked claims {here:.2f} - about {here - emailed:.2f} "
                                "is NOT covered by final pay. INCLUDE the uncovered portion in the payroll file; "
                                "agree the split with HR.")
            else:
                action, guidance = 'Decide', "Reconcile the leaver's final-pay email against the claims with HR."
            return {'disposition': 'blocked', 'action': action, 'amount': here, 'guidance': guidance}
        if check in HELD_ERROR_CHECKS:
            return {'disposition': 'held', 'action': 'Fix & re-run', 'amount': held_amt.get(e),
                    'guidance': HELD_GUIDANCE.get(check, 'Fix the source data and re-run.')}
        if check == 'LEAVER CATEGORY MISMATCH':
            return {'disposition': 'excluded', 'action': 'Review', 'amount': None,
                    'guidance': WARN_GUIDANCE[check]}
        return {'disposition': 'warn', 'action': 'Review', 'amount': None,
                'guidance': WARN_GUIDANCE.get(check, 'Review before submitting.')}

    validation = []
    if n_err or n_warn:
        allv = pd.concat([errors_df.assign(Sev='ERROR'), warnings_df.assign(Sev='WARNING')], ignore_index=True)
        amount_seen = set()   # (disposition, EEID): show each employee's total once so group sums stay authoritative
        for _, r in allv.iterrows():
            d = _disposition(r['Check'], r['EEID'])
            amount = d['amount']
            if amount is not None:
                akey = (d['disposition'], _k6(r['EEID']) if not pd.isna(r['EEID']) else '')
                if akey in amount_seen:
                    amount = None
                else:
                    amount_seen.add(akey)
            validation.append({
                'Sev': r['Sev'], 'Check': r['Check'],
                'EEID': '' if pd.isna(r['EEID']) else str(r['EEID']),
                'Name': '' if pd.isna(r.get('Name')) else str(r.get('Name')),
                'Detail': '' if pd.isna(r['Detail']) else str(r['Detail']),
                'amount': None if amount is None else round(float(amount), 2),
                'disposition': d['disposition'], 'action': d['action'], 'guidance': d['guidance'],
            })

    outputs = [sum_path, pay_path, master_path] + sea_paths
    if vpath:
        outputs.append(vpath)

    return {
        'outputs': outputs,
        'log': LOG,
        'errors': int(n_err),
        'warnings': int(n_warn),
        'validation': validation,
        'held_total': float(held['Payment Amt'].sum()) if n_err and len(held) else 0.0,
        'held_rows': int(len(held)) if n_err else 0,
        'grand_total': float(bd['Amount'].sum()),
        'payroll_total': float(pay['Amount'].sum()),
        'breakdown_rows': int(len(bd)),
        'payroll_rows': int(len(agg)),
        'employees': int(bd['EEID'].nunique()),
        'excluded_final_pay': float(excluded['Amount'].sum()),
        'blocked_mismatch': float(blocked['Amount'].sum()),
    }
