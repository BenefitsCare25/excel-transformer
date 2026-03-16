"""
Renewal Comparison Processor

Compares insurance renewal listing Excel files between two years and generates
an Adjustment Breakdown Excel report using Cancel and Re-enroll method.

Product types:
- Type 1 (Sum Insured): GTL, GDD, GPA, GDI — uses Sum Insured x Premium Rate
- Type 2 (Premium): GHS, GMM, GP, SP, GD — uses Annual Premium + GST

Abbreviations:
  GTL  = Group Term Life
  GDD  = Group Dread Disease
  GPA  = Group Personal Accident
  GDI  = Group Disability Income Benefit
  GHS  = Group Hospital & Surgical
  GMM  = Group Major Medical
  GP   = Group Clinical General Practitioner Insurance
  SP   = Group Clinical Specialist Insurance
  GD   = Group Dental Insurance
"""

import openpyxl
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import os
import re
import logging

logger = logging.getLogger(__name__)

EMPLOYEE_LISTING_SHEET_PREFIX = "Employee Listing"

SKIP_SECTIONS = {
    "employee", "dependant", "termination", "re-employment",
    "remarks", "internal", "re-enrolment"
}

# Known product type hints — used as fallback when column headers are ambiguous.
# Maps lowercase name patterns → product type (1 = Sum Insured, 2 = Premium)
# Product abbreviations: GTL, GDD, GPA, GDI, GHS, GMM, GP, SP, GD
PRODUCT_TYPE_HINTS: dict = {
    # Type 1 — Sum Insured based
    'gtl': 1, 'term life': 1,
    'gdd': 1, 'dread disease': 1,
    'gpa': 1, 'personal accident': 1,
    'gdi': 1, 'disability income': 1,
    # Type 2 — Premium based
    'ghs': 2, 'hospital': 2, 'surgical': 2,
    'gmm': 2, 'major medical': 2,
    'clinical': 2, 'general practitioner': 2,
    'sp': 2, 'specialist': 2,
    'gd': 2, 'dental': 2,
}


def _infer_product_type_from_name(name: str) -> int:
    """Return product type (1 or 2) from known name patterns, 0 if unrecognised."""
    name_lower = name.lower()
    for pattern, ptype in PRODUCT_TYPE_HINTS.items():
        if pattern in name_lower:
            return ptype
    return 0

ACCOUNTING_FORMAT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'


@dataclass
class DetectedProduct:
    name: str
    product_type: int  # 1 = Sum Insured, 2 = Premium
    col_start: int
    col_end: int
    admin_type_col: Optional[int] = None
    category_col: Optional[int] = None
    value_col: Optional[int] = None  # Sum Insured (type1) or Premium (type2)
    premium_rate: Optional[float] = None  # Type 1 only, from row 12
    annual_premium_col: Optional[int] = None  # Type 1: column with per-row calculated annual premium


@dataclass
class EmployeeRecord:
    name: str
    dob: str
    employee_id: str
    cost_centre: str
    department: str
    category: str = ''
    nric: str = ''
    email: str = ''
    product_data: Dict[str, dict] = field(default_factory=dict)

    def unique_key(self) -> str:
        """Matching priority: NRIC → email → name+DOB."""
        nric = self.nric.upper().strip()
        if nric:
            return f"NRIC|{nric}"
        email = self.email.lower().strip()
        if email:
            return f"EMAIL|{email}"
        name_norm = ' '.join(self.name.upper().split())
        return f"NAME|{name_norm}|{self.dob.strip()}"


@dataclass
class ClassificationChange:
    product: str
    name: str
    dob: str
    prev_type: str
    curr_type: str


@dataclass
class NamedExclusion:
    product: str
    name: str
    dob: str
    year: int


def _normalize(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _normalize_date(value) -> str:
    """Return date as dd/mm/yyyy string for consistent key matching across files."""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    s = str(value).strip()
    if not s:
        return ''
    # Try common string formats so both files resolve to the same key
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y',
                '%d/%m/%y', '%Y/%m/%d', '%d-%b-%Y', '%d %b %Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%d/%m/%Y')
        except ValueError:
            continue
    return s  # return as-is if no format matched (avoids silent empty)


def _to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).replace(',', '').replace('$', '').strip()
        if cleaned == '' or cleaned == '-':
            return None
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _year_from_val(val) -> Optional[int]:
    """Extract a plausible policy year (2000-2100) from a cell value."""
    if isinstance(val, datetime):
        if 2000 <= val.year <= 2100:
            return val.year
        return None
    if val:
        match = re.search(r'(20\d{2})', str(val))
        if match:
            return int(match.group(1))
    return None


def _find_employee_listing_sheet(wb, filename: str) -> Tuple[str, int]:
    """
    Locate the 'Employee Listing YYYY' sheet and return (sheet_name, year).
    Raises ValueError with a clear, actionable message if not found.
    """
    for name in wb.sheetnames:
        m = re.match(r'^Employee Listing\s+(\d{4})\s*$', name.strip(), re.IGNORECASE)
        if m:
            year = int(m.group(1))
            if 2000 <= year <= 2100:
                logger.info(f"Found sheet '{name}' → year {year} in {filename}")
                return name, year

    # Helpful fallback messages
    old_style = next((n for n in wb.sheetnames if n.strip().lower() == 'employee listing'), None)
    if old_style:
        raise ValueError(
            f"File '{filename}': sheet is named '{old_style}' but must include the policy year. "
            f"Please rename it to 'Employee Listing YYYY' — e.g. 'Employee Listing 2025'."
        )
    raise ValueError(
        f"File '{filename}': no 'Employee Listing YYYY' sheet found. "
        f"Found sheets: {', '.join(wb.sheetnames)}. "
        f"Please rename the sheet to 'Employee Listing YYYY' — e.g. 'Employee Listing 2025'."
    )


def _detect_header_rows(ws) -> Tuple[int, int, int]:
    """
    Dynamically locate (product_header_row, subheader_row, data_start_row).

    Strategy:
    1. Find the row whose merged cells contain the most product keywords
       (e.g. 'GTL', 'GHS', 'term life', 'hospital') → product_header_row.
    2. subheader_row  = product_header_row + 1
    3. data_start_row = product_header_row + 2

    Falls back to (13, 14, 15) when nothing is found.
    """
    PRODUCT_KEYWORDS = {
        'gtl', 'gdd', 'gpa', 'gdi', 'ghs', 'gmm', 'sp', 'gd',
        'term life', 'dread disease', 'personal accident', 'disability income',
        'hospital', 'surgical', 'major medical', 'dental', 'specialist', 'clinical',
        'group term', 'group dread', 'group hospital', 'group major',
    }
    row_scores: Dict[int, int] = {}
    for mr in ws.merged_cells.ranges:
        val = ws.cell(row=mr.min_row, column=mr.min_col).value
        if val:
            val_lower = str(val).lower()
            if any(kw in val_lower for kw in PRODUCT_KEYWORDS):
                row_scores[mr.min_row] = row_scores.get(mr.min_row, 0) + 1

    if not row_scores:
        logger.warning("Could not auto-detect product header row; defaulting to row 13/14/15")
        return 13, 14, 15

    product_header_row = max(row_scores, key=row_scores.get)
    logger.info(f"Auto-detected product header row: {product_header_row} (scores={row_scores})")
    return product_header_row, product_header_row + 1, product_header_row + 2


def _detect_year(ws, product_header_row: int = 13) -> Optional[int]:
    """
    Detect policy year by scanning metadata rows BEFORE the product/data area.
    Only rows 1..(product_header_row-1) are scanned so employee DOBs and hire
    dates (which live in the data rows) are never mistaken for the policy year.
    Falls back to a full-sheet scan if nothing found in the metadata rows.
    """
    # Phase 1: metadata rows only (safest)
    for row in range(1, product_header_row):
        for col in range(1, ws.max_column + 1):
            year = _year_from_val(ws.cell(row=row, column=col).value)
            if year:
                logger.info(f"Year {year} detected at R{row}C{col}")
                return year

    # Phase 2: scan the header rows themselves
    for row in range(product_header_row, product_header_row + 3):
        for col in range(1, ws.max_column + 1):
            year = _year_from_val(ws.cell(row=row, column=col).value)
            if year:
                logger.info(f"Year {year} detected (header row scan) at R{row}C{col}")
                return year

    return None


def _detect_products(ws, product_header_row: int = 13, subheader_row: int = 14) -> List[DetectedProduct]:
    """
    Detect products dynamically using merged cells on product_header_row and
    column sub-headers on subheader_row.

    Key behaviours:
    - Each section's col_end is extended to just before the next section's
      col_start so that 'orphan' premium columns outside a merged range are
      captured (common when only the label columns are merged).
    - admin_type_col is assigned as the nearest 'Type of Administration'
      column to the LEFT of each section.
    - Premium rate (Type-1) is read from the row immediately above the
      product_header_row.
    """
    products = []
    merged_sections = []
    rate_row = product_header_row - 1  # row that holds premium rates for Type-1 products

    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row == product_header_row and merged_range.max_row == product_header_row:
            cell_val = ws.cell(row=product_header_row, column=merged_range.min_col).value
            if cell_val:
                merged_sections.append({
                    'name': str(cell_val).strip(),
                    'col_start': merged_range.min_col,
                    'col_end': merged_range.max_col
                })

    if not merged_sections:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=product_header_row, column=col).value
            if val:
                name = str(val).strip()
                is_merged = any(
                    mr.min_row <= product_header_row <= mr.max_row and
                    mr.min_col <= col <= mr.max_col
                    for mr in ws.merged_cells.ranges
                )
                if not is_merged:
                    merged_sections.append({'name': name, 'col_start': col, 'col_end': col})

    merged_sections.sort(key=lambda x: x['col_start'])

    # Extend each section's col_end to just before the next section starts so
    # orphan value columns (Premium, Sum Insured) outside the merge are included
    for i, section in enumerate(merged_sections):
        if i + 1 < len(merged_sections):
            next_start = merged_sections[i + 1]['col_start']
            if next_start - 1 > section['col_end']:
                section['col_end'] = next_start - 1
        else:
            if ws.max_column > section['col_end']:
                section['col_end'] = ws.max_column

    # Pre-detect all "Type of Administration" columns across the subheader row
    admin_type_cols = []
    for col in range(1, ws.max_column + 1):
        header = _normalize(ws.cell(row=subheader_row, column=col).value).lower()
        if 'type of administration' in header:
            admin_type_cols.append(col)

    for section in merged_sections:
        name_lower = section['name'].lower()
        if any(skip in name_lower for skip in SKIP_SECTIONS):
            continue

        product = DetectedProduct(
            name=section['name'],
            product_type=0,
            col_start=section['col_start'],
            col_end=section['col_end']
        )

        # Assign nearest admin_type col to the LEFT of this section
        candidates = [c for c in admin_type_cols if c < section['col_start']]
        if candidates:
            product.admin_type_col = max(candidates)

        has_sum_insured = False
        eligible_si_col = None

        for col in range(section['col_start'], section['col_end'] + 1):
            header = _normalize(ws.cell(row=subheader_row, column=col).value).lower()
            if not header:
                continue

            if 'category' in header:
                product.category_col = col
            elif 'eligible sum insured' in header:
                eligible_si_col = col
                has_sum_insured = True
            elif 'sum insured' in header and 'pending' not in header:
                if not eligible_si_col:
                    product.value_col = col
                has_sum_insured = True
            elif 'premium' in header and 'gst' not in header and 'w/' not in header:
                if not has_sum_insured:
                    product.value_col = col
                else:
                    product.annual_premium_col = col  # Type 1: per-row calculated premium

        if eligible_si_col:
            product.value_col = eligible_si_col

        if has_sum_insured:
            product.product_type = 1
            if rate_row >= 1:
                for col in range(section['col_start'], section['col_end'] + 1):
                    rate_val = _to_float(ws.cell(row=rate_row, column=col).value)
                    if rate_val is not None and 0 < rate_val < 1:
                        product.premium_rate = rate_val
                        break
        elif product.value_col:
            product.product_type = 2

        # Fallback: infer type from product name when column headers are ambiguous
        if product.product_type == 0:
            hinted = _infer_product_type_from_name(product.name)
            if hinted > 0:
                product.product_type = hinted
                logger.info(f"Product type for '{product.name}' inferred from name hint: Type {hinted}")
                if not product.value_col:
                    skip_kw = {'category', 'name', 'type', 'administration', 'nric', 'id', 'dob', 'date',
                               'sex', 'age', 'marital', 'nationality', 'designation', 'employment',
                               'mu status', 'mu decision', 'pending', 'acceptance', 'last accepted'}
                    for col in range(section['col_start'], section['col_end'] + 1):
                        header = _normalize(ws.cell(row=subheader_row, column=col).value).lower()
                        if header and not any(kw in header for kw in skip_kw):
                            product.value_col = col
                            logger.info(f"Value col for '{product.name}' set by broad scan: col {col} ('{header}')")
                            break

        if product.product_type > 0 and product.value_col:
            products.append(product)
            logger.info(
                f"Detected product: {product.name} (Type {product.product_type}), "
                f"cols {product.col_start}-{product.col_end}, "
                f"value_col={product.value_col}, admin_col={product.admin_type_col}"
                + (f", rate={product.premium_rate}" if product.premium_rate else "")
            )
        else:
            logger.warning(
                f"Skipping section '{section['name']}': "
                f"type={product.product_type}, value_col={product.value_col}"
            )

    return products



def _find_employee_columns(ws, subheader_row: int = 14) -> dict:
    """Find common employee data columns by scanning the subheader row."""
    col_map = {}
    name_fallback = None

    # Scan all columns — employee headers can be anywhere in the subheader row
    for col in range(1, ws.max_column + 1):
        header = _normalize(ws.cell(row=subheader_row, column=col).value).lower()
        if not header:
            continue

        if 'name' not in col_map and 'name' in header and ('surname' in header or 'first' in header):
            col_map['name'] = col
        elif name_fallback is None and 'name' in header and 'employee' not in header and 'dependant' not in header:
            name_fallback = col

        if 'dob' not in col_map and ('date of birth' in header or 'dob' in header or header in ('d.o.b', 'birth date', 'birthdate')):
            col_map['dob'] = col
        elif 'employee id' in header or 'emp id' in header or 'staff id' in header or 'employee id no' in header:
            col_map['employee_id'] = col  # check BEFORE nric to prevent 'id no' false match
        elif 'nric' not in col_map and ('nric' in header or 'ic no' in header or 'id no' in header or 'national id' in header or 'passport' in header or header in ('nric/fin', 'nric / fin', 'fin', 'nric/passport')):
            col_map['nric'] = col
        elif 'email' not in col_map and ('email' in header or 'e-mail' in header or 'e mail' in header):
            col_map['email'] = col
        elif 'cost centre' in header or 'cost center' in header:
            col_map['cost_centre'] = col
        elif 'department' in header:
            col_map['department'] = col
        elif 'category' in header and 'category' not in col_map:
            col_map['category'] = col

    if 'name' not in col_map and name_fallback:
        col_map['name'] = name_fallback

    logger.info(f"Employee column map (subheader_row={subheader_row}): {col_map}")
    return col_map



def _extract_employees(ws, products: List[DetectedProduct], emp_cols: dict,
                       data_start_row: int = 15) -> Dict[str, EmployeeRecord]:
    """Extract employee records with per-product data."""
    employees = {}

    for row in range(data_start_row, ws.max_row + 1):
        name_val = _normalize(ws.cell(row=row, column=emp_cols.get('name', 2)).value)
        if not name_val:
            continue

        dob_val = _normalize_date(ws.cell(row=row, column=emp_cols.get('dob', 8)).value)

        emp = EmployeeRecord(
            name=name_val,
            dob=dob_val,
            employee_id=_normalize(ws.cell(row=row, column=emp_cols['employee_id']).value) if 'employee_id' in emp_cols else '',
            cost_centre=_normalize(ws.cell(row=row, column=emp_cols['cost_centre']).value) if 'cost_centre' in emp_cols else '',
            department=_normalize(ws.cell(row=row, column=emp_cols['department']).value) if 'department' in emp_cols else '',
            category=_normalize(ws.cell(row=row, column=emp_cols.get('category', 12)).value),
            nric=_normalize(ws.cell(row=row, column=emp_cols['nric']).value) if 'nric' in emp_cols else '',
            email=_normalize(ws.cell(row=row, column=emp_cols['email']).value) if 'email' in emp_cols else '',
        )

        for product in products:
            admin_type = ''
            if product.admin_type_col:
                admin_type = _normalize(ws.cell(row=row, column=product.admin_type_col).value)

            category = ''
            if product.category_col:
                category = _normalize(ws.cell(row=row, column=product.category_col).value)

            value = _to_float(ws.cell(row=row, column=product.value_col).value)
            annual_premium = None
            if product.annual_premium_col:
                annual_premium = _to_float(ws.cell(row=row, column=product.annual_premium_col).value)

            if admin_type or value is not None:
                emp.product_data[product.name] = {
                    'admin_type': admin_type,
                    'category': category,
                    'value': value,
                    'annual_premium': annual_premium,
                }

        key = emp.unique_key()
        if len(employees) < 3:
            logger.info(f"Sample key: {repr(key)}")
        if key not in employees:
            employees[key] = emp
        else:
            # Same name+DOB appears twice — keep max value per product
            for pname, pdata in emp.product_data.items():
                existing = employees[key].product_data
                if pname not in existing:
                    existing[pname] = pdata
                elif pdata.get('value') is not None:
                    ev = existing[pname].get('value')
                    if ev is None or pdata['value'] > ev:
                        existing[pname] = pdata

    # Post-process: for family rows (no DOB), add type-2 (premium) product entries
    # from the EO row if not already present. This handles products like Group Clinical
    # where the family row has zero/None premium but still appears in the output.
    # Type-1 (sum insured) products are excluded — only the employee is insured.
    type2_names = {p.name for p in products if p.product_type == 2}
    name_to_eo: Dict[str, EmployeeRecord] = {}
    for emp in employees.values():
        if emp.dob:
            name_norm = emp.name.upper().strip()
            if name_norm not in name_to_eo:
                name_to_eo[name_norm] = emp

    for emp in employees.values():
        if emp.dob:
            continue
        eo_emp = name_to_eo.get(emp.name.upper().strip())
        if not eo_emp:
            continue
        for pname in type2_names:
            if pname in eo_emp.product_data and pname not in emp.product_data:
                eo_pdata = eo_emp.product_data[pname]
                if eo_pdata.get('admin_type', '').lower() != 'named':
                    emp.product_data[pname] = {
                        'admin_type': '',
                        'category': eo_pdata.get('category', ''),
                        'value': None,
                        'annual_premium': None,
                    }

    return employees



def _generate_product_sheet(
    wb: openpyxl.Workbook,
    product: DetectedProduct,
    prev_employees: Dict[str, EmployeeRecord],
    curr_employees: Dict[str, EmployeeRecord],
    prev_year: int,
    curr_year: int,
    divisor: int,
    prev_rate: Optional[float] = None,
) -> Tuple[int, int]:
    """Generate a product adjustment sheet. Returns (prev_rows, curr_rows) written."""
    sheet_name = re.sub(r'[\\/*?:\[\]&]', '', product.name)[:31]
    ws = wb.create_sheet(title=sheet_name)

    is_type1 = product.product_type == 1
    # Use prev year rate for all rows to match reference convention
    rate = (prev_rate or product.premium_rate) if is_type1 else None

    # Short product abbreviation for column headers
    name = product.name
    col_h_label = f"{name} {prev_year}"
    col_i_label = f"{name} {curr_year}"

    if is_type1:
        col_j_label = f"{name}\n(sum insured)"
        col_k_label = f"{name} Annual Premium"
        col_l_label = "Adj Premium\n(w/o GST)"
        col_m_label = "Adj Premium\n(w/ GST)"
    else:
        col_j_label = f"{name} Annual Premium"
        col_k_label = "GST 9%"
        col_l_label = "Adj Premium / 2"

    base_headers = ["Remarks", "Name \n(Surname, First Name) ", "NRIC No. or FIN No.\n(eg. S1234567F)",
                    "EMP ID", "Cost Centre", "Department",
                    "Category                        \n(Pls Select Basis Accordingly)",
                    col_h_label, col_i_label, col_j_label, col_k_label, col_l_label]
    headers = base_headers + [col_m_label] if is_type1 else base_headers

    header_font = Font(bold=True, size=11)
    header_align = Alignment(wrap_text=True, vertical='center')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_align

    row_num = 2
    data_start_row = 2

    prev_headcount = []
    for key, emp in prev_employees.items():
        pdata = emp.product_data.get(product.name)
        if not pdata:
            continue
        admin = pdata.get('admin_type', '').strip().lower()
        if admin == 'named':
            continue
        # Skip None-value rows unless it's a family row (no DOB) — family rows
        # appear in all product sheets even with zero premium
        is_family_row = not emp.dob
        if pdata.get('value') is None and not is_family_row:
            continue
        prev_headcount.append((key, emp, pdata))

    prev_headcount.sort(key=lambda x: (x[1].department.lower(), x[1].name.upper()))

    for key, emp, pdata in prev_headcount:
        ws.cell(row=row_num, column=1, value=f"Renewal {prev_year}")
        ws.cell(row=row_num, column=2, value=emp.name)
        ws.cell(row=row_num, column=3, value=emp.nric)
        ws.cell(row=row_num, column=4, value=emp.employee_id)
        ws.cell(row=row_num, column=5, value=emp.cost_centre)
        ws.cell(row=row_num, column=6, value=emp.department)
        ws.cell(row=row_num, column=7, value=pdata.get('category') or emp.category)

        val = pdata.get('value')
        if val is not None:
            ws.cell(row=row_num, column=8, value=val)
        # Col I empty for prev year; Col J = I - H
        ws.cell(row=row_num, column=10).value = f"=I{row_num}-H{row_num}"

        if is_type1:
            ap = pdata.get('annual_premium')
            if ap is not None:
                ws.cell(row=row_num, column=11, value=-ap)  # cancel = negative
            elif rate:
                ws.cell(row=row_num, column=11).value = f"=J{row_num}*{rate}"
            ws.cell(row=row_num, column=12).value = f"=K{row_num}/{divisor}"  # Adj w/o GST
            ws.cell(row=row_num, column=13).value = f"=L{row_num}*1.09"        # Adj w/ GST
        else:
            ws.cell(row=row_num, column=11).value = f"=J{row_num}*0.09"
            ws.cell(row=row_num, column=12).value = f"=J{row_num}/{divisor}"

        fmt_end = 14 if is_type1 else 13
        for c in range(8, fmt_end):
            ws.cell(row=row_num, column=c).number_format = ACCOUNTING_FORMAT

        row_num += 1

    curr_headcount = []
    for key, emp in curr_employees.items():
        pdata = emp.product_data.get(product.name)
        if not pdata:
            continue
        admin = pdata.get('admin_type', '').strip().lower()
        if admin == 'named':
            # Only exclude if Named in both years (permanent Named employee).
            # If they were Headcount in prev year (classification change), include.
            prev_emp = prev_employees.get(key)
            if prev_emp is None:
                continue  # new Named employee → exclude
            prev_admin = prev_emp.product_data.get(product.name, {}).get('admin_type', '').strip().lower()
            if prev_admin == 'named':
                continue  # both years Named → exclude
            # else: switched from HC to Named → include in both blocks
        is_family_row = not emp.dob
        if pdata.get('value') is None and not is_family_row:
            continue
        curr_headcount.append((key, emp, pdata))

    curr_headcount.sort(key=lambda x: (x[1].department.lower(), x[1].name.upper()))

    for key, emp, pdata in curr_headcount:
        ws.cell(row=row_num, column=1, value=f"Renewal {curr_year}")
        ws.cell(row=row_num, column=2, value=emp.name)
        ws.cell(row=row_num, column=3, value=emp.nric)
        ws.cell(row=row_num, column=4, value=emp.employee_id)
        ws.cell(row=row_num, column=5, value=emp.cost_centre)
        ws.cell(row=row_num, column=6, value=emp.department)
        ws.cell(row=row_num, column=7, value=pdata.get('category') or emp.category)

        # Col H empty for curr year; Col J = I - H
        # Type 1 (sum insured): always use curr year value so SI changes are captured.
        # Type 2 (premium): use prev year value for renewals so rate-only changes are
        #   excluded from the adjustment (billed separately at renewal).
        if is_type1:
            val = pdata.get('value')
        elif key in prev_employees:
            prev_pdata = prev_employees[key].product_data.get(product.name, {})
            val = prev_pdata.get('value')
        else:
            val = pdata.get('value')
        if val is not None:
            ws.cell(row=row_num, column=9, value=val)
        ws.cell(row=row_num, column=10).value = f"=I{row_num}-H{row_num}"

        if is_type1:
            # Always use curr year annual premium so SI changes flow through correctly.
            # For unchanged SI: curr_premium == prev_premium → cancel+re-enroll nets to 0.
            # For changed SI: curr_premium reflects new SI → adjustment is captured.
            ap = pdata.get('annual_premium')
            if ap is not None:
                ws.cell(row=row_num, column=11, value=ap)  # enroll = positive
            elif rate:
                ws.cell(row=row_num, column=11).value = f"=J{row_num}*{rate}"
            ws.cell(row=row_num, column=12).value = f"=K{row_num}/{divisor}"  # Adj w/o GST
            ws.cell(row=row_num, column=13).value = f"=L{row_num}*1.09"        # Adj w/ GST
        else:
            ws.cell(row=row_num, column=11).value = f"=J{row_num}*0.09"
            ws.cell(row=row_num, column=12).value = f"=J{row_num}/{divisor}"

        fmt_end = 14 if is_type1 else 13
        for c in range(8, fmt_end):
            ws.cell(row=row_num, column=c).number_format = ACCOUNTING_FORMAT

        row_num += 1

    last_data_row = row_num - 1

    if last_data_row < data_start_row:
        return len(prev_headcount), len(curr_headcount)

    row_num += 1

    bold_font = Font(bold=True)
    if is_type1:
        ws.cell(row=row_num, column=11, value="Adjustment Breakdown").font = bold_font
        ws.cell(row=row_num, column=12).value = f"=SUM(L{data_start_row}:L{last_data_row})"
        ws.cell(row=row_num, column=12).number_format = ACCOUNTING_FORMAT
        ws.cell(row=row_num, column=12).font = bold_font
        ws.cell(row=row_num, column=13).value = f"=SUM(M{data_start_row}:M{last_data_row})"
        ws.cell(row=row_num, column=13).number_format = ACCOUNTING_FORMAT
        ws.cell(row=row_num, column=13).font = bold_font
    else:
        ws.cell(row=row_num, column=11, value="Adjustment Premium").font = bold_font
        ws.cell(row=row_num, column=12).value = f"=SUM(L{data_start_row}:L{last_data_row})"
        ws.cell(row=row_num, column=12).number_format = ACCOUNTING_FORMAT
        ws.cell(row=row_num, column=12).font = bold_font

        row_num += 1
        ws.cell(row=row_num, column=11, value="GST").font = bold_font
        ws.cell(row=row_num, column=12).value = f"=L{row_num-1}*0.09"
        ws.cell(row=row_num, column=12).number_format = ACCOUNTING_FORMAT

        row_num += 1
        ws.cell(row=row_num, column=11, value="Adjustment Premium with GST").font = bold_font
        ws.cell(row=row_num, column=12).value = f"=L{row_num-2}+L{row_num-1}"
        ws.cell(row=row_num, column=12).number_format = ACCOUNTING_FORMAT
        ws.cell(row=row_num, column=12).font = bold_font

    col_widths = {'A': 18, 'B': 30, 'C': 18, 'D': 12, 'E': 14, 'F': 20,
                  'G': 14, 'H': 18, 'I': 18, 'J': 18, 'K': 18, 'L': 18}
    if is_type1:
        col_widths['M'] = 18
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    return len(prev_headcount), len(curr_headcount)


def _generate_summary_sheet(
    wb: openpyxl.Workbook,
    products: List[DetectedProduct],
    prev_employees: Dict[str, EmployeeRecord],
    curr_employees: Dict[str, EmployeeRecord],
    prev_year: int,
    curr_year: int,
    prev_filename: str,
    curr_filename: str,
    classification_changes: List[ClassificationChange],
    named_exclusions: List[NamedExclusion],
):
    """Generate the Summary sheet as first sheet."""
    ws = wb.active
    ws.title = "Summary"

    bold_font = Font(bold=True, size=11)
    header_font = Font(bold=True, size=12)
    row = 1

    # File Info
    ws.cell(row=row, column=1, value="File Information").font = header_font
    row += 1
    ws.cell(row=row, column=1, value="Previous Year:")
    ws.cell(row=row, column=2, value=f"{prev_year} ({prev_filename})")
    row += 1
    ws.cell(row=row, column=1, value="Current Year:")
    ws.cell(row=row, column=2, value=f"{curr_year} ({curr_filename})")
    row += 2

    # Products Detected
    ws.cell(row=row, column=1, value="Products Detected").font = header_font
    row += 1
    prod_headers = ["Product", "Type", "Prev HC", "Curr HC", "Prev Named", "Curr Named"]
    for ci, h in enumerate(prod_headers, 1):
        ws.cell(row=row, column=ci, value=h).font = bold_font
    row += 1

    for product in products:
        ptype = "Sum Insured" if product.product_type == 1 else "Premium"
        # Headcount = all with the product who are not explicitly 'named'
        prev_hc = sum(1 for e in prev_employees.values()
                      if product.name in e.product_data
                      and e.product_data[product.name].get('admin_type', '').strip().lower() != 'named')
        curr_hc = sum(1 for e in curr_employees.values()
                      if product.name in e.product_data
                      and e.product_data[product.name].get('admin_type', '').strip().lower() != 'named')
        prev_named = sum(1 for e in prev_employees.values()
                         if product.name in e.product_data
                         and e.product_data[product.name].get('admin_type', '').strip().lower() == 'named')
        curr_named = sum(1 for e in curr_employees.values()
                         if product.name in e.product_data
                         and e.product_data[product.name].get('admin_type', '').strip().lower() == 'named')

        ws.cell(row=row, column=1, value=product.name)
        ws.cell(row=row, column=2, value=ptype)
        ws.cell(row=row, column=3, value=prev_hc)
        ws.cell(row=row, column=4, value=curr_hc)
        ws.cell(row=row, column=5, value=prev_named)
        ws.cell(row=row, column=6, value=curr_named)
        row += 1

    row += 1

    # Employee Overview
    ws.cell(row=row, column=1, value="Employee Overview").font = header_font
    row += 1
    prev_keys = set(prev_employees.keys())
    curr_keys = set(curr_employees.keys())
    common = prev_keys & curr_keys
    added = curr_keys - prev_keys
    removed = prev_keys - curr_keys

    for label, val, note in [
        ("Previous year employees:",
         len(prev_keys),
         f"Total headcount in {prev_year} policy file"),
        ("Current year employees:",
         len(curr_keys),
         f"Total headcount in {curr_year} policy file"),
        ("Common (matched):",
         len(common),
         "Employees found in BOTH years — matched by name & date of birth"),
        ("New employees:",
         len(added),
         f"In {curr_year} but NOT in {prev_year} — enrolled after previous policy (Cancel & Re-enroll as NEW)"),
        ("Left employees:",
         len(removed),
         f"In {prev_year} but NOT in {curr_year} — no longer on current policy (Cancel & Re-enroll as CANCEL)"),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=3, value=note).font = Font(italic=True, color="595959", size=9)
        row += 1

    row += 1

    # Classification Changes
    if classification_changes:
        ws.cell(row=row, column=1, value="Classification Changes").font = header_font
        row += 1
        change_headers = ["Product", "Name", "DOB", "Previous Type", "Current Type"]
        for ci, h in enumerate(change_headers, 1):
            ws.cell(row=row, column=ci, value=h).font = bold_font
        row += 1
        for cc in classification_changes:
            ws.cell(row=row, column=1, value=cc.product)
            ws.cell(row=row, column=2, value=cc.name)
            ws.cell(row=row, column=3, value=cc.dob)
            ws.cell(row=row, column=4, value=cc.prev_type)
            ws.cell(row=row, column=5, value=cc.curr_type)
            row += 1
        row += 1

    # Named Employees Excluded
    if named_exclusions:
        ws.cell(row=row, column=1, value="Named Employees (Excluded)").font = header_font
        row += 1
        ne_headers = ["Product", "Name", "DOB", "Year"]
        for ci, h in enumerate(ne_headers, 1):
            ws.cell(row=row, column=ci, value=h).font = bold_font
        row += 1
        for ne in named_exclusions:
            ws.cell(row=row, column=1, value=ne.product)
            ws.cell(row=row, column=2, value=ne.name)
            ws.cell(row=row, column=3, value=ne.dob)
            ws.cell(row=row, column=4, value=ne.year)
            row += 1

    col_widths = {'A': 28, 'B': 30, 'C': 16, 'D': 16, 'E': 16, 'F': 16}
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width


def process_renewal_comparison(
    file1_path: str,
    file2_path: str,
    file1_name: str,
    file2_name: str,
    pro_rata_divisor: int = 2,
    output_path: str = None,
) -> dict:
    """
    Main entry point: compare two renewal listing files and generate output.

    Returns dict with summary info for the API response.
    """
    wb1 = openpyxl.load_workbook(file1_path, data_only=True)
    wb2 = openpyxl.load_workbook(file2_path, data_only=True)

    try:
        sheet1_name, year1 = _find_employee_listing_sheet(wb1, file1_name)
    except ValueError:
        wb1.close()
        wb2.close()
        raise
    try:
        sheet2_name, year2 = _find_employee_listing_sheet(wb2, file2_name)
    except ValueError:
        wb1.close()
        wb2.close()
        raise

    ws1 = wb1[sheet1_name]
    ws2 = wb2[sheet2_name]

    # Auto-detect row structure for each file independently (formats vary by client)
    ph1, sh1, ds1 = _detect_header_rows(ws1)
    ph2, sh2, ds2 = _detect_header_rows(ws2)
    logger.info(f"File1 rows — product_header:{ph1} subheader:{sh1} data_start:{ds1}")
    logger.info(f"File2 rows — product_header:{ph2} subheader:{sh2} data_start:{ds2}")

    if year1 == year2:
        wb1.close()
        wb2.close()
        raise ValueError(f"Both files appear to be from the same year ({year1}).")

    if year1 < year2:
        prev_ws, curr_ws = ws1, ws2
        prev_year, curr_year = year1, year2
        prev_filename, curr_filename = file1_name, file2_name
        prev_ph, prev_sh, prev_ds = ph1, sh1, ds1
        curr_ph, curr_sh, curr_ds = ph2, sh2, ds2
    else:
        prev_ws, curr_ws = ws2, ws1
        prev_year, curr_year = year2, year1
        prev_filename, curr_filename = file2_name, file1_name
        prev_ph, prev_sh, prev_ds = ph2, sh2, ds2
        curr_ph, curr_sh, curr_ds = ph1, sh1, ds1

    logger.info(f"Previous year: {prev_year} ({prev_filename})")
    logger.info(f"Current year: {curr_year} ({curr_filename})")

    # Detect products from both files using their respective row structures
    prev_products = _detect_products(prev_ws, prev_ph, prev_sh)
    curr_products = _detect_products(curr_ws, curr_ph, curr_sh)

    if not prev_products and not curr_products:
        wb1.close()
        wb2.close()
        raise ValueError(
            "No products detected in either file. "
            "Ensure product names (e.g. GTL, GHS) appear as merged cell headers above the column headers."
        )

    # Use current year products as primary for structure; track prev rates separately
    products_map = {}
    prev_rates_map: Dict[str, Optional[float]] = {}
    for p in prev_products:
        products_map[p.name] = p
        prev_rates_map[p.name] = p.premium_rate
    for p in curr_products:
        products_map[p.name] = p
        if p.name not in prev_rates_map:
            prev_rates_map[p.name] = p.premium_rate

    all_product_names = list(products_map.keys())
    logger.info(f"Products detected: {all_product_names}")

    # Find employee columns using the detected subheader row for each file
    prev_emp_cols = _find_employee_columns(prev_ws, prev_sh)
    curr_emp_cols = _find_employee_columns(curr_ws, curr_sh)

    # Extract employees using the detected data start row for each file
    prev_employees = _extract_employees(prev_ws, prev_products, prev_emp_cols, prev_ds)
    curr_employees = _extract_employees(curr_ws, curr_products, curr_emp_cols, curr_ds)

    logger.info(f"Previous year employees: {len(prev_employees)}")
    logger.info(f"Current year employees: {len(curr_employees)}")

    wb1.close()
    wb2.close()

    # Detect classification changes and named exclusions
    classification_changes = []
    named_exclusions = []
    prev_keys = set(prev_employees.keys())
    curr_keys = set(curr_employees.keys())
    common_keys = prev_keys & curr_keys

    for product_name in all_product_names:
        for key in common_keys:
            prev_emp = prev_employees[key]
            curr_emp = curr_employees[key]

            prev_pdata = prev_emp.product_data.get(product_name, {})
            curr_pdata = curr_emp.product_data.get(product_name, {})

            prev_admin = prev_pdata.get('admin_type', '').strip().lower()
            curr_admin = curr_pdata.get('admin_type', '').strip().lower()

            if prev_admin and curr_admin and prev_admin != curr_admin:
                classification_changes.append(ClassificationChange(
                    product=product_name,
                    name=prev_emp.name,
                    dob=prev_emp.dob,
                    prev_type=prev_pdata.get('admin_type', ''),
                    curr_type=curr_pdata.get('admin_type', ''),
                ))

        # Track named exclusions
        for key in prev_keys:
            emp = prev_employees[key]
            pdata = emp.product_data.get(product_name, {})
            admin = pdata.get('admin_type', '').strip().lower()
            if admin == 'named':
                named_exclusions.append(NamedExclusion(
                    product=product_name, name=emp.name, dob=emp.dob, year=prev_year
                ))
        for key in curr_keys:
            emp = curr_employees[key]
            pdata = emp.product_data.get(product_name, {})
            admin = pdata.get('admin_type', '').strip().lower()
            if admin == 'named':
                named_exclusions.append(NamedExclusion(
                    product=product_name, name=emp.name, dob=emp.dob, year=curr_year
                ))

    # Generate output Excel
    out_wb = openpyxl.Workbook()

    _generate_summary_sheet(
        out_wb, list(products_map.values()),
        prev_employees, curr_employees,
        prev_year, curr_year,
        prev_filename, curr_filename,
        classification_changes, named_exclusions,
    )

    sheet_row_counts = {}
    for product_name in all_product_names:
        product = products_map[product_name]
        result = _generate_product_sheet(
            out_wb, product,
            prev_employees, curr_employees,
            prev_year, curr_year,
            pro_rata_divisor,
            prev_rate=prev_rates_map.get(product_name),
        )
        sheet_row_counts[product_name] = result if result else (0, 0)

    out_wb.save(output_path)
    out_wb.close()

    logger.info(f"Output saved to: {output_path}")

    # Build response summary
    product_summaries = []
    for product_name in all_product_names:
        product = products_map[product_name]
        prev_hc = sum(1 for e in prev_employees.values()
                      if product_name in e.product_data
                      and e.product_data[product_name].get('admin_type', '').strip().lower() != 'named')
        curr_hc = sum(1 for e in curr_employees.values()
                      if product_name in e.product_data
                      and e.product_data[product_name].get('admin_type', '').strip().lower() != 'named')
        prev_named = sum(1 for e in prev_employees.values()
                         if product_name in e.product_data
                         and e.product_data[product_name].get('admin_type', '').strip().lower() == 'named')
        curr_named = sum(1 for e in curr_employees.values()
                         if product_name in e.product_data
                         and e.product_data[product_name].get('admin_type', '').strip().lower() == 'named')

        sheet_prev, sheet_curr = sheet_row_counts.get(product_name, (0, 0))
        product_summaries.append({
            'name': product_name,
            'type': 'Sum Insured' if product.product_type == 1 else 'Premium',
            'prev_headcount': prev_hc,
            'curr_headcount': curr_hc,
            'prev_named': prev_named,
            'curr_named': curr_named,
            'sheet_prev_rows': sheet_prev,
            'sheet_curr_rows': sheet_curr,
        })

    return {
        'previous_year': prev_year,
        'current_year': curr_year,
        'previous_filename': prev_filename,
        'current_filename': curr_filename,
        'products': product_summaries,
        'employee_summary': {
            'prev_total': len(prev_keys),
            'curr_total': len(curr_keys),
            'common': len(common_keys),
            'added': len(curr_keys - prev_keys),
            'removed': len(prev_keys - curr_keys),
        },
        'classification_changes': [
            {'product': cc.product, 'name': cc.name, 'dob': cc.dob,
             'prev_type': cc.prev_type, 'curr_type': cc.curr_type}
            for cc in classification_changes
        ],
        'named_excluded': [
            {'product': ne.product, 'name': ne.name, 'dob': ne.dob, 'year': ne.year}
            for ne in named_exclusions
        ],
    }
