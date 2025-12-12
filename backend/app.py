from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import os
import sys
import uuid
import re
from datetime import datetime
import traceback
import googlemaps
from geopy.geocoders import GoogleV3
from dotenv import load_dotenv
from functools import lru_cache
import logging
import threading
import time
try:
    from concurrent.futures import ThreadPoolExecutor, as_completed
    CONCURRENT_SUPPORT = True
except ImportError:
    # Fallback for environments without concurrent.futures
    CONCURRENT_SUPPORT = False
    import threading

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

load_dotenv()

app = Flask(__name__)
CORS(app)

# Configuration
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', 'uploads')
PROCESSED_FOLDER = os.getenv('PROCESSED_FOLDER', 'processed')

# Postal code master file paths (in order of preference)
POSTAL_CODE_PATHS = [
    os.getenv('POSTAL_CODE_MASTER_FILE'),  # Environment variable (highest priority)
    os.path.join('..', 'data', 'SG_postal.csv'),  # Full SG postal codes CSV (121K+ codes)
    os.path.join('..', 'data', 'postal_code_master.xlsx'),  # Project data folder (fallback)
    os.getenv('POSTAL_CODE_FALLBACK_PATH'),  # Configurable fallback path
    'postal_code_master.xlsx',  # Current directory
]

# Government hospitals to exclude from clinic matching (with common abbreviations)
GOVERNMENT_HOSPITALS = {
    'alexandra hospital',
    'changi general hospital', 'cgh',
    'institute of mental health', 'imh',
    'khoo teck puat hospital', 'ktph',
    'kk women\'s and children\'s hospital', 'kkh',
    'national university hospital', 'nuh',
    'ng teng fong general hospital',
    'jurong community hospital',  # Part of combined entry
    'sengkang general hospital', 'skh',
    'singapore general hospital', 'sgh',
    'tan tock seng hospital', 'ttsh',
    'woodlands health',
}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

# Global postal code lookup - loaded once at module startup
_POSTAL_CODE_LOOKUP_CACHE = None

def _load_postal_code_lookup_once():
    """Load postal code lookup table once at module level - shared across all instances"""
    global _POSTAL_CODE_LOOKUP_CACHE

    if _POSTAL_CODE_LOOKUP_CACHE is not None:
        return _POSTAL_CODE_LOOKUP_CACHE

    try:
        logger.info("=" * 60)
        logger.info("POSTAL CODE LOOKUP INITIALIZATION")
        logger.info("=" * 60)

        # Log all paths being checked
        logger.info(f"Checking {len(POSTAL_CODE_PATHS)} postal code file paths:")
        for idx, path in enumerate(POSTAL_CODE_PATHS, 1):
            if path:
                exists = os.path.exists(path)
                logger.info(f"  {idx}. {path} - {'EXISTS' if exists else 'NOT FOUND'}")
            else:
                logger.info(f"  {idx}. (None/Empty)")

        # Try each path in order of preference with fast fail
        master_file_path = None
        for path in POSTAL_CODE_PATHS:
            if path and os.path.exists(path):
                # Quick file size check to avoid loading huge files
                try:
                    file_size = os.path.getsize(path)
                    if file_size > 50 * 1024 * 1024:  # Skip files > 50MB
                        logger.warning(f"Postal Code Lookup: File too large ({file_size/1024/1024:.1f}MB), skipping: {path}")
                        continue
                    logger.info(f"Selected postal code file: {path} ({file_size/1024:.1f} KB)")
                    master_file_path = path
                    break
                except OSError as e:
                    logger.warning(f"Error accessing {path}: {e}")
                    continue

        if not master_file_path:
            logger.warning("Postal Code Lookup: NO SUITABLE FILE FOUND - Using Google Maps API only")
            logger.info("=" * 60)
            _POSTAL_CODE_LOOKUP_CACHE = {}
            return _POSTAL_CODE_LOOKUP_CACHE

        logger.info(f"Loading postal code data...")

        # Detect file type and load accordingly
        file_extension = os.path.splitext(master_file_path)[1].lower()
        logger.info(f"File format: {file_extension.upper()}")

        if file_extension == '.csv':
            df = pd.read_csv(master_file_path, dtype={'postal_code': str})
            postal_col = 'postal_code'
            lat_col = 'Latitude'
            lng_col = 'Longitude'
        else:  # Excel format (.xlsx, .xls)
            df = pd.read_excel(master_file_path, dtype={'PostalCode': str})
            postal_col = 'PostalCode'
            lat_col = 'Latitude'
            lng_col = 'Longitude'

        total_rows = len(df)
        logger.info(f"Total rows in file: {total_rows:,}")
        logger.info(f"Columns: {postal_col}, {lat_col}, {lng_col}")

        # Create dictionary for fast lookup: {postal_code: (lat, lng)}
        lookup = {}
        row_count = 0
        skipped_rows = 0
        sample_codes = []

        for _, row in df.iterrows():
            row_count += 1
            if row_count % 20000 == 0:  # Progress indicator (adjusted for larger files)
                logger.info(f"Processing row {row_count:,}/{total_rows:,} ({row_count*100//total_rows}%)...")

            # Handle postal code formatting
            postal_code_raw = row.get(postal_col)
            if pd.notna(postal_code_raw) and postal_code_raw:
                try:
                    # Ensure 6-digit format (handles 5-digit codes by adding leading zero)
                    if isinstance(postal_code_raw, str):
                        postal_code = postal_code_raw.strip().zfill(6)
                    else:
                        postal_code = f"{int(float(postal_code_raw)):06d}"
                except (ValueError, TypeError):
                    skipped_rows += 1
                    continue
            else:
                skipped_rows += 1
                continue

            lat = row.get(lat_col)
            lng = row.get(lng_col)
            if pd.notna(lat) and pd.notna(lng):
                try:
                    lookup[postal_code] = (float(lat), float(lng))
                    # Store first 3 postal codes as samples
                    if len(sample_codes) < 3:
                        sample_codes.append((postal_code, lat, lng))
                except (ValueError, TypeError):
                    skipped_rows += 1
                    continue
            else:
                skipped_rows += 1

        logger.info("=" * 60)
        logger.info(f"POSTAL CODE LOADING COMPLETE")
        logger.info(f"Successfully loaded: {len(lookup):,} postal codes")
        logger.info(f"Skipped rows: {skipped_rows:,}")
        logger.info(f"Coverage: {len(lookup)*100//total_rows}% of file rows")

        # Log sample postal codes for verification
        if sample_codes:
            logger.info(f"Sample postal codes loaded:")
            for code, lat, lng in sample_codes:
                logger.info(f"  {code}: ({lat}, {lng})")

        logger.info("=" * 60)
        _POSTAL_CODE_LOOKUP_CACHE = lookup
        return _POSTAL_CODE_LOOKUP_CACHE

    except Exception as e:
        logger.error(f"Postal Code Lookup: FAILED - {e}")
        logger.warning("Continuing without postal code lookup, using Google Maps API only")
        _POSTAL_CODE_LOOKUP_CACHE = {}
        return _POSTAL_CODE_LOOKUP_CACHE

class GeocodingService:
    def __init__(self, use_google_api=True):
        self.use_google_api = use_google_api
        self.google_api_key = os.getenv('GOOGLE_MAPS_API_KEY')
        self._initialize_google_maps_clients()
        self.postal_code_lookup = _load_postal_code_lookup_once()  # Use shared global lookup
        self.geocode_stats = {'postal_matches': 0, 'api_calls': 0, 'failures': 0}
    
    def _initialize_google_maps_clients(self):
        """Initialize Google Maps API clients with status logging"""
        if not self.use_google_api:
            logger.info("Google Maps API: DISABLED BY USER - Using postal code lookup only")
            self.gmaps = None
            self.geolocator = None
            return

        if not self.google_api_key:
            logger.info("Google Maps API: NOT CONFIGURED - Using postal code lookup only")
            self.gmaps = None
            self.geolocator = None
            return

        try:
            self.gmaps = googlemaps.Client(key=self.google_api_key)
            self.geolocator = GoogleV3(api_key=self.google_api_key)
            logger.info("Google Maps API: CONFIGURED AND ENABLED")
        except Exception as e:
            logger.error(f"Google Maps API: FAILED - {e}")
            self.gmaps = None
            self.geolocator = None
    
    def geocode_by_postal_code(self, postal_code):
        """Get coordinates by postal code lookup"""
        if not postal_code or str(postal_code).strip() in ('None', '', 'nan'):
            return None, None
        
        try:
            # Normalize postal code to 6-digit format with leading zeros
            postal_code_str = str(postal_code).strip()
            # Handle Singapore postal codes with 'S' prefix
            if postal_code_str.upper().startswith('S') and len(postal_code_str) > 1:
                postal_code_str = postal_code_str[1:]  # Remove 'S' prefix
            postal_code_normalized = f"{int(float(postal_code_str)):06d}"
            
            if postal_code_normalized in self.postal_code_lookup:
                self.geocode_stats['postal_matches'] += 1
                lat, lng = self.postal_code_lookup[postal_code_normalized]
                return lat, lng
        except (ValueError, TypeError, AttributeError) as e:
            # Log invalid postal code format for debugging
            logger.warning(f"Invalid postal code format: {postal_code} - {e}")
        
        return None, None
    
    def geocode_by_address(self, address, country=None):
        """Get coordinates by Google Maps API using full address

        Args:
            address: The address to geocode
            country: Optional country code ('SINGAPORE' or 'MALAYSIA') to force region bias
        """
        if not address or str(address).strip() in ('', 'None', 'nan') or not self.geolocator:
            return None, None

        try:
            self.geocode_stats['api_calls'] += 1

            # Clean address and detect country
            address_str = str(address).strip()
            address_lower = address_str.lower()

            # Determine region parameter for API call
            region = None

            # PRIORITY: Use explicitly provided country parameter
            if country:
                if country == 'MALAYSIA':
                    region = 'my'  # Force Malaysia region bias
                    # Ensure Malaysia is in the address string
                    if 'malaysia' not in address_lower:
                        address_str += ', Malaysia'
                elif country == 'SINGAPORE':
                    region = 'sg'  # Force Singapore region bias
                    if 'singapore' not in address_lower:
                        address_str += ', Singapore'
            else:
                # Fallback: Detect from address text
                has_singapore = 'singapore' in address_lower
                has_malaysia = 'malaysia' in address_lower

                # If no country specified, try to detect from context
                if not has_singapore and not has_malaysia:
                    # Check for Malaysian states/regions in address
                    malaysian_indicators = ['johor', 'kuala lumpur', 'selangor', 'penang', 'perak', 'kedah', 'kelantan', 'terengganu', 'pahang', 'negeri sembilan', 'melaka', 'sabah', 'sarawak', 'perlis', 'putrajaya', 'labuan']
                    if any(indicator in address_lower for indicator in malaysian_indicators):
                        address_str += ', Malaysia'
                        region = 'my'
                    else:
                        # Default to Singapore for addresses without clear country indicators
                        address_str += ', Singapore'
                        region = 'sg'
                elif has_malaysia:
                    region = 'my'
                elif has_singapore:
                    region = 'sg'

            # Call geocoder with region parameter if available
            if region:
                location = self.geolocator.geocode(address_str, timeout=10, region=region)
                logger.debug(f"Geocoding with region bias: {region.upper()} for address: {address_str}")
            else:
                location = self.geolocator.geocode(address_str, timeout=10)

            if location:
                return location.latitude, location.longitude
            else:
                self.geocode_stats['failures'] += 1
                return None, None

        except Exception as e:
            self.geocode_stats['failures'] += 1
            logger.warning(f"Address geocoding failed for '{address}': {e}")
            return None, None
    
    def geocode(self, postal_code, address, country=None):
        """Main geocoding method: country-aware geocoding strategy

        Args:
            postal_code: Postal code to lookup
            address: Full address string
            country: Optional country code ('SINGAPORE' or 'MALAYSIA') for region bias

        Strategy:
            - SINGAPORE: Try postal code lookup first, then Google Maps API fallback
            - MALAYSIA: Only use Google Maps API (postal code lookup is Singapore-only)
            - No country specified: Try postal code lookup, then API fallback
        """
        # MALAYSIA: Skip postal code lookup (Singapore-only), use Google Maps API only
        if country == 'MALAYSIA':
            if self.use_google_api:
                lat, lng = self.geocode_by_address(address, country=country)
                if lat is not None and lng is not None:
                    return lat, lng, 'address'
            # If API disabled or geocoding failed, return None for Malaysia
            return None, None, 'failed'

        # SINGAPORE or unspecified country: Try postal code lookup first
        lat, lng = self.geocode_by_postal_code(postal_code)
        if lat is not None and lng is not None:
            return lat, lng, 'postal_code'

        # Fallback to address geocoding only if use_google_api is enabled
        if self.use_google_api:
            # Pass country parameter to force region bias (especially important for Malaysia)
            lat, lng = self.geocode_by_address(address, country=country)
            if lat is not None and lng is not None:
                return lat, lng, 'address'

        return None, None, 'failed'
    
    def get_stats(self):
        """Get geocoding statistics"""
        return self.geocode_stats.copy()

class ExcelTransformer:
    @staticmethod
    def detect_alliance_tokio_format(ws):
        """
        Detect Alliance-Tokio Marine format by checking:
        1. Merged cell K1:N1 (operating hours disclaimer)
        2. Row 1 has 'POSTAL\\nCODE' or 'POSTAL CODE' header
        3. Row 2 has 'MON - FRI' sub-header
        4. ZONE and ESTATE columns present
        """
        try:
            # Check for merged cells pattern
            has_operating_hours_merge = False
            for merged_range in ws.merged_cells.ranges:
                range_str = str(merged_range)
                if 'K1:N1' in range_str or 'K1' in range_str:
                    has_operating_hours_merge = True
                    break

            # Check row 1 headers
            row1 = [cell.value for cell in ws[1]]
            has_postal_code = any(
                'POSTAL' in str(h).upper() for h in row1 if h
            )
            has_zone_estate = (
                any('ZONE' in str(h).upper() for h in row1 if h) and
                any('ESTATE' in str(h).upper() for h in row1 if h)
            )

            # Check row 2 sub-headers
            row2 = [cell.value for cell in ws[2]]
            has_mon_fri_subheader = any('MON - FRI' in str(h) for h in row2 if h)

            is_alliance_tokio = (
                has_operating_hours_merge and
                has_postal_code and
                has_zone_estate and
                has_mon_fri_subheader
            )

            if is_alliance_tokio:
                logger.info("Detected Alliance-Tokio Marine format (multi-level headers with merged cells)")

            return is_alliance_tokio

        except Exception as e:
            logger.debug(f"Alliance-Tokio format detection failed: {e}")
            return False

    @staticmethod
    def unmerge_and_fill_cells(ws):
        """
        Unmerge all cells and fill merged cell values into all cells in the range.
        This ensures row-by-row reading works correctly.
        """
        import openpyxl

        # Get all merged ranges (need to copy list as we'll modify it)
        merged_ranges = list(ws.merged_cells.ranges)

        logger.info(f"Unmerging {len(merged_ranges)} merged cell ranges")

        for merged_range in merged_ranges:
            # Get the value from top-left cell
            top_left = merged_range.start_cell
            merge_value = top_left.value

            # Unmerge the range
            ws.unmerge_cells(str(merged_range))

            # Fill the value into all cells that were part of the merge
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    ws.cell(row=row, column=col).value = merge_value

        return ws

    @staticmethod
    def get_alliance_tokio_headers(ws):
        """
        Reconstruct headers from unmerged rows 1-2.
        For most columns: use row 1 value
        For columns K-N: use row 2 values (MON-FRI, SAT, SUN, PUBLIC HOLIDAYS)
        """
        headers = []

        for col_idx in range(1, 16):  # Columns A-O (1-15)
            if col_idx <= 10 or col_idx == 15:  # Columns A-J and O
                # Use row 1 value, clean it
                value = ws.cell(row=1, column=col_idx).value
                if value:
                    value = str(value).replace('\n', ' ').replace('\r', ' ').strip()
                headers.append(value)
            else:  # Columns K-N (11-14)
                # Use row 2 sub-headers for operating hours
                value = ws.cell(row=2, column=col_idx).value
                if value:
                    value = str(value).strip()
                headers.append(value)

        logger.debug(f"Alliance-Tokio headers: {headers}")
        return headers

    @staticmethod
    def convert_alliance_hours_to_standard(mon_fri, sat, sun, ph):
        """
        Convert Alliance-Tokio operating hours format to standard AM/PM/NIGHT format.

        Alliance format: 4 columns with detailed hours or "Closed"
        Standard format: "AM hours/PM hours/NIGHT hours" or "CLOSED/CLOSED/CLOSED"

        Strategy: Extract time ranges and categorize into AM/PM/NIGHT slots
        """
        def parse_time_slots(hours_str):
            """Parse hours string into AM/PM/NIGHT slots"""
            if pd.isna(hours_str) or str(hours_str).strip().upper() in ('CLOSED', '', 'NAN'):
                return 'CLOSED', 'CLOSED', 'CLOSED'

            hours_str = str(hours_str).strip()

            # Extract all time ranges
            import re
            time_pattern = r'(\d{1,2}[:.]\d{2}\s*(?:am|pm))\s*-\s*(\d{1,2}[:.]\d{2}\s*(?:am|pm))'
            matches = re.findall(time_pattern, hours_str, re.IGNORECASE)

            if not matches:
                # No parseable times, return as-is for AM slot
                return hours_str, 'CLOSED', 'CLOSED'

            am_slots = []
            pm_slots = []
            night_slots = []

            for start, end in matches:
                # Parse start time
                start_clean = start.lower().replace('.', ':')

                # Categorize by start time
                if 'am' in start_clean or ('12:' in start_clean and 'pm' in start_clean):
                    # Morning slot (before noon)
                    am_slots.append(f"{start} - {end}")
                elif '6' in start_clean.split(':')[0] and 'pm' in start_clean:
                    # Evening/Night slot (6pm or later)
                    night_slots.append(f"{start} - {end}")
                else:
                    # Afternoon slot
                    pm_slots.append(f"{start} - {end}")

            # Combine slots
            am_result = ', '.join(am_slots) if am_slots else 'CLOSED'
            pm_result = ', '.join(pm_slots) if pm_slots else 'CLOSED'
            night_result = ', '.join(night_slots) if night_slots else 'CLOSED'

            return am_result, pm_result, night_result

        # Parse each day type
        mon_fri_am, mon_fri_pm, mon_fri_night = parse_time_slots(mon_fri)
        sat_am, sat_pm, sat_night = parse_time_slots(sat)
        sun_am, sun_pm, sun_night = parse_time_slots(sun)
        ph_am, ph_pm, ph_night = parse_time_slots(ph)

        # Build standard format strings
        weekday = f"{mon_fri_am}/{mon_fri_pm}/{mon_fri_night}"
        saturday = f"{sat_am}/{sat_pm}/{sat_night}"
        sunday = f"{sun_am}/{sun_pm}/{sun_night}"
        holiday = f"{ph_am}/{ph_pm}/{ph_night}"

        return weekday, saturday, sunday, holiday

    @staticmethod
    def safe_read_excel(file_path, sheet_name=None, **kwargs):
        """Safely read Excel file with fallback for corrupted XML metadata"""
        try:
            # First attempt: Normal read
            return pd.read_excel(file_path, sheet_name=sheet_name, **kwargs)
        except ValueError as e:
            if "could not assign names" in str(e) or "invalid XML" in str(e):
                # Openpyxl XML corruption - monkey patch the problematic function
                logger.warning(f"Excel file has corrupted metadata (print titles/names), using patched openpyxl")
                import openpyxl
                from openpyxl.reader.workbook import WorkbookParser

                # Save original method
                original_assign_names = WorkbookParser.assign_names

                def patched_assign_names(self):
                    """Patched version that skips invalid print titles"""
                    try:
                        original_assign_names(self)
                    except ValueError as ve:
                        if "not a valid print titles definition" in str(ve):
                            logger.warning(f"Skipping invalid print title: {ve}")
                            # Continue without assigning this name
                            pass
                        else:
                            raise

                # Apply patch
                WorkbookParser.assign_names = patched_assign_names

                try:
                    # Retry with patched openpyxl
                    result = pd.read_excel(file_path, sheet_name=sheet_name, **kwargs)
                    return result
                finally:
                    # Restore original method
                    WorkbookParser.assign_names = original_assign_names
            else:
                raise

    @staticmethod
    def find_header_row(file_path, sheet_name=None):
        """Find the actual header row by looking for clinic-related keywords"""
        df_raw = ExcelTransformer.safe_read_excel(file_path, sheet_name=sheet_name, header=None)

        for idx, row in df_raw.iterrows():
            row_values = [str(val) for val in row.values if pd.notna(val)]
            row_text = ' '.join(row_values).lower()

            # Enhanced header detection patterns
            header_patterns = [
                # Primary pattern: S/N with clinic ID
                ('s/n' in row_text and 'clinic' in row_text and 'id' in row_text),
                # SP Clinic specific patterns
                ('s/n' in row_text and 'specialty' in row_text and 'sp code' in row_text and 'doctor' in row_text),
                ('s/n' in row_text and 'sp code' in row_text and 'clinic name' in row_text and 'address1' in row_text),
                # MY GP List specific patterns
                ('s/n' in row_text and 'clinic code' in row_text and 'city' in row_text and 'state' in row_text),
                ('s/n' in row_text and 'clinic name' in row_text and 'address1' in row_text),
                ('s/n' in row_text and 'clinic' in row_text and 'tel' in row_text and 'operation' in row_text),
                # Alternative patterns for different sheet layouts
                ('s/n' in row_text and 'region' in row_text and 'area' in row_text),
                ('s/n' in row_text and 'clinic' in row_text and 'name' in row_text),
                ('no.' in row_text and 'clinic' in row_text and 'name' in row_text),
                # For termination sheets
                ('no.' in row_text and 'region' in row_text and 'area' in row_text),
                # TCM sheet specific patterns
                ('s/n' in row_text and 'clinic' in row_text and 'postal' in row_text),
                ('master code' in row_text and 'clinic' in row_text and 'postal' in row_text),
                ('master code' in row_text and 'physician' in row_text and 'charge' in row_text),
                ('master code' in row_text and 'tel' in row_text and len(row_values) >= 8),
                # General patterns that indicate header rows
                ('clinic' in row_text and 'postal' in row_text and 'tel' in row_text),
                ('code' in row_text and 'clinic' in row_text and 'address' in row_text),
                ('provider' in row_text and 'name' in row_text and len(row_values) >= 5),
                # Enhanced minimum viable header patterns
                ('region' in row_text and 'clinic' in row_text and len(row_values) >= 5),
                ('city' in row_text and 'clinic' in row_text and 'tel' in row_text),
                ('state' in row_text and 'clinic name' in row_text and len(row_values) >= 6)
            ]

            if any(pattern for pattern in header_patterns):
                return idx

        # If no clear header found, look for the first row with substantial data
        for idx, row in df_raw.iterrows():
            non_null_count = sum(1 for val in row.values if pd.notna(val) and str(val).strip())
            if non_null_count >= 5:  # At least 5 non-empty columns
                return idx

        # Final fallback
        return 4

    @staticmethod
    def classify_sheets(sheet_names):
        """Enhanced classification for various sheet types"""
        panel_sheets = []
        termination_sheets = []

        for sheet in sheet_names:
            sheet_lower = sheet.lower()
            # Identify termination sheets
            if any(term in sheet_lower for term in ['terminat', 'remov', 'cancel', 'delist']):
                termination_sheets.append(sheet)
            # Enhanced panel sheet identification
            elif any(panel in sheet_lower for panel in [
                'gp', 'tcm', 'dental', 'clinic', 'panel',  # Original patterns
                'sp list', 'sp clinic', 'specialist',      # Specialist patterns
                'sg', 'my', 'msia', 'malaysia', 'singapore', # Country patterns
                'blue', 'red', 'flexi', 'aia',            # Plan type patterns
                'medical', 'health', 'doctor',             # Healthcare patterns
                'alliance', 'tokio', 'marine', 'provider'  # Insurance/Provider patterns
            ]):
                panel_sheets.append(sheet)

        return panel_sheets, termination_sheets

    @staticmethod
    def normalize_code(value):
        """Normalize provider code or postal code by removing unnecessary decimal points

        Excel often stores numbers as floats (e.g., 40088.0, 518180.0)
        This function converts them to clean strings without decimals (e.g., "40088", "518180")
        """
        if value is None or pd.isna(value):
            return None

        value_str = str(value).strip()

        if value_str.lower() == 'nan' or value_str == '':
            return None

        # Remove .0 suffix if present (e.g., "40088.0" -> "40088")
        if '.' in value_str:
            try:
                # Try to convert to float then int to remove decimal
                float_val = float(value_str)
                # Check if it's a whole number
                if float_val == int(float_val):
                    return str(int(float_val))
            except (ValueError, OverflowError):
                pass

        return value_str

    @staticmethod
    def extract_terminated_clinic_ids(file_path, termination_sheets):
        """Extract clinic IDs and postal codes from termination sheets

        Returns a set of tuples: {(provider_code, postal_code), ...}
        This ensures termination only occurs when BOTH provider code AND postal code match
        """
        terminated_entries = set()

        for sheet in termination_sheets:
            try:
                # Find header row for termination sheet
                header_row = ExcelTransformer.find_header_row(file_path, sheet)
                df = ExcelTransformer.safe_read_excel(file_path, sheet_name=sheet, header=header_row)
                df.columns = df.columns.str.strip()

                # Look for clinic ID/provider code column (various possible names)
                id_columns = [col for col in df.columns if 'clinic' in col.lower() and 'id' in col.lower()]
                if not id_columns:
                    id_columns = [col for col in df.columns if 'provider' in col.lower() and ('code' in col.lower() or 'id' in col.lower())]
                if not id_columns:
                    id_columns = [col for col in df.columns if 'code' in col.lower()]

                # Look for postal code column (various naming patterns)
                postal_columns = [col for col in df.columns if 'postal' in col.lower() and 'code' in col.lower()]
                if not postal_columns:
                    postal_columns = [col for col in df.columns if 'post' in col.lower() and 'code' in col.lower()]
                if not postal_columns:
                    postal_columns = [col for col in df.columns if col.lower() == 'postalcode']
                if not postal_columns:
                    # Match standalone "POSTAL" or "POST" column
                    postal_columns = [col for col in df.columns if col.lower().strip() == 'postal']
                if not postal_columns:
                    postal_columns = [col for col in df.columns if col.lower().strip() == 'post']

                if id_columns and postal_columns:
                    # Both provider code and postal code found - use dual matching
                    id_col = id_columns[0]
                    postal_col = postal_columns[0]

                    for _, row in df.iterrows():
                        # Normalize both provider code and postal code
                        provider_code = ExcelTransformer.normalize_code(row[id_col])
                        postal_code = ExcelTransformer.normalize_code(row[postal_col])

                        # Only add if both values are valid
                        if provider_code and postal_code:
                            terminated_entries.add((provider_code, postal_code))

                    logger.info(f"Extracted {len(terminated_entries)} terminated entries (provider code + postal code) from sheet '{sheet}'")

                elif id_columns:
                    # Postal code column not found - try to extract from address columns
                    logger.warning(f"Postal code column not found in termination sheet '{sheet}' - attempting extraction from address columns")

                    # Look for address columns
                    address_columns = [col for col in df.columns if 'address' in col.lower()]

                    extracted_count = 0
                    provider_only_count = 0

                    for _, row in df.iterrows():
                        provider_code = ExcelTransformer.normalize_code(row[id_columns[0]])
                        if not provider_code:
                            continue

                        # Try to extract postal code from address columns
                        postal_code = None
                        if address_columns:
                            # Combine all address columns
                            address_parts = [str(row[col]) for col in address_columns if pd.notna(row[col])]
                            combined_address = ' '.join(address_parts)

                            # Extract postal code from combined address
                            postal_code = ExcelTransformer.extract_postal_code(combined_address)
                            if postal_code:
                                postal_code = ExcelTransformer.normalize_code(postal_code)

                        if postal_code:
                            # Successfully extracted postal code - use dual matching
                            terminated_entries.add((provider_code, postal_code))
                            extracted_count += 1
                        else:
                            # No postal code found - fallback to provider code only
                            terminated_entries.add((provider_code, None))
                            provider_only_count += 1

                    if extracted_count > 0:
                        logger.info(f"Extracted {extracted_count} terminated entries with postal codes from address columns")
                    if provider_only_count > 0:
                        logger.info(f"Extracted {provider_only_count} terminated entries with provider code only (no postal code found)")

                    logger.info(f"Total from sheet '{sheet}': {extracted_count + provider_only_count} entries ({extracted_count} with postal, {provider_only_count} without)")
                else:
                    logger.warning(f"Could not find clinic ID column in termination sheet '{sheet}'")

            except Exception as e:
                logger.error(f"Error processing termination sheet '{sheet}': {e}")

        logger.info(f"Total terminated entries: {len(terminated_entries)}")
        return terminated_entries

    @staticmethod
    def sanitize_filename(name):
        """Sanitize sheet name for use in filename"""
        # Remove invalid characters and replace spaces
        sanitized = re.sub(r'[<>:"/\\|?*]', '', name)
        sanitized = re.sub(r'\s+', '_', sanitized)
        return sanitized.strip('_')

    @staticmethod
    def format_postal_codes(df):
        """Format PostalCode column as string to preserve leading zeros and prevent decimal notation"""
        if 'PostalCode' in df.columns:
            def format_code(x):
                if pd.isna(x) or str(x).strip() in ('', 'None', 'nan'):
                    return ''
                # Convert to string and remove any decimal points
                code_str = str(x).strip()
                # If it's a number with decimal (e.g., "330047.0" or "80050.0"), convert to int first
                if '.' in code_str:
                    try:
                        code_str = str(int(float(code_str)))
                    except (ValueError, OverflowError):
                        pass
                # Return as-is without padding - preserve original length (5 digits for Malaysia, 6 for Singapore)
                return code_str
            
            df['PostalCode'] = df['PostalCode'].apply(format_code)
        return df
    
    @staticmethod
    def write_excel_with_text_postal_codes(df, file_path):
        """Write DataFrame to Excel with PostalCode column formatted as text"""
        from openpyxl import load_workbook
        from openpyxl.styles import numbers
        
        # Write the DataFrame to Excel
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        # Open the workbook and apply text formatting to PostalCode column
        if 'PostalCode' in df.columns:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Find the PostalCode column index
            postal_col_idx = None
            for idx, cell in enumerate(ws[1], start=1):
                if cell.value == 'PostalCode':
                    postal_col_idx = idx
                    break
            
            if postal_col_idx:
                # Apply text format to the entire PostalCode column
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=postal_col_idx)
                    cell.number_format = '@'  # '@' is the Excel format code for text
            
            wb.save(file_path)
        
        return file_path


    @staticmethod
    def map_columns(df_columns):
        """Robust column mapping with fuzzy matching and multiple file format support"""
        import re
        from difflib import get_close_matches

        column_mapping = {}

        # Enhanced column mapping patterns with more variations
        mappings = {
            'clinic_id': [
                'ihp clinic id', 'provider code', 'clinic id', 'id', 'clinic code',
                'provider id', 'clinic identifier', 'code', 'clinic no', 'clinic number',
                'master code', 'master id',  # TCM sheet specific
                'sp code', 'sp id'  # SP clinic specific
            ],
            'clinic_name': [
                'clinic name', 'name', 'clinic', 'provider name', 'facility name',
                'medical center', 'medical centre', 'center name', 'centre name'
            ],
            'region': [
                'region', 'zone', 'district', 'sector', 'territory', 'location',
                'geographical region', 'geo region', 'state', 'province', 'city'  # Added city for MY sheets
            ],
            'area': [
                'area', 'estate', 'neighbourhood', 'neighborhood', 'locality',
                'precinct', 'town', 'suburb', 'community', 'district area', 'state'  # Added state for MY sheets
            ],
            'address': [
                'address', 'full address', 'complete address', 'location address',
                'physical address', 'street address', 'mailing address', 'address1'
            ],
            'telephone': [
                'tel no.', 'tel', 'phone', 'telephone', 'contact', 'contact no',
                'contact number', 'phone number', 'tel number', 'mobile', 'contact no.', 'tel no'
            ],
            'remarks': [
                'remarks', 'comment', 'note', 'remark', 'comments', 'notes',
                'additional info', 'special notes', 'observation', 'memo'
            ],
            'operating_hours': [
                'operating hours', 'hours', 'business hours', 'clinic hours',
                'opening hours', 'operation hours', 'working hours', 'service hours'
            ],
            'mon_fri_am': [
                'mon - fri (am)', 'monday - friday', 'weekday am', 'mon-fri am',
                'operating hours\nmon-fri', 'operating hours mon-fri', 'weekdays am',
                'operating hours \nmon - fri', 'operating hours mon - fri',  # TCM format
                'operating hours (monday - friday)',  # Income/Adept format
                'operation hours', 'mon to fri',  # MY GP List format
                'mon - fri',  # AIA SP format
                'weekdays',  # AIA dental format - direct weekdays column
                'operating hour monday - friday'  # Singlife format - direct text extraction (cleaned with spaces)
            ],
            'mon_fri_pm': [
                'mon - fri (pm)', 'monday - friday (evening)', 'weekday pm', 'mon-fri pm', 'weekdays pm'
            ],
            'mon_fri_night': [
                'mon - fri (night)', 'weekday night', 'mon-fri night', 'weekdays night'
            ],
            'sat_am': ['sat (am)', 'saturday', 'sat am', 'saturday am'],
            'sat_pm': ['sat (pm)', 'sat pm', 'saturday pm'],
            'sat_night': ['sat (night)', 'sat night', 'saturday night'],
            'sat_simple': ['sat', 'operating hours (saturday)'],  # Simple Saturday column + Income format
            'sun_am': ['sun (am)', 'sunday', 'sun am', 'sunday am'],
            'sun_pm': ['sun (pm)', 'sun pm', 'sunday pm'],
            'sun_night': ['sun (night)', 'sun night', 'sunday night'],
            'sun_simple': ['sun', 'operating hours (sunday)'],  # Simple Sunday column + Income format
            'holiday_am': ['public holiday (am)', 'public holiday', 'holiday am', 'ph am', 'ph', 'publicday'],
            'holiday_pm': ['public holiday (pm)', 'holiday pm', 'ph pm'],
            'holiday_night': ['public holiday (night)', 'holiday night', 'ph night'],
            'holiday_simple': ['holiday', 'ph', 'operating hours (holiday(s))', 'operating hours (holidays)'],  # Simple Holiday column + Income format
            # Address components for composite address construction
            'address_blk': ['blk', 'block', 'building no', 'bldg no', 'unit block', 'blk & road name'],
            'address_road': ['road name', 'street name', 'street', 'road', 'avenue', 'ave', 'blk & road name'],
            'address_unit': ['unit no.', 'unit no', 'unit', '#', 'suite', 'level', 'unit & building name'],
            'address_building': ['building name', 'building', 'bldg name', 'complex name', 'unit & building name'],
            'postal_code': ['postal code', 'postcode', 'zip code', 'zip', 'postal'],
            # TCM-specific field for doctor information
            'doctor_name': ['physician - in - charge', 'physician in charge', 'doctor', 'physician', 'practitioner'],
            # SP clinic specific fields
            'specialty': ['specialty', 'speciality', 'medical specialty', 'specialization', 'department'],
            'address1': ['address1', 'address 1', 'primary address', 'street address'],
            'address2': ['address2', 'address 2', 'secondary address', 'unit number'],
            'address3': ['address3', 'address 3', 'building name', 'complex name'],
            'address4': ['address4', 'address 4', 'postal address', 'location detail']
        }

        # Convert all column names to lowercase and clean for comparison
        df_cols_lower = {}
        df_cols_cleaned = {}
        df_cols_original = {}  # Keep track of original column names

        for col in df_columns:
            if pd.notna(col) and isinstance(col, str):
                col_clean = col.lower().strip()
                # Remove extra whitespace and newlines
                col_clean = re.sub(r'\s+', ' ', col_clean)
                df_cols_lower[col_clean] = col
                df_cols_cleaned[col_clean] = col
                df_cols_original[col] = col_clean

        logger.debug(f"Available columns (cleaned): {list(df_cols_cleaned.keys())}")

        # Phase 1: Exact pattern matching
        for expected_col, patterns in mappings.items():
            for pattern in patterns:
                if pattern in df_cols_lower:
                    column_mapping[expected_col] = df_cols_lower[pattern]
                    logger.debug(f"Exact match: {expected_col} -> {pattern} -> {df_cols_lower[pattern]}")
                    break

        # Phase 2: Handle sequential operating hours columns (MY GP List format)
        # Look for operation hours followed by unnamed columns
        if 'mon_fri_am' in column_mapping:
            # Find the index of operation hours column
            operation_hours_col = column_mapping['mon_fri_am']
            col_list = list(df_columns)

            try:
                operation_hours_idx = col_list.index(operation_hours_col)
                logger.debug(f"Found operation hours at index {operation_hours_idx}: {operation_hours_col}")

                # Check for sequential unnamed columns
                if operation_hours_idx + 1 < len(col_list):
                    next_col = col_list[operation_hours_idx + 1]
                    if pd.notna(next_col) and ('unnamed' in str(next_col).lower() or str(next_col).lower().strip() == 'sat'):
                        column_mapping['sat_simple'] = next_col
                        logger.debug(f"Sequential match: sat_simple -> {next_col}")

                if operation_hours_idx + 2 < len(col_list):
                    next_col = col_list[operation_hours_idx + 2]
                    if pd.notna(next_col) and ('unnamed' in str(next_col).lower() or 'sun' in str(next_col).lower()):
                        column_mapping['sun_simple'] = next_col
                        logger.debug(f"Sequential match: sun_simple -> {next_col}")

                if operation_hours_idx + 3 < len(col_list):
                    next_col = col_list[operation_hours_idx + 3]
                    if pd.notna(next_col) and ('unnamed' in str(next_col).lower() or 'ph' in str(next_col).lower() or 'holiday' in str(next_col).lower()):
                        column_mapping['holiday_simple'] = next_col
                        logger.debug(f"Sequential match: holiday_simple -> {next_col}")

            except ValueError:
                pass  # Column not found in list

        # Phase 3: Fuzzy matching for unmapped essential columns
        essential_columns = ['clinic_name', 'telephone']
        # Only add region/area if we have good candidates (avoid false matches)
        has_region_candidate = any(pattern in col.lower() for col in df_cols_cleaned.keys() for pattern in ['region', 'zone', 'state', 'city'])
        has_area_candidate = any(pattern in col.lower() for col in df_cols_cleaned.keys() for pattern in ['area', 'district', 'state'])

        if has_region_candidate:
            essential_columns.append('region')
        if has_area_candidate:
            essential_columns.append('area')

        for expected_col in essential_columns:
            if expected_col not in column_mapping:
                # Get all patterns for this column
                all_patterns = mappings.get(expected_col, [])

                # Try fuzzy matching with higher threshold for region/area
                cutoff = 0.8 if expected_col in ['region', 'area'] else 0.6
                for pattern in all_patterns:
                    # Find close matches with similarity threshold
                    close_matches = get_close_matches(
                        pattern,
                        list(df_cols_cleaned.keys()),
                        n=1,
                        cutoff=cutoff
                    )
                    if close_matches:
                        matched_col = df_cols_cleaned[close_matches[0]]
                        column_mapping[expected_col] = matched_col
                        logger.debug(f"Fuzzy match: {expected_col} -> {pattern} ~= {close_matches[0]} -> {matched_col}")
                        break

        # Phase 4: Keyword-based matching for remaining columns (with better specificity)
        remaining_mappings = {
            'clinic_id': ['clinic code', 'clinic id', 'provider id', 'provider code'],  # More specific patterns
            'address': ['address1', 'address', 'location'],
            'remarks': ['remark', 'comment', 'note'],
        }

        for expected_col, keywords in remaining_mappings.items():
            if expected_col not in column_mapping:
                for col_name in df_cols_cleaned.keys():
                    # Use more specific matching - keyword must be substantial part of column name
                    for keyword in keywords:
                        if keyword in col_name and len(keyword) / len(col_name) > 0.4:  # At least 40% match
                            column_mapping[expected_col] = df_cols_cleaned[col_name]
                            logger.debug(f"Keyword match: {expected_col} -> {keyword} in {col_name} -> {df_cols_cleaned[col_name]}")
                            break
                    if expected_col in column_mapping:
                        break

        return column_mapping

    @staticmethod
    def infer_columns_from_data(df):
        """Infer column structure when headers are not clear"""
        inferred_mapping = {}

        # Check if we need to infer columns (either few meaningful names OR column names look like data)
        meaningful_cols = [col for col in df.columns if pd.notna(col) and isinstance(col, str) and len(str(col).strip()) > 0]

        # Detect if column names look like data rather than headers
        data_like_columns = 0
        for col in meaningful_cols[:6]:  # Check first 6 columns
            col_str = str(col).upper()
            if any(indicator in col_str for indicator in ['SINGAPORE', 'CLINIC', 'MEDICAL', 'CENTRE', 'AVENUE', 'ROAD', 'STREET', 'AM', 'PM']):
                data_like_columns += 1

        should_infer = len(meaningful_cols) < 3 or data_like_columns >= 2

        if should_infer:
            # Assume standard clinic data structure based on position
            cols = list(df.columns)

            if len(cols) >= 5:
                # Standard pattern: [S/N, Region, Area, Clinic_ID, Clinic_Name, Address, ...]
                inferred_mapping = {
                    'region': cols[1] if len(cols) > 1 else None,
                    'area': cols[2] if len(cols) > 2 else None,
                    'clinic_id': cols[3] if len(cols) > 3 else None,
                    'clinic_name': cols[4] if len(cols) > 4 else None,
                    'address': cols[5] if len(cols) > 5 else None,
                    'telephone': cols[6] if len(cols) > 6 else None,
                    'mon_fri_am': cols[7] if len(cols) > 7 else None,
                    'sat_am': cols[9] if len(cols) > 9 else None,
                    'sun_am': cols[10] if len(cols) > 10 else None,
                    'holiday_am': cols[11] if len(cols) > 11 else None
                }

                # Filter out None values
                inferred_mapping = {k: v for k, v in inferred_mapping.items() if v is not None}

                logger.debug(f"Inferred column mapping based on position: {inferred_mapping}")

        return inferred_mapping
    
    @staticmethod
    def extract_postal_code(address, country=None):
        """Extract postal code from address based on country format"""
        if pd.isna(address):
            return None

        address_str = str(address).strip()

        # Detect country from address if not provided
        if country is None:
            address_lower = address_str.lower()

            # PRIORITY 1: Check for explicit "SINGAPORE" keyword first
            # This prevents false positives like "Penang Road, Singapore" being detected as Malaysia
            if 'singapore' in address_lower:
                country = 'SINGAPORE'
            else:
                # PRIORITY 2: Check for Malaysian indicators
                malaysian_indicators = [
                    'malaysia', 'johor', 'kuala lumpur', 'selangor', 'penang', 'perak',
                    'kedah', 'kelantan', 'terengganu', 'pahang', 'negeri sembilan',
                    'melaka', 'sabah', 'sarawak', 'perlis', 'putrajaya', 'labuan',
                    # Additional Malaysian cities and areas
                    'kulai', 'skudai', 'pasir gudang', 'ulu tiram', 'masai', 'gelang patah',
                    'johor bahru', 'kl', 'shah alam', 'petaling jaya', 'bandar', 'taman'
                ]
                # Also check for Malaysian postal code patterns (5 digits vs Singapore's 6)
                has_5_digit = bool(re.search(r'\b\d{5}\b', address_str))
                has_6_digit = bool(re.search(r'\b\d{6}\b', address_str))

                is_malaysian = any(indicator in address_lower for indicator in malaysian_indicators)
                # If we find 5-digit codes but no 6-digit codes, likely Malaysian
                if has_5_digit and not has_6_digit:
                    is_malaysian = True

                country = 'MALAYSIA' if is_malaysian else 'SINGAPORE'

        if country == 'SINGAPORE':
            # Singapore: Look for SINGAPORE followed by 6 digits
            match = re.search(r'SINGAPORE\s+(\d{6})', address_str, re.IGNORECASE)
            if match:
                return match.group(1)
            # Fallback: Look for 6-digit patterns in Singapore addresses
            matches = re.findall(r'\b(\d{6})\b', address_str)
            return matches[-1] if matches else None  # Return last 6-digit number found

        elif country == 'MALAYSIA':
            # Malaysia: Look for 5-digit postal codes in various formats
            # Use multiple patterns to catch different formats
            patterns = [
                # Pattern 1: Standalone 5-digit codes (81300 SKUDAI, JOHOR)
                r'\b(\d{5})\b',
                # Pattern 2: City followed by postal code (KULAI 81000)
                r'\b[A-Za-z\s]+\s+(\d{5})',
                # Pattern 3: Postal code at end (TAMAN PERLING, 81200)
                r',\s*(\d{5})\s*$',
                # Pattern 4: Postal code before end tokens
                r'(\d{5})(?=\s*(?:$|,|\s+(?:JOHOR|SELANGOR|MALAYSIA)))',
                # Pattern 5: Any 5-digit sequence (most permissive)
                r'(\d{5})'
            ]

            for pattern in patterns:
                matches = re.findall(pattern, address_str, re.IGNORECASE)
                if matches:
                    # Return the first match found
                    return matches[0]

        return None
    
    @staticmethod
    def combine_phone_remarks(phone, remarks):
        """Combine telephone number with remarks"""
        phone_str = str(phone) if pd.notna(phone) else ""
        remarks_str = str(remarks) if pd.notna(remarks) else ""

        if remarks_str and remarks_str.lower() != 'nan':
            return f"{phone_str} - {remarks_str}"
        return phone_str

    @staticmethod
    def _is_truly_empty(value):
        """
        Check if value is truly empty (NaN/null or empty string/whitespace).
        Note: 'CLOSED' is NOT considered empty as it's explicit data.

        Args:
            value: Value to check

        Returns:
            bool: True if value is truly empty, False otherwise
        """
        if pd.isna(value):
            return True
        value_str = str(value).strip().lower()
        return value_str in ('', 'nan', 'none')

    @staticmethod
    def normalize_time_range(start, end):
        """
        Normalize time range to standard format.
        Handles various formats: 0900, 9AM, 9:00AM, 17, etc.

        Args:
            start: Start time string
            end: End time string

        Returns:
            str: Normalized time range (e.g., "0900-1230" or "9:00AM-5:00PM")
        """
        def normalize_single_time(time_str):
            """Normalize a single time value"""
            time_str = str(time_str).strip()

            # Check if it has AM/PM
            has_ampm = bool(re.search(r'(am|pm)', time_str, re.IGNORECASE))

            if has_ampm:
                # Preserve AM/PM format, just ensure colon for readability
                if ':' not in time_str and len(re.findall(r'\d', time_str)) >= 3:
                    # Insert colon: 900AM -> 9:00AM
                    match = re.match(r'(\d{1,2})(\d{2})\s*(am|pm)', time_str, re.IGNORECASE)
                    if match:
                        return f"{match.group(1)}:{match.group(2)}{match.group(3).upper()}"
                return time_str.upper()
            else:
                # 24hr format - ensure 4 digits
                digits_only = re.sub(r'\D', '', time_str)
                if len(digits_only) == 3:
                    # 900 -> 0900
                    digits_only = '0' + digits_only
                elif len(digits_only) == 2:
                    # Assume hours only: 9 -> 0900, 17 -> 1700
                    digits_only = digits_only + '00'
                elif len(digits_only) == 1:
                    # Single digit hour: 9 -> 0900
                    digits_only = '0' + digits_only + '00'

                return digits_only[:4].zfill(4)

        normalized_start = normalize_single_time(start)
        normalized_end = normalize_single_time(end)

        return f"{normalized_start}-{normalized_end}"

    @staticmethod
    def extract_hours_from_remarks(remarks_text):
        """
        Extract operating hours from remarks text with intelligent parsing.
        Handles various patterns like:
        - "(dental) Mon-Fri:0900-1230,1400-1600;Sat/Sun:0900-1200"
        - "Mon - Wed : 1030 to 2200"
        - "THUR: 0830 TO 1230, 1400 TO 1700"
        - "EVE OF PH: HALF DAY 9AM TO 1PM"

        Args:
            remarks_text: Raw remarks string from Excel file

        Returns:
            dict: {
                'weekdays': str or None,    # Mon-Fri hours
                'saturday': str or None,    # Saturday hours
                'sunday': str or None,      # Sunday hours
                'publicday': str or None,   # Public holiday hours
                'metadata': dict            # Parsing info
            }
        """
        logger = logging.getLogger(__name__)

        # Initialize empty result
        result = {
            'weekdays': None,
            'saturday': None,
            'sunday': None,
            'publicday': None,
            'metadata': {'confidence': 0.0, 'patterns_found': []}
        }

        # Validate input
        if pd.isna(remarks_text) or not str(remarks_text).strip():
            result['metadata']['reason'] = 'empty_input'
            return result

        try:
            # Normalize text
            text = str(remarks_text).strip()
            # Remove common prefixes
            text = re.sub(r'^\(dental\)\s*', '', text, flags=re.IGNORECASE)
            text = re.sub(r'^\(DENTAL\)\s*', '', text)

            # Define day patterns (case-insensitive)
            # Order matters - check ranges and combos first, then individual days
            day_patterns = {
                # Ranges (check these first)
                'mon_to_fri': r'mon(?:day)?\s*(?:-|to)\s*fri(?:day)?',
                'mon_to_wed': r'mon(?:day)?\s*(?:-|to)\s*wed(?:nesday)?',
                'thu_to_fri': r'thu(?:r(?:s)?)?\s*(?:-|to)\s*fri(?:day)?',
                # Combinations (slash-separated)
                'day_combo': r'(?:mon|tue|wed|thu(?:r)?|fri|sat|sun)(?:\s*/\s*(?:mon|tue|wed|thu(?:r)?|fri|sat|sun))+',
                # General keywords
                'weekdays': r'weekdays?',
                # Special cases
                'public_holiday': r'(?:public\s+holiday|ph|public\s+hol(?:iday)?)',
                'eve_of_ph': r'eve\s+of\s+(?:public\s+holiday|ph)',
                # Individual days (check last)
                'monday': r'\bmon(?:day)?\b',
                'tuesday': r'\btue(?:s(?:day)?)?\b',
                'wednesday': r'\bwed(?:nesday)?\b',
                'thursday': r'\bthu(?:r(?:s(?:day)?)?)?\b',
                'friday': r'\bfri(?:day)?\b',
                'saturday': r'\bsat(?:urday)?\b',
                'sunday': r'\bsun(?:day)?\b',
            }

            # Define time patterns (order matters - more specific first)
            time_patterns = [
                r'(\d{1,2}(?::\d{2})?\s*(?:am|pm))\s*(?:-|to)\s*(\d{1,2}(?::\d{2})?\s*(?:am|pm))',  # 9AM-5PM or 9AM to 5PM
                r'(\d{4})\s*-\s*(\d{4})',  # 0900-1230
                r'(\d{4})\s+(?:to|TO)\s+(\d{4})',  # 0900 TO 1230
                r'(\d{1,2})\s+(?:to|TO)\s+(\d{1,2})(?=\s|,|;|$)',  # 9 to 17 (with lookahead to ensure it's not part of a word)
            ]

            # Split by semicolons and newlines to get segments
            segments = re.split(r'[;\n]', text)

            for segment in segments:
                segment = segment.strip()
                if not segment:
                    continue

                # Try to match day pattern followed by colon and time ranges
                for day_key, day_pattern in day_patterns.items():
                    # Match pattern: "Day: time_ranges"
                    match = re.search(f'({day_pattern})\\s*:\\s*([^;\\n]+)', segment, re.IGNORECASE)
                    if match:
                        day_str = match.group(1).strip().lower()
                        time_str = match.group(2).strip()

                        # Extract all time ranges from time_str
                        time_ranges = []
                        for time_pattern in time_patterns:
                            for time_match in re.finditer(time_pattern, time_str, re.IGNORECASE):
                                start = time_match.group(1)
                                end = time_match.group(2)
                                normalized = ExcelTransformer.normalize_time_range(start, end)
                                time_ranges.append(normalized)

                        # If no time ranges found, store the raw time string
                        if not time_ranges:
                            # Check for special cases like "HALF DAY", "By appointment"
                            if re.search(r'half\s+day|appointment|closed', time_str, re.IGNORECASE):
                                time_ranges.append(time_str)

                        # Join multiple time slots with comma
                        final_time = ','.join(time_ranges) if time_ranges else None

                        if final_time:
                            # Map to result categories
                            if day_key in ['mon_to_fri', 'mon_to_wed', 'thu_to_fri', 'weekdays']:
                                result['weekdays'] = final_time
                                result['metadata']['patterns_found'].append(day_key)
                            elif day_key == 'saturday':
                                result['saturday'] = final_time
                                result['metadata']['patterns_found'].append('saturday')
                            elif day_key == 'sunday':
                                result['sunday'] = final_time
                                result['metadata']['patterns_found'].append('sunday')
                            elif day_key in ['public_holiday', 'eve_of_ph']:
                                result['publicday'] = final_time
                                result['metadata']['patterns_found'].append('publicday')
                            elif day_key == 'day_combo':
                                # Handle combinations like "Sat/Sun" or "Mon/Tue/Fri"
                                if re.search(r'sat', day_str, re.IGNORECASE) and re.search(r'sun', day_str, re.IGNORECASE):
                                    # Both Sat and Sun
                                    result['saturday'] = final_time
                                    result['sunday'] = final_time
                                    result['metadata']['patterns_found'].append('sat_sun_combo')
                                else:
                                    # Weekday combination - map to weekdays
                                    result['weekdays'] = final_time
                                    result['metadata']['patterns_found'].append('weekday_combo')
                            elif day_key in ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']:
                                # Specific weekday - map to weekdays category
                                if not result['weekdays']:  # Only if not already set
                                    result['weekdays'] = final_time
                                    result['metadata']['patterns_found'].append(f'specific_{day_key}')

            # Calculate confidence
            found_count = sum(1 for v in [result['weekdays'], result['saturday'], result['sunday'], result['publicday']] if v is not None)
            result['metadata']['confidence'] = min(1.0, found_count * 0.25)

            if found_count > 0:
                logger.debug(f"Extracted hours from remarks: {result}")

        except Exception as e:
            logger.warning(f"Remarks extraction failed: {e}")
            result['metadata']['reason'] = 'parse_error'
            result['metadata']['error'] = str(e)

        return result

    @staticmethod
    def combine_operating_hours_flexible(df_source, col_map, day_type):
        """Smart operating hours combination supporting both complex (AM/PM/NIGHT) and simple formats with remarks fallback"""
        logger = logging.getLogger(__name__)

        # Define potential column keys for each day type
        if day_type == 'weekday':
            complex_keys = ('mon_fri_am', 'mon_fri_pm', 'mon_fri_night')
            simple_key = 'mon_fri_am'  # Sometimes Mon-Fri is in a single column
        elif day_type == 'saturday':
            complex_keys = ('sat_am', 'sat_pm', 'sat_night')
            simple_key = 'sat_simple'
        elif day_type == 'sunday':
            complex_keys = ('sun_am', 'sun_pm', 'sun_night')
            simple_key = 'sun_simple'
        elif day_type == 'public_holiday':
            complex_keys = ('holiday_am', 'holiday_pm', 'holiday_night')
            simple_key = 'holiday_simple'
        else:
            return ['CLOSED/CLOSED/CLOSED'] * len(df_source)

        # Map day_type to extraction key for remarks fallback
        fallback_map = {
            'weekday': 'weekdays',
            'saturday': 'saturday',
            'sunday': 'sunday',
            'public_holiday': 'publicday'
        }

        result = []
        for idx, row in df_source.iterrows():
            current_result = None

            # Strategy 1: Try complex format (AM/PM/NIGHT columns)
            am_key, pm_key, night_key = complex_keys
            # Only use complex format if we have multiple time periods (AM/PM/Night)
            complex_keys_found = [key for key in complex_keys if key in col_map]
            has_complex = len(complex_keys_found) > 1

            if has_complex:
                # Use complex format
                am = row.get(col_map.get(am_key, ''), 'CLOSED') if am_key in col_map else 'CLOSED'
                pm = row.get(col_map.get(pm_key, ''), 'CLOSED') if pm_key in col_map else 'CLOSED'
                night = row.get(col_map.get(night_key, ''), 'CLOSED') if night_key in col_map else 'CLOSED'

                # Handle NaN values
                am = 'CLOSED' if pd.isna(am) else str(am)
                pm = 'CLOSED' if pd.isna(pm) else str(pm)
                night = 'CLOSED' if pd.isna(night) else str(night)

                current_result = f"{am}/{pm}/{night}"

            else:
                # Strategy 2: Use simple format - check any available key
                hours_str = 'CLOSED'
                found_key = False
                for key_to_check in [simple_key] + list(complex_keys):
                    if key_to_check in col_map:
                        hours_value = row.get(col_map[key_to_check], 'CLOSED')
                        hours_str = 'CLOSED' if pd.isna(hours_value) else str(hours_value)
                        found_key = True
                        break

                if found_key:
                    # For SP clinic format, use clean hours without /CLOSED/CLOSED suffix
                    current_result = hours_str
                else:
                    # Strategy 3: No mapping found, default to CLOSED
                    current_result = 'CLOSED'

            # NEW: Strategy 4 - Fallback to remarks if day column is truly empty
            if ExcelTransformer._is_truly_empty(current_result) and 'remarks' in col_map:
                remarks = row.get(col_map['remarks'], None)
                if pd.notna(remarks):
                    try:
                        extracted = ExcelTransformer.extract_hours_from_remarks(remarks)

                        if day_type in fallback_map:
                            fallback_value = extracted.get(fallback_map[day_type])
                            if fallback_value:
                                current_result = fallback_value
                                logger.debug(f"Row {idx}: Used remarks fallback for {day_type} - extracted: {fallback_value}")
                    except Exception as e:
                        logger.warning(f"Row {idx}: Remarks extraction failed for {day_type}: {e}")

            result.append(current_result)

        return result

    @staticmethod
    def construct_address(df_source, col_map):
        """Smart address construction from multiple columns"""
        addresses = []

        # Define address component priorities
        address_components = [
            ('address_blk', 'Blk'),     # Block number
            ('address_unit', '#'),       # Unit number
            ('address_road', ''),        # Road name
            ('address_building', ''),    # Building name
        ]

        for _, row in df_source.iterrows():
            address_parts = []

            # Check if we have a complete address column first
            if 'address' in col_map and pd.notna(row.get(col_map['address'], '')):
                address_str = str(row[col_map['address']]).strip()
                if address_str and address_str.lower() not in ['nan', '', 'none']:
                    addresses.append(address_str)
                    continue

            # SP Clinic format: Use only Address1 for primary address (Address2/Address3 handled separately)
            if 'address1' in col_map:
                value = row.get(col_map['address1'], '')
                if pd.notna(value) and str(value).strip() and str(value).lower() not in ['nan', '', 'none']:
                    addresses.append(str(value).strip())
                    continue

            # Construct address from components (avoid duplicates)
            used_columns = set()
            for comp_key, prefix in address_components:
                if comp_key in col_map and col_map[comp_key] not in used_columns:
                    used_columns.add(col_map[comp_key])
                    value = row.get(col_map[comp_key], '')
                    if pd.notna(value) and str(value).strip() and str(value).lower() not in ['nan', '', 'none']:
                        value_str = str(value).strip()
                        # For TCM sheets, don't add prefix if the value already contains structural info
                        if comp_key == 'address_blk' and ('blk' in value_str.lower() or 'block' in value_str.lower()):
                            address_parts.append(value_str)
                        elif comp_key == 'address_unit' and ('#' in value_str or 'unit' in value_str.lower()):
                            address_parts.append(value_str)
                        elif prefix and not value_str.startswith(prefix):
                            address_parts.append(f"{prefix} {value_str}")
                        else:
                            address_parts.append(value_str)

            # Combine parts
            if address_parts:
                addresses.append(' '.join(address_parts))
            else:
                addresses.append('')

        return addresses

    @staticmethod
    def smart_column_fallback(df_source, col_map, field_type):
        """Intelligent fallback for missing critical fields"""
        if field_type == 'clinic_id':
            # Try to generate ID from clinic name or use row index
            if 'clinic_name' in col_map:
                return [f"AUTO_{i+1:04d}_{str(name).replace(' ', '_').upper()[:10]}"
                       for i, name in enumerate(df_source[col_map['clinic_name']])]
            else:
                return [f"AUTO_{i+1:04d}" for i in range(len(df_source))]

        elif field_type == 'region':
            # For TCM sheets without region, use 'TCM' as default
            if 'area' in col_map:
                return df_source[col_map['area']].fillna('TCM')
            else:
                return ['TCM'] * len(df_source)

        elif field_type == 'area':
            # For TCM sheets without area, use clinic location or 'SINGAPORE'
            if 'region' in col_map:
                return df_source[col_map['region']].fillna('SINGAPORE')
            else:
                return ['SINGAPORE'] * len(df_source)

        return [''] * len(df_source)

    @staticmethod
    def transform_sheet(input_path, sheet_name, terminated_ids=None, use_google_api=True):
        """Transform a single sheet to target template format with geocoding"""
        try:
            # Initialize geocoding service with user preference
            geocoding_service = GeocodingService(use_google_api=use_google_api)

            # Check if this is Alliance-Tokio Marine format
            # Note: Only .xlsx files support Alliance-Tokio format (uses merged cells)
            import openpyxl
            file_extension = os.path.splitext(input_path)[1].lower()

            if file_extension == '.xlsx':
                # Only check for Alliance-Tokio format in .xlsx files
                wb = openpyxl.load_workbook(input_path)
                ws = wb[sheet_name]
                is_alliance_tokio = ExcelTransformer.detect_alliance_tokio_format(ws)
            else:
                # Legacy .xls files don't support Alliance-Tokio merged cell format
                is_alliance_tokio = False

            if is_alliance_tokio:
                logger.info(f"Processing Alliance-Tokio Marine format for sheet: {sheet_name}")

                # Unmerge cells and fill values
                ws = ExcelTransformer.unmerge_and_fill_cells(ws)

                # Save to temporary file
                import tempfile
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                wb.save(temp_file.name)
                temp_file.close()

                # Read with pandas (data starts at row 3, after headers at rows 1-2)
                df_source = pd.read_excel(temp_file.name, sheet_name=sheet_name, header=None, skiprows=2)

                # Set column names from reconstructed headers
                headers = ExcelTransformer.get_alliance_tokio_headers(ws)
                df_source.columns = headers[:len(df_source.columns)]

                # Clean up temp file
                os.unlink(temp_file.name)

                header_row = None  # Already handled
            else:
                # Standard format - find the correct header row
                header_row = ExcelTransformer.find_header_row(input_path, sheet_name)

                # Read the source sheet
                df_source = ExcelTransformer.safe_read_excel(input_path, sheet_name=sheet_name, header=header_row)
                df_source.columns = df_source.columns.str.strip()
            
            # Create transformed dataframe
            df_transformed = pd.DataFrame()
            
            # Map columns flexibly
            col_map = ExcelTransformer.map_columns(df_source.columns)
            logger.debug(f"Column mapping for sheet '{sheet_name}': {col_map}")

            # If column mapping failed, try to infer from data structure
            if 'clinic_name' not in col_map:
                logger.info(f"Standard column mapping failed for '{sheet_name}', attempting inference...")
                inferred_map = ExcelTransformer.infer_columns_from_data(df_source)
                if 'clinic_name' in inferred_map:
                    col_map.update(inferred_map)
                    logger.info(f"Successfully inferred columns for sheet '{sheet_name}'")
                else:
                    raise ValueError(f"Sheet '{sheet_name}' missing required 'clinic name' column after inference attempt")

            # Filter out metadata rows for Alliance-Tokio format
            if is_alliance_tokio:
                initial_count = len(df_source)
                # Filter out rows where ZONE contains metadata keywords
                metadata_keywords = ['Legend:', 'Remarks', '24 Hours', 'Newly joined', 'JB Clinics', 'clinic']
                zone_col = 'ZONE'
                if zone_col in df_source.columns:
                    metadata_mask = df_source[zone_col].astype(str).str.contains('|'.join(metadata_keywords), case=False, na=False)
                    df_source = df_source[~metadata_mask]
                    filtered_metadata = initial_count - len(df_source)
                    if filtered_metadata > 0:
                        logger.info(f"Filtered out {filtered_metadata} metadata rows from Alliance-Tokio sheet")

            # NOTE: Termination filtering will be done AFTER postal code extraction
            # This ensures we can match both provider code AND postal code
            terminated_count = 0
            filtered_provider_codes = []

            # Filter out empty/invalid rows - keep only rows with valid clinic data
            initial_count = len(df_source)
            if 'clinic_name' in col_map:
                # Filter by clinic_name (essential identifier)
                # Provider code is optional - some valid clinics may have missing codes
                valid_mask = df_source[col_map['clinic_name']].notna() & (df_source[col_map['clinic_name']].astype(str).str.strip() != '')
            elif 'clinic_id' in col_map:
                # Fallback: if no clinic_name, use clinic_id
                valid_mask = df_source[col_map['clinic_id']].notna() & (df_source[col_map['clinic_id']].astype(str).str.strip() != '')
            else:
                # Should not happen, but keep all rows if no clinic identifier found
                valid_mask = pd.Series([True] * len(df_source))

            df_source = df_source[valid_mask]
            filtered_count = len(df_source)
            if initial_count != filtered_count:
                logger.info(f"Filtered out {initial_count - filtered_count} empty/invalid rows from sheet '{sheet_name}', keeping {filtered_count} valid records")

            # Robust field mapping with fallbacks
            # Clinic ID with smart fallback
            if 'clinic_id' in col_map:
                df_transformed['Code'] = df_source[col_map['clinic_id']]
            else:
                df_transformed['Code'] = ExcelTransformer.smart_column_fallback(df_source, col_map, 'clinic_id')
                logger.info(f"Generated auto clinic IDs for {len(df_source)} records")

            # Clinic Name (required field)
            df_transformed['Name'] = df_source[col_map['clinic_name']]

            # Region with smart fallback
            if 'region' in col_map:
                df_transformed['Zone'] = df_source[col_map['region']]
            else:
                df_transformed['Zone'] = ExcelTransformer.smart_column_fallback(df_source, col_map, 'region')

            # Area with smart fallback
            if 'area' in col_map:
                df_transformed['Area'] = df_source[col_map['area']]
            else:
                df_transformed['Area'] = ExcelTransformer.smart_column_fallback(df_source, col_map, 'area')

            # Specialty field (available in SP clinic sheets)
            if 'specialty' in col_map:
                df_transformed['Specialty'] = df_source[col_map['specialty']]
            else:
                df_transformed['Specialty'] = None

            # Doctor field (available in TCM and SP clinic sheets)
            if 'doctor_name' in col_map:
                df_transformed['Doctor'] = df_source[col_map['doctor_name']]
            else:
                df_transformed['Doctor'] = None

            # Smart address construction
            df_transformed['Address1'] = ExcelTransformer.construct_address(df_source, col_map)

            # Extract Address2 and Address3 from source data if available
            if 'address2' in col_map:
                df_transformed['Address2'] = df_source[col_map['address2']].fillna('')
            else:
                df_transformed['Address2'] = None

            if 'address3' in col_map:
                df_transformed['Address3'] = df_source[col_map['address3']].fillna('')
            else:
                df_transformed['Address3'] = None

            # Extract postal codes from addresses (supports both Singapore and Malaysia)
            logger.info("=" * 60)
            logger.info("POSTAL CODE EXTRACTION STARTED")
            logger.info(f"Total records: {len(df_transformed)}")
            logger.info(f"Postal code column mapped: {'postal_code' in col_map}")
            if 'postal_code' in col_map:
                logger.info(f"  Source column: {col_map['postal_code']}")
            logger.info("=" * 60)

            # Enhanced postal code extraction
            postal_codes = []
            extraction_methods = {'dedicated_column': 0, 'address4': 0, 'address1': 0, 'failed': 0}

            for index, address in enumerate(df_transformed['Address1']):
                postal_code = None
                extraction_method = None

                # Try dedicated postal code column first (if it has valid data)
                if 'postal_code' in col_map:
                    postal_col_value = df_source.iloc[index][col_map['postal_code']]
                    if pd.notna(postal_col_value) and str(postal_col_value).strip() not in ('', 'nan', 'None'):
                        postal_code = str(postal_col_value).strip()
                        extraction_method = 'dedicated_column'

                # If no valid postal code from dedicated column, extract from address
                if not postal_code:
                    # For SP clinic format, check Address4 first (contains "SINGAPORE 247909")
                    if 'address4' in col_map:
                        address4_value = df_source.iloc[index].get(col_map['address4'], '')
                        if pd.notna(address4_value) and str(address4_value).strip():
                            postal_code = ExcelTransformer.extract_postal_code(str(address4_value))
                            if postal_code:
                                extraction_method = 'address4'

                    # Fallback to extracting from combined address
                    if not postal_code and pd.notna(address) and str(address).strip():
                        postal_code = ExcelTransformer.extract_postal_code(address)
                        if postal_code:
                            extraction_method = 'address1'

                if extraction_method:
                    extraction_methods[extraction_method] += 1
                else:
                    extraction_methods['failed'] += 1

                postal_codes.append(postal_code)

            df_transformed['PostalCode'] = postal_codes

            # Log extraction results
            logger.info("=" * 60)
            logger.info("POSTAL CODE EXTRACTION COMPLETE")
            logger.info(f"Extraction results:")
            logger.info(f"  From dedicated column: {extraction_methods['dedicated_column']}")
            logger.info(f"  From Address4 field: {extraction_methods['address4']}")
            logger.info(f"  From Address1 field: {extraction_methods['address1']}")
            logger.info(f"  Failed extractions: {extraction_methods['failed']}")
            total_extracted = extraction_methods['dedicated_column'] + extraction_methods['address4'] + extraction_methods['address1']
            logger.info(f"Success rate: {(total_extracted/len(df_transformed)*100):.1f}%" if len(df_transformed) > 0 else "N/A")
            logger.info("=" * 60)

            # Filter out terminated clinics using dual-parameter matching (provider code + postal code)
            if terminated_ids and 'clinic_id' in col_map:
                initial_count = len(df_transformed)
                clinic_id_col = col_map['clinic_id']

                # Create termination matching mask
                terminated_mask = []
                for idx, row in df_transformed.iterrows():
                    # Normalize provider code and postal code for consistent matching
                    provider_code = ExcelTransformer.normalize_code(df_source.loc[idx][clinic_id_col])
                    postal_code = ExcelTransformer.normalize_code(row['PostalCode'])

                    is_terminated = False
                    if provider_code and postal_code:
                        # Check for exact match (provider_code, postal_code)
                        if (provider_code, postal_code) in terminated_ids:
                            is_terminated = True
                            if provider_code not in filtered_provider_codes:
                                filtered_provider_codes.append(provider_code)
                        # Check for fallback single-parameter match (provider_code, None)
                        elif (provider_code, None) in terminated_ids:
                            is_terminated = True
                            if provider_code not in filtered_provider_codes:
                                filtered_provider_codes.append(provider_code)

                    terminated_mask.append(is_terminated)

                # Apply filter to both dataframes
                df_transformed = df_transformed[~pd.Series(terminated_mask, index=df_transformed.index)]
                df_source = df_source[~pd.Series(terminated_mask, index=df_source.index)]

                # Reset indices
                df_transformed = df_transformed.reset_index(drop=True)
                df_source = df_source.reset_index(drop=True)

                terminated_count = initial_count - len(df_transformed)

                logger.info("=" * 60)
                logger.info(f"TERMINATION FILTERING COMPLETE")
                logger.info(f"Filtered out {terminated_count} terminated clinics from sheet '{sheet_name}'")
                logger.info(f"Matching criteria: Provider Code + Postal Code (dual-parameter)")
                if filtered_provider_codes:
                    logger.info(f"Filtered provider codes: {', '.join(filtered_provider_codes[:10])}" +
                               (f" ... and {len(filtered_provider_codes)-10} more" if len(filtered_provider_codes) > 10 else ""))
                logger.info("=" * 60)

            # Detect country from address information
            def detect_country(address):
                if pd.isna(address) or str(address).strip() == '':
                    return 'SINGAPORE'  # Default

                address_lower = str(address).lower()

                # PRIORITY 1: Check for explicit "SINGAPORE" keyword first
                # This prevents false positives like "Penang Road, Singapore" being detected as Malaysia
                if 'singapore' in address_lower:
                    return 'SINGAPORE'

                # PRIORITY 2: Check for Malaysian indicators using word boundaries
                # Multi-word phrases (check first to avoid partial matches)
                multi_word_indicators = [
                    'kuala lumpur', 'johor bahru', 'negeri sembilan',
                    'shah alam', 'petaling jaya', 'johor darul', 'iskandar puteri',
                    'taman daya', 'bandar indahpura', 'ulu tiram'
                ]

                for indicator in multi_word_indicators:
                    if indicator in address_lower:
                        return 'MALAYSIA'

                # Single word indicators (use word boundaries to avoid false positives)
                import re
                single_word_indicators = [
                    'malaysia', 'johor', 'selangor', 'penang', 'perak',
                    'kedah', 'kelantan', 'terengganu', 'pahang',
                    'melaka', 'sabah', 'sarawak', 'perlis', 'putrajaya', 'labuan',
                    # Additional Johor cities/towns
                    'kulai', 'skudai', 'senai', 'pasir gudang', 'pontian',
                    'batu pahat', 'muar', 'segamat', 'kluang', 'kota tinggi'
                ]

                for indicator in single_word_indicators:
                    # Use word boundary regex to match whole words only
                    if re.search(r'\b' + re.escape(indicator) + r'\b', address_lower):
                        return 'MALAYSIA'

                # Special case: "kl" should only match as standalone abbreviation
                # Check for "kl" with word boundaries OR preceded/followed by comma, space, or end
                if re.search(r'\bkl\b', address_lower):
                    return 'MALAYSIA'

                return 'SINGAPORE'

            # Detect country from combined address fields (Address1, Address2, Address3, PostalCode, Zone, Region, Area)
            def detect_country_from_row(row):
                # Combine all address-related fields AND region/zone/area to check for country indicators
                # This catches cases where "JOHOR" is in Zone/Region but not in the address itself
                combined_address = ' '.join([
                    str(row.get('Zone', '')),
                    str(row.get('Region', '')),
                    str(row.get('Area', '')),
                    str(row.get('Address1', '')),
                    str(row.get('Address2', '')),
                    str(row.get('Address3', '')),
                    str(row.get('PostalCode', ''))
                ])
                return detect_country(combined_address)

            df_transformed['Country'] = df_transformed.apply(detect_country_from_row, axis=1)

            # Combine phone and remarks (if available)
            if 'telephone' in col_map and col_map['telephone'] is not None and pd.notna(col_map['telephone']):
                if 'remarks' in col_map and col_map['remarks'] is not None:
                    df_transformed['PhoneNumber'] = df_source.apply(
                        lambda row: ExcelTransformer.combine_phone_remarks(
                            row[col_map['telephone']], row[col_map['remarks']]
                        ), axis=1
                    )
                else:
                    df_transformed['PhoneNumber'] = df_source[col_map['telephone']].astype(str)
            else:
                df_transformed['PhoneNumber'] = ''

            # Operating hours handling
            if is_alliance_tokio:
                # Alliance-Tokio format: Convert 4-column format to AM/PM/NIGHT
                logger.info("Converting Alliance-Tokio operating hours to standard AM/PM/NIGHT format")

                weekdays = []
                saturdays = []
                sundays = []
                holidays = []

                for _, row in df_source.iterrows():
                    mon_fri = row.get('MON - FRI', '')
                    sat = row.get('SAT', '')
                    sun = row.get('SUN', '')
                    ph = row.get('PUBLIC HOLIDAYS', '')

                    wd, sat_result, sun_result, hol_result = ExcelTransformer.convert_alliance_hours_to_standard(
                        mon_fri, sat, sun, ph
                    )

                    weekdays.append(wd)
                    saturdays.append(sat_result)
                    sundays.append(sun_result)
                    holidays.append(hol_result)

                df_transformed['MonToFri'] = weekdays
                df_transformed['Saturday'] = saturdays
                df_transformed['Sunday'] = sundays
                df_transformed['PublicHoliday'] = holidays
            else:
                # Standard format: flexible mapping
                df_transformed['MonToFri'] = ExcelTransformer.combine_operating_hours_flexible(df_source, col_map, 'weekday')
                df_transformed['Saturday'] = ExcelTransformer.combine_operating_hours_flexible(df_source, col_map, 'saturday')
                df_transformed['Sunday'] = ExcelTransformer.combine_operating_hours_flexible(df_source, col_map, 'sunday')
                df_transformed['PublicHoliday'] = ExcelTransformer.combine_operating_hours_flexible(df_source, col_map, 'public_holiday')
            
            # Geocoding: Populate Latitude and Longitude
            logger.info("=" * 60)
            logger.info(f"GEOCODING PROCESS STARTED - Sheet: {sheet_name}")
            logger.info(f"Total records to geocode: {len(df_transformed)}")
            logger.info("=" * 60)

            latitudes = []
            longitudes = []
            geocoding_methods = []

            # Log sample of postal codes being processed
            sample_size = min(3, len(df_transformed))
            if sample_size > 0:
                logger.info(f"Sample postal codes extracted:")
                for i in range(sample_size):
                    postal = df_transformed.iloc[i]['PostalCode']
                    addr = df_transformed.iloc[i]['Address1']
                    logger.info(f"  Row {i+1}: PostalCode='{postal}', Address='{addr[:50]}...' " if len(str(addr)) > 50 else f"  Row {i+1}: PostalCode='{postal}', Address='{addr}'")

            for index, row in df_transformed.iterrows():
                postal_code = row['PostalCode']
                address = row['Address1']
                country = row['Country']  # Get country to force region bias for Malaysia

                # Pass country to geocode method to force Malaysia region for Malaysian addresses
                lat, lng, method = geocoding_service.geocode(postal_code, address, country=country)
                latitudes.append(lat)
                longitudes.append(lng)
                geocoding_methods.append(method)

                # Log progress every 50 records
                if (index + 1) % 50 == 0:
                    success_count = sum(1 for l in latitudes if l is not None)
                    logger.info(f"Geocoding progress: {index + 1}/{len(df_transformed)} records ({success_count} successful)")

            df_transformed['Latitude'] = latitudes
            df_transformed['Longitude'] = longitudes
            
            # Generate Google Maps URLs for successfully geocoded locations
            def generate_google_maps_url(lat, lng):
                if lat is not None and lng is not None:
                    return f"https://maps.google.com/?q={lat},{lng}"
                return None
            
            df_transformed['GoogleMapURL'] = df_transformed.apply(
                lambda row: generate_google_maps_url(row['Latitude'], row['Longitude']), axis=1
            )
            
            # Return the transformed dataframe instead of saving
            # Saving will be handled by the multi-sheet processor
            
            # Get geocoding statistics
            stats = geocoding_service.get_stats()
            successful_geocodes = sum(1 for lat in latitudes if lat is not None)
            postal_matches = len([m for m in geocoding_methods if m == 'postal_code'])
            address_matches = len([m for m in geocoding_methods if m == 'address'])
            failed_geocodes = len(df_transformed) - successful_geocodes
            success_rate = (successful_geocodes/len(df_transformed)*100) if len(df_transformed) > 0 else 0

            # Log final geocoding statistics
            logger.info("=" * 60)
            logger.info(f"GEOCODING COMPLETE - Sheet: {sheet_name}")
            logger.info(f"Total records processed: {len(df_transformed)}")
            logger.info(f"Successful geocodes: {successful_geocodes} ({success_rate:.1f}%)")
            logger.info(f"  - Via postal code lookup: {postal_matches}")
            logger.info(f"  - Via Google Maps API: {address_matches}")
            logger.info(f"Failed geocodes: {failed_geocodes}")

            if failed_geocodes > 0:
                # Log sample of failed postal codes for debugging
                failed_samples = []
                for idx, (lat, postal, addr) in enumerate(zip(latitudes, df_transformed['PostalCode'], df_transformed['Address1'])):
                    if lat is None and len(failed_samples) < 3:
                        failed_samples.append((postal, str(addr)[:40]))
                if failed_samples:
                    logger.warning(f"Sample failed postal codes:")
                    for postal, addr in failed_samples:
                        logger.warning(f"  PostalCode='{postal}', Address='{addr}...'")

            logger.info("=" * 60)

            return {
                'success': True,
                'dataframe': df_transformed,
                'message': f'Successfully transformed {len(df_transformed)} records',
                'records_processed': len(df_transformed),
                'terminated_clinics_filtered': terminated_count,
                'filtered_provider_codes': filtered_provider_codes,
                'geocoding_stats': {
                    'total_records': len(df_transformed),
                    'successful_geocodes': successful_geocodes,
                    'postal_code_matches': postal_matches,
                    'address_geocodes': address_matches,
                    'failed_geocodes': failed_geocodes,
                    'success_rate': f"{success_rate:.1f}%"
                }
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f'Error transforming sheet {sheet_name}: {str(e)}',
                'error_details': traceback.format_exc()
            }

    @staticmethod
    def transform_excel_multi_sheet(input_path, output_dir, job_id, use_google_api=True):
        """Transform Excel file with multiple sheets to multiple output files"""
        try:
            # Get all sheet names (with fallback for corrupted XML)
            try:
                xl_file = pd.ExcelFile(input_path)
                sheet_names = xl_file.sheet_names
            except ValueError as e:
                if "could not assign names" in str(e) or "invalid XML" in str(e):
                    logger.warning(f"Excel file has corrupted metadata, using patched openpyxl to read sheet names")
                    import openpyxl
                    from openpyxl.reader.workbook import WorkbookParser

                    # Save original method
                    original_assign_names = WorkbookParser.assign_names

                    def patched_assign_names(self):
                        """Patched version that skips invalid print titles"""
                        try:
                            original_assign_names(self)
                        except ValueError as ve:
                            if "not a valid print titles definition" in str(ve):
                                logger.warning(f"Skipping invalid print title: {ve}")
                                pass
                            else:
                                raise

                    # Apply patch
                    WorkbookParser.assign_names = patched_assign_names

                    try:
                        xl_file = pd.ExcelFile(input_path)
                        sheet_names = xl_file.sheet_names
                    finally:
                        # Restore original method
                        WorkbookParser.assign_names = original_assign_names
                else:
                    raise

            # Classify sheets
            panel_sheets, termination_sheets = ExcelTransformer.classify_sheets(sheet_names)

            logger.info(f"Detected {len(panel_sheets)} panel sheets: {panel_sheets}")
            logger.info(f"Detected {len(termination_sheets)} termination sheets: {termination_sheets}")

            # Extract terminated clinic IDs
            terminated_ids = ExcelTransformer.extract_terminated_clinic_ids(input_path, termination_sheets)

            # Process each panel sheet
            results = []
            output_files = []

            for sheet in panel_sheets:
                logger.info(f"Processing sheet: {sheet}")

                # Transform the sheet with geocoding preference
                result = ExcelTransformer.transform_sheet(input_path, sheet, terminated_ids, use_google_api)

                if result['success']:
                    df = result['dataframe']

                    # Check if this sheet has mixed Singapore/Malaysia data
                    has_country_column = 'Country' in df.columns
                    has_mixed_countries = False

                    if has_country_column:
                        countries = df['Country'].unique()
                        has_mixed_countries = len([c for c in countries if c in ['SINGAPORE', 'MALAYSIA']]) > 1
                        logger.info(f"Sheet '{sheet}' countries detected: {list(countries)}")

                    if has_mixed_countries:
                        # Separate Singapore and Malaysia clinics
                        df_sg = df[df['Country'] == 'SINGAPORE'].copy()
                        df_my = df[df['Country'] == 'MALAYSIA'].copy()

                        logger.info(f"Separating sheet '{sheet}': {len(df_sg)} Singapore clinics, {len(df_my)} Malaysia clinics")

                        # For mixed country sheets, use clean country names as display names
                        # This prevents "SINGAPORE (Malaysia)" or "MALAYSIA (Singapore)" confusion
                        sheet_upper = sheet.upper()

                        # Save Singapore file
                        if len(df_sg) > 0:
                            sanitized_name = ExcelTransformer.sanitize_filename(sheet)
                            sg_filename = f"{job_id}_{sanitized_name}_Singapore.xlsx"
                            sg_path = os.path.join(output_dir, sg_filename)
                            df_sg = ExcelTransformer.format_postal_codes(df_sg)
                            ExcelTransformer.write_excel_with_text_postal_codes(df_sg, sg_path)

                            # Use "SINGAPORE" for Singapore data, regardless of original sheet name
                            sg_display_name = "SINGAPORE"

                            results.append({
                                'sheet_name': sg_display_name,
                                'output_filename': sg_filename,
                                'output_path': sg_path,
                                'records_processed': len(df_sg),
                                'terminated_clinics_filtered': result['terminated_clinics_filtered'],
                                'filtered_provider_codes': result.get('filtered_provider_codes', []),
                                'geocoding_stats': {
                                    'total_records': len(df_sg),
                                    'successful_geocodes': int(df_sg['Latitude'].notna().sum()),
                                    'success_rate': f"{(df_sg['Latitude'].notna().sum()/len(df_sg)*100):.1f}%"
                                }
                            })
                            output_files.append(sg_filename)

                        # Save Malaysia file
                        if len(df_my) > 0:
                            sanitized_name = ExcelTransformer.sanitize_filename(sheet)
                            my_filename = f"{job_id}_{sanitized_name}_Malaysia.xlsx"
                            my_path = os.path.join(output_dir, my_filename)
                            df_my = ExcelTransformer.format_postal_codes(df_my)
                            ExcelTransformer.write_excel_with_text_postal_codes(df_my, my_path)

                            # Use "MALAYSIA" for Malaysia data, regardless of original sheet name
                            my_display_name = "MALAYSIA"

                            results.append({
                                'sheet_name': my_display_name,
                                'output_filename': my_filename,
                                'output_path': my_path,
                                'records_processed': len(df_my),
                                'terminated_clinics_filtered': 0,
                                'filtered_provider_codes': [],
                                'geocoding_stats': {
                                    'total_records': len(df_my),
                                    'successful_geocodes': int(df_my['Latitude'].notna().sum()),
                                    'success_rate': f"{(df_my['Latitude'].notna().sum()/len(df_my)*100):.1f}%"
                                }
                            })
                            output_files.append(my_filename)

                        logger.info(f"SUCCESS: Separated sheet '{sheet}' into Singapore ({len(df_sg)} records) and Malaysia ({len(df_my)} records)")
                    else:
                        # Single country or no country column - save as one file
                        sanitized_name = ExcelTransformer.sanitize_filename(sheet)
                        output_filename = f"{job_id}_{sanitized_name}.xlsx"
                        output_path = os.path.join(output_dir, output_filename)

                        # Save the transformed dataframe
                        df = ExcelTransformer.format_postal_codes(df)
                        ExcelTransformer.write_excel_with_text_postal_codes(df, output_path)

                        # Store result info
                        sheet_result = {
                            'sheet_name': sheet,
                            'output_filename': output_filename,
                            'output_path': output_path,
                            'records_processed': result['records_processed'],
                            'terminated_clinics_filtered': result['terminated_clinics_filtered'],
                            'filtered_provider_codes': result.get('filtered_provider_codes', []),
                            'geocoding_stats': result['geocoding_stats']
                        }
                        results.append(sheet_result)
                        output_files.append(output_filename)

                        logger.info(f"SUCCESS: Processed sheet '{sheet}': {result['records_processed']} records")
                else:
                    logger.error(f"FAILED: Could not process sheet '{sheet}': {result['message']}")
                    return {
                        'success': False,
                        'message': f"Failed to process sheet '{sheet}': {result['message']}",
                        'error_details': result.get('error_details', '')
                    }

            if not results:
                return {
                    'success': False,
                    'message': 'No panel sheets found to process. Expected sheets with names containing: GP, TCM, dental, clinic, or panel.'
                }

            # Calculate summary statistics
            total_records = sum(r['records_processed'] for r in results)
            total_geocodes = sum(r['geocoding_stats']['successful_geocodes'] for r in results)
            total_terminated = sum(r.get('terminated_clinics_filtered', 0) for r in results)

            # Aggregate all filtered provider codes from all sheets
            all_filtered_codes = []
            for r in results:
                all_filtered_codes.extend(r.get('filtered_provider_codes', []))
            # Remove duplicates while preserving order
            unique_filtered_codes = list(dict.fromkeys(all_filtered_codes))

            return {
                'success': True,
                'message': f'Successfully processed {len(results)} sheets with {total_records} total records',
                'sheets_processed': len(results),
                'total_records': total_records,
                'terminated_clinics_filtered': total_terminated,
                'filtered_provider_codes': unique_filtered_codes,
                'output_files': output_files,
                'results': results,
                'summary_stats': {
                    'total_successful_geocodes': total_geocodes,
                    'overall_geocoding_rate': f"{(total_geocodes/total_records*100):.1f}%" if total_records > 0 else "0%"
                }
            }

        except Exception as e:
            return {
                'success': False,
                'message': f'Error processing multi-sheet file: {str(e)}',
                'error_details': traceback.format_exc()
            }

class BatchProcessor:
    def __init__(self):
        self.batch_jobs = {}  # Store batch job status
        self.lock = threading.Lock()

    def create_batch_job(self, batch_id, file_count):
        """Create a new batch processing job"""
        with self.lock:
            self.batch_jobs[batch_id] = {
                'status': 'processing',
                'total_files': file_count,
                'completed_files': 0,
                'failed_files': 0,
                'results': [],
                'created_at': datetime.now().isoformat()
            }

    def update_batch_progress(self, batch_id, file_result):
        """Update progress for a batch job"""
        with self.lock:
            if batch_id in self.batch_jobs:
                self.batch_jobs[batch_id]['results'].append(file_result)
                if file_result.get('success', False):
                    self.batch_jobs[batch_id]['completed_files'] += 1
                else:
                    self.batch_jobs[batch_id]['failed_files'] += 1

                # Check if batch is complete
                total_processed = self.batch_jobs[batch_id]['completed_files'] + self.batch_jobs[batch_id]['failed_files']
                if total_processed >= self.batch_jobs[batch_id]['total_files']:
                    self.batch_jobs[batch_id]['status'] = 'completed'

    def get_batch_status(self, batch_id):
        """Get status of a batch job"""
        with self.lock:
            return self.batch_jobs.get(batch_id, None)

# Global batch processor instance
batch_processor = BatchProcessor()

def process_single_file_in_batch(file_data, batch_id, use_google_api=True):
    """Process a single file as part of a batch job"""
    try:
        file_content = file_data['content']
        original_filename = file_data['filename']
        job_id = str(uuid.uuid4())

        # Save uploaded file with original extension
        original_ext = os.path.splitext(original_filename)[1] or '.xlsx'  # Default to .xlsx if no extension
        input_filename = f"{job_id}_input{original_ext}"
        input_path = os.path.join(UPLOAD_FOLDER, input_filename)

        with open(input_path, 'wb') as f:
            f.write(file_content)

        # Validate file content (using safe method that handles corrupted XML)
        try:
            # Use the same safe reading approach as transform_excel_multi_sheet
            try:
                xl_file = pd.ExcelFile(input_path)
                sheet_names = xl_file.sheet_names
            except ValueError as e:
                if "could not assign names" in str(e) or "invalid XML" in str(e):
                    logger.warning(f"Batch file has corrupted metadata, using patched validation for {original_filename}")
                    import openpyxl
                    from openpyxl.reader.workbook import WorkbookParser

                    # Save original method
                    original_assign_names = WorkbookParser.assign_names

                    def patched_assign_names(self):
                        """Patched version that skips invalid print titles"""
                        try:
                            original_assign_names(self)
                        except ValueError as ve:
                            if "not a valid print titles definition" in str(ve):
                                logger.warning(f"Skipping invalid print title during validation: {ve}")
                                pass
                            else:
                                raise

                    # Apply patch
                    WorkbookParser.assign_names = patched_assign_names

                    try:
                        xl_file = pd.ExcelFile(input_path)
                        sheet_names = xl_file.sheet_names
                    finally:
                        # Restore original method
                        WorkbookParser.assign_names = original_assign_names
                else:
                    raise
            logger.info(f"Batch file validation passed for {original_filename} (job {job_id})")
        except Exception as validation_error:
            if os.path.exists(input_path):
                os.remove(input_path)
            raise ValueError(f"Invalid Excel file: {original_filename}")

        # Transform file with geocoding preference
        result = ExcelTransformer.transform_excel_multi_sheet(input_path, PROCESSED_FOLDER, job_id, use_google_api)

        # Clean up input file
        if os.path.exists(input_path):
            os.remove(input_path)

        if result['success']:
            file_result = {
                'success': True,
                'filename': original_filename,
                'job_id': job_id,
                'sheets_processed': result['sheets_processed'],
                'total_records': result['total_records'],
                'terminated_clinics_filtered': result['terminated_clinics_filtered'],
                'filtered_provider_codes': result.get('filtered_provider_codes', []),
                'output_files': result['output_files'],
                'download_urls': [f'/download/{job_id}/{filename}' for filename in result['output_files']],
                'results': result['results'],  # Individual sheet results
                'summary_stats': result['summary_stats']
            }
        else:
            file_result = {
                'success': False,
                'filename': original_filename,
                'error': result['message'],
                'details': result.get('error_details', '')
            }

        # Update batch progress
        batch_processor.update_batch_progress(batch_id, file_result)

        return file_result

    except Exception as e:
        error_result = {
            'success': False,
            'filename': file_data.get('filename', 'unknown'),
            'error': str(e),
            'details': traceback.format_exc()
        }
        batch_processor.update_batch_progress(batch_id, error_result)
        return error_result

@app.route('/upload/batch', methods=['POST'])
def upload_batch():
    """Handle batch upload of multiple Excel files"""
    try:
        # Check if files are provided
        if 'files' not in request.files:
            return jsonify({'error': 'No files provided'}), 400

        files = request.files.getlist('files')
        if not files or len(files) == 0:
            return jsonify({'error': 'No files selected'}), 400

        # Extract geocoding preference (default: True)
        use_google_api = request.form.get('use_google_api', 'true').lower() == 'true'
        logger.info(f"Batch upload - Google Maps API geocoding: {'ENABLED' if use_google_api else 'DISABLED'}")

        # Validate file count (limit to 10 files per batch)
        MAX_BATCH_SIZE = 10
        if len(files) > MAX_BATCH_SIZE:
            return jsonify({'error': f'Too many files. Maximum {MAX_BATCH_SIZE} files allowed per batch.'}), 400

        # Validate each file
        file_data_list = []
        for file in files:
            if file.filename == '':
                continue

            if not file.filename.lower().endswith(('.xlsx', '.xls')):
                return jsonify({'error': f'Invalid file format: {file.filename}. Please upload Excel files only.'}), 400

            # File size validation (50MB limit per file)
            file.seek(0, 2)  # Seek to end
            file_size = file.tell()
            file.seek(0)  # Seek back to beginning

            MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
            if file_size > MAX_FILE_SIZE:
                return jsonify({'error': f'File too large: {file.filename}. Maximum size is 50MB per file.'}), 413

            # Read file content
            file_content = file.read()
            file_data_list.append({
                'content': file_content,
                'filename': file.filename,
                'size': file_size
            })

        if not file_data_list:
            return jsonify({'error': 'No valid files provided'}), 400

        # Generate batch ID
        batch_id = str(uuid.uuid4())

        # Create batch job
        batch_processor.create_batch_job(batch_id, len(file_data_list))

        # Process files - use concurrent processing if available, otherwise sequential
        if CONCURRENT_SUPPORT and len(file_data_list) > 1:
            # Concurrent processing for better performance
            max_workers = min(len(file_data_list), 2)  # Reduced from 3 to 2 for free tier
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Submit all files for processing with geocoding preference
                future_to_file = {
                    executor.submit(process_single_file_in_batch, file_data, batch_id, use_google_api): file_data['filename']
                    for file_data in file_data_list
                }

                # Collect results as they complete
                results = []
                for future in as_completed(future_to_file):
                    try:
                        result = future.result(timeout=300)  # 5 minute timeout per file
                        results.append(result)
                    except Exception as e:
                        filename = future_to_file[future]
                        error_result = {
                            'success': False,
                            'filename': filename,
                            'error': f'Processing timeout or error: {str(e)}'
                        }
                        results.append(error_result)
                        batch_processor.update_batch_progress(batch_id, error_result)
        else:
            # Sequential processing fallback
            results = []
            for file_data in file_data_list:
                try:
                    result = process_single_file_in_batch(file_data, batch_id, use_google_api)
                    results.append(result)
                except Exception as e:
                    error_result = {
                        'success': False,
                        'filename': file_data['filename'],
                        'error': f'Processing error: {str(e)}'
                    }
                    results.append(error_result)
                    batch_processor.update_batch_progress(batch_id, error_result)

        # Get final batch status
        batch_status = batch_processor.get_batch_status(batch_id)

        successful_files = [r for r in results if r.get('success', False)]
        failed_files = [r for r in results if not r.get('success', False)]

        return jsonify({
            'batch_id': batch_id,
            'message': f'Batch processing completed. {len(successful_files)} files processed successfully, {len(failed_files)} failed.',
            'total_files': len(file_data_list),
            'successful_files': len(successful_files),
            'failed_files': len(failed_files),
            'results': results,
            'batch_status': batch_status
        })

    except Exception as e:
        logger.error(f"Batch upload error: {str(e)}")
        return jsonify({
            'error': 'Internal server error during batch processing',
            'details': str(e)
        }), 500

@app.route('/batch/status/<batch_id>', methods=['GET'])
def get_batch_status(batch_id):
    """Get status of a batch processing job"""
    try:
        status = batch_processor.get_batch_status(batch_id)
        if not status:
            return jsonify({'error': 'Batch job not found'}), 404

        return jsonify(status)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/batch/download/<batch_id>', methods=['GET'])
def download_batch(batch_id):
    """Download all files from a batch processing job as ZIP"""
    try:
        batch_status = batch_processor.get_batch_status(batch_id)
        if not batch_status:
            return jsonify({'error': 'Batch job not found'}), 404

        if batch_status['status'] != 'completed':
            return jsonify({'error': 'Batch processing not completed yet'}), 400

        # Collect all successful results
        successful_results = [r for r in batch_status['results'] if r.get('success', False)]

        if not successful_results:
            return jsonify({'error': 'No successful files to download'}), 404

        import zipfile
        import tempfile
        import glob

        # Create temporary zip file
        temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')

        with zipfile.ZipFile(temp_zip.name, 'w') as zipf:
            for result in successful_results:
                job_id = result['job_id']
                original_filename = result['filename']

                # Find all files for this job
                pattern = os.path.join(PROCESSED_FOLDER, f"{job_id}_*.xlsx")
                matching_files = glob.glob(pattern)

                for file_path in matching_files:
                    filename = os.path.basename(file_path)
                    # Create descriptive archive names: originalname_sheetname.xlsx
                    sheet_part = filename.replace(f"{job_id}_", "").replace(".xlsx", "")
                    base_name = os.path.splitext(original_filename)[0]
                    arc_name = f"{base_name}_{sheet_part}.xlsx"
                    zipf.write(file_path, arc_name)

        return send_file(
            temp_zip.name,
            as_attachment=True,
            download_name=f'batch_{batch_id}_results.zip',
            mimetype='application/zip'
        )

    except Exception as e:
        logger.error(f"Batch download error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/geocode', methods=['POST'])
def geocode_address():
    """Standalone geocoding endpoint for testing"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No JSON data provided'}), 400
        
        postal_code = data.get('postal_code')
        address = data.get('address')
        country = data.get('country')  # Optional: 'SINGAPORE' or 'MALAYSIA' to force region bias

        if not postal_code and not address:
            return jsonify({'error': 'Either postal_code or address must be provided'}), 400

        geocoding_service = GeocodingService()
        lat, lng, method = geocoding_service.geocode(postal_code, address, country=country)
        
        if lat is not None and lng is not None:
            google_maps_url = f"https://maps.google.com/?q={lat},{lng}"
            return jsonify({
                'success': True,
                'latitude': lat,
                'longitude': lng,
                'method': method,
                'google_maps_url': google_maps_url,
                'postal_code': postal_code,
                'address': address
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Could not geocode the provided postal code or address',
                'postal_code': postal_code,
                'address': address
            }), 404
            
    except Exception as e:
        return jsonify({
            'error': 'Internal server error',
            'details': str(e)
        }), 500

@app.route('/health', methods=['GET'])
def health_check():
    """Enhanced health check with deployment diagnostics"""
    try:
        start_time = datetime.now()

        # Quick health check first
        health_info = {
            'status': 'healthy',
            'timestamp': start_time.isoformat(),
            'deployment': {
                'environment': 'deployment' if (os.getenv('RENDER') or os.getenv('VERCEL') or os.getenv('HEROKU')) else 'local',
                'python_version': f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
                'flask_env': os.getenv('FLASK_ENV', 'development')
            }
        }

        # Try geocoding service initialization (with timeout protection)
        try:
            geocoding_service = GeocodingService()

            # Check postal code lookup status
            postal_status = len(geocoding_service.postal_code_lookup) > 0

            # Check Google Maps API status
            google_api_configured = geocoding_service.google_api_key is not None
            google_api_working = geocoding_service.geolocator is not None

            health_info['geocoding'] = {
                'postal_code_lookup': {
                    'enabled': postal_status,
                    'postal_codes_loaded': len(geocoding_service.postal_code_lookup)
                },
                'google_maps_api': {
                    'configured': google_api_configured,
                    'working': google_api_working,
                    'api_key_present': '****' + geocoding_service.google_api_key[-4:] if google_api_configured else None
                }
            }

        except Exception as geocoding_error:
            health_info['geocoding'] = {
                'error': str(geocoding_error),
                'status': 'geocoding_service_failed_but_app_running'
            }

        # Calculate response time
        end_time = datetime.now()
        health_info['response_time_ms'] = int((end_time - start_time).total_seconds() * 1000)

        return jsonify(health_info)

    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        # Basic request validation
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        # File extension validation
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload Excel files only.'}), 400

        # File size validation (50MB limit)
        MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
        if hasattr(file, 'content_length') and file.content_length > MAX_FILE_SIZE:
            return jsonify({'error': 'File too large. Maximum size is 50MB.'}), 413

        # Content length validation (for chunked requests)
        if request.content_length and request.content_length > MAX_FILE_SIZE:
            return jsonify({'error': 'File too large. Maximum size is 50MB.'}), 413

        # Filename validation
        if len(file.filename) > 255:
            return jsonify({'error': 'Filename too long. Maximum 255 characters.'}), 400

        # Security: prevent path traversal
        import re
        if re.search(r'[<>:"|?*]|\.\.', file.filename):
            return jsonify({'error': 'Invalid characters in filename.'}), 400

        # Extract geocoding preference (default: True)
        use_google_api = request.form.get('use_google_api', 'true').lower() == 'true'
        logger.info(f"Single upload - Google Maps API geocoding: {'ENABLED' if use_google_api else 'DISABLED'}")

        # Generate unique job ID
        job_id = str(uuid.uuid4())

        # Save uploaded file with original extension
        original_ext = os.path.splitext(file.filename)[1] or '.xlsx'  # Default to .xlsx if no extension
        input_filename = f"{job_id}_input{original_ext}"
        input_path = os.path.join(UPLOAD_FOLDER, input_filename)
        file.save(input_path)

        # Validate file content is actually Excel
        try:
            # Quick validation by attempting to read file structure
            pd.ExcelFile(input_path).sheet_names
            logger.info(f"File validation passed for job {job_id}: {file.filename}")
        except Exception as validation_error:
            # Clean up invalid file
            if os.path.exists(input_path):
                os.remove(input_path)
            logger.warning(f"File validation failed for {file.filename}: {validation_error}")
            return jsonify({'error': 'Invalid Excel file. File appears to be corrupted or not a valid Excel format.'}), 400

        # Transform file with multi-sheet support and geocoding preference
        result = ExcelTransformer.transform_excel_multi_sheet(input_path, PROCESSED_FOLDER, job_id, use_google_api)

        if result['success']:
            # Build download URLs for each output file
            download_urls = [f'/download/{job_id}/{filename}' for filename in result['output_files']]

            return jsonify({
                'job_id': job_id,
                'message': result['message'],
                'sheets_processed': result['sheets_processed'],
                'total_records': result['total_records'],
                'terminated_clinics_filtered': result['terminated_clinics_filtered'],
                'filtered_provider_codes': result.get('filtered_provider_codes', []),
                'output_files': result['output_files'],
                'download_urls': download_urls,
                'results': result['results'],
                'summary_stats': result['summary_stats']
            })
        else:
            return jsonify({
                'error': result['message'],
                'details': result.get('error_details', '')
            }), 500

    except Exception as e:
        return jsonify({
            'error': 'Internal server error',
            'details': str(e)
        }), 500

@app.route('/download/<job_id>/<filename>', methods=['GET'])
def download_specific_file(job_id, filename):
    """Download a specific output file by job ID and filename"""
    try:
        # Validate filename belongs to this job ID
        if not filename.startswith(job_id):
            return jsonify({'error': 'Invalid file for this job'}), 400

        output_path = os.path.join(PROCESSED_FOLDER, filename)

        if not os.path.exists(output_path):
            return jsonify({'error': 'File not found'}), 404

        # Extract sheet name from filename for better download name
        # Format: {job_id}_{sheet_name}.xlsx
        sheet_part = filename.replace(f"{job_id}_", "").replace(".xlsx", "")
        download_name = f"transformed_{sheet_part}.xlsx"

        return send_file(
            output_path,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download/<job_id>', methods=['GET'])
def download_file(job_id):
    """Legacy endpoint - download first available file or create zip if multiple files"""
    try:
        # Look for files with this job_id
        import glob
        pattern = os.path.join(PROCESSED_FOLDER, f"{job_id}_*.xlsx")
        matching_files = glob.glob(pattern)

        if not matching_files:
            return jsonify({'error': 'No files found for this job'}), 404

        if len(matching_files) == 1:
            # Single file - return directly
            output_path = matching_files[0]
            filename = os.path.basename(output_path)
            sheet_part = filename.replace(f"{job_id}_", "").replace(".xlsx", "")
            download_name = f"transformed_{sheet_part}.xlsx"

            return send_file(
                output_path,
                as_attachment=True,
                download_name=download_name,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            # Multiple files - create zip
            import zipfile
            import tempfile

            # Create temporary zip file
            temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')

            with zipfile.ZipFile(temp_zip.name, 'w') as zipf:
                for file_path in matching_files:
                    filename = os.path.basename(file_path)
                    sheet_part = filename.replace(f"{job_id}_", "").replace(".xlsx", "")
                    arc_name = f"transformed_{sheet_part}.xlsx"
                    zipf.write(file_path, arc_name)

            return send_file(
                temp_zip.name,
                as_attachment=True,
                download_name='transformed_templates.zip',
                mimetype='application/zip'
            )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/status/<job_id>', methods=['GET'])
def job_status(job_id):
    """Get status of processing job with support for multiple output files"""
    try:
        import glob

        # Look for files with this job_id
        pattern = os.path.join(PROCESSED_FOLDER, f"{job_id}_*.xlsx")
        matching_files = glob.glob(pattern)

        if not matching_files:
            return jsonify({'status': 'not_found'}), 404

        # Gather information about all files
        files_info = []
        total_size = 0
        earliest_time = None

        for file_path in matching_files:
            file_stats = os.stat(file_path)
            filename = os.path.basename(file_path)
            sheet_part = filename.replace(f"{job_id}_", "").replace(".xlsx", "")

            file_info = {
                'filename': filename,
                'sheet_name': sheet_part,
                'file_size': file_stats.st_size,
                'created_at': datetime.fromtimestamp(file_stats.st_ctime).isoformat(),
                'download_url': f'/download/{job_id}/{filename}'
            }
            files_info.append(file_info)
            total_size += file_stats.st_size

            if earliest_time is None or file_stats.st_ctime < earliest_time:
                earliest_time = file_stats.st_ctime

        return jsonify({
            'status': 'completed',
            'files_count': len(files_info),
            'total_size': total_size,
            'created_at': datetime.fromtimestamp(earliest_time).isoformat() if earliest_time else None,
            'files': files_info,
            'download_all_url': f'/download/{job_id}'  # Legacy endpoint for zip download
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

def filter_excluded_clinics(clinic_names_set, exclude_polyclinics, exclude_hospitals):
    """
    Filter clinic names by removing polyclinics and/or government hospitals.

    Args:
        clinic_names_set: Set of lowercase clinic names
        exclude_polyclinics: Boolean flag to remove polyclinics
        exclude_hospitals: Boolean flag to remove government hospitals

    Returns:
        tuple: (filtered_set, polyclinic_count, hospital_count)
    """
    filtered_set = clinic_names_set.copy()
    polyclinic_count = 0
    hospital_count = 0

    if exclude_polyclinics:
        polyclinics = {name for name in filtered_set if 'polyclinic' in name}
        polyclinic_count = len(polyclinics)
        filtered_set -= polyclinics
        logger.debug(f"Filtered {polyclinic_count} polyclinics")

    if exclude_hospitals:
        hospitals = {name for name in filtered_set
                     if any(hosp in name for hosp in GOVERNMENT_HOSPITALS)}
        hospital_count = len(hospitals)
        filtered_set -= hospitals
        logger.debug(f"Filtered {hospital_count} hospitals")

    return filtered_set, polyclinic_count, hospital_count

def extract_clinic_names_from_excel(file_path):
    """
    Extract clinic names from an Excel file by automatically detecting
    the clinic name column across all sheets.
    Returns a set of lowercase clinic names for case-insensitive matching.
    """
    clinic_names = set()

    try:
        # Read all sheets from the Excel file
        excel_file = pd.ExcelFile(file_path)
        logger.info(f"Processing {len(excel_file.sheet_names)} sheets from {os.path.basename(file_path)}")

        for sheet_name in excel_file.sheet_names:
            try:
                # Read the sheet without header first to find the actual header row
                df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                logger.info(f"Sheet '{sheet_name}': {len(df_raw)} rows, {len(df_raw.columns)} columns")

                # Find the header row by looking for 'Clinic Name' column
                header_row = None
                for idx in range(min(20, len(df_raw))):
                    row_values = [str(val) for val in df_raw.iloc[idx].values if pd.notna(val)]
                    row_text_lower = ' | '.join(row_values).lower()
                    # Look for specific header indicators
                    if ('clinic name' in row_text_lower or 'clinic_name' in row_text_lower) and \
                       ('s/n' in row_text_lower or 'code' in row_text_lower):
                        header_row = idx
                        logger.info(f"Found header row at index {idx} in sheet '{sheet_name}'")
                        logger.info(f"Header values: {row_values[:10]}")
                        break

                # Read with correct header
                if header_row is not None:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                else:
                    # Fallback: assume first row is header
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    logger.warning(f"Could not find header row in sheet '{sheet_name}', using default")

                # Find clinic name column with better pattern matching
                clinic_name_col = None

                # Log available columns for debugging
                logger.info(f"Available columns in sheet '{sheet_name}': {list(df.columns)[:10]}")

                for col in df.columns:
                    col_str = str(col).lower().strip()
                    # Check for various patterns
                    if (('clinic' in col_str and 'name' in col_str) or
                        col_str == 'clinic name' or
                        col_str == 'clinic' or
                        col_str == 'name' or
                        'clinic_name' in col_str.replace(' ', '_')):
                        clinic_name_col = col
                        break

                if clinic_name_col:
                    logger.info(f"Found clinic name column: '{clinic_name_col}' in sheet '{sheet_name}'")
                    # Extract clinic names and normalize (lowercase, strip whitespace)
                    for name in df[clinic_name_col].dropna():
                        if isinstance(name, str) and name.strip():
                            clinic_names.add(name.strip().lower())
                else:
                    logger.warning(f"Could not find clinic name column in sheet '{sheet_name}'")

            except Exception as sheet_error:
                logger.error(f"Error processing sheet '{sheet_name}': {sheet_error}")
                continue

        logger.info(f"Extracted {len(clinic_names)} unique clinic names from {os.path.basename(file_path)}")
        return clinic_names

    except Exception as e:
        logger.error(f"Error extracting clinic names from {file_path}: {e}")
        raise

def extract_clinics_with_visit_counts(file_path):
    """
    Extract clinic names with visit counts from Excel file.
    Similar to utilisation report logic but returns structured data for top N selection.

    Args:
        file_path: Path to the Excel file with utilization data

    Returns:
        list of tuples: [(clinic_name_normalized, visit_count), ...]
        sorted by visit_count descending
        Returns empty list if extraction fails
    """
    try:
        # Read all sheets from the Excel file
        excel_file = pd.ExcelFile(file_path)
        all_data = []

        logger.info(f"Extracting clinics with visit counts from {os.path.basename(file_path)}")
        logger.info(f"Processing {len(excel_file.sheet_names)} sheets")

        # Combine all sheets into single DataFrame
        for sheet_name in excel_file.sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                logger.info(f"Sheet '{sheet_name}': {len(df)} rows")
                all_data.append(df)
            except Exception as sheet_error:
                logger.error(f"Error reading sheet '{sheet_name}': {sheet_error}")
                continue

        if not all_data:
            logger.warning("No valid data found in Excel file for visit count extraction")
            return []

        # Combine all sheets
        combined_df = pd.concat(all_data, ignore_index=True)
        logger.info(f"Combined data: {len(combined_df)} total rows")

        # Find clinic name column
        clinic_col = None
        for col in combined_df.columns:
            col_str = str(col).lower().strip()
            if 'clinic' in col_str and 'name' in col_str:
                clinic_col = col
                break

        if clinic_col is None:
            # Fallback: look for just 'clinic' or 'name'
            for col in combined_df.columns:
                col_str = str(col).lower().strip()
                if col_str in ['clinic name', 'clinic', 'name']:
                    clinic_col = col
                    break

        if clinic_col is None:
            logger.warning("Could not find 'Clinic Name' column, falling back to alphabetical ordering")
            return []

        logger.info(f"Using clinic name column: '{clinic_col}'")

        # Clean data: remove rows with empty clinic names
        combined_df = combined_df[combined_df[clinic_col].notna()]
        combined_df = combined_df[combined_df[clinic_col].astype(str).str.strip() != '']

        # Normalize clinic names (strip whitespace and lowercase)
        combined_df['Clinic Name Normalized'] = combined_df[clinic_col].astype(str).str.strip().str.lower()

        logger.info(f"After cleaning: {len(combined_df)} rows")

        # Group by clinic name and count visits
        visit_counts = combined_df.groupby('Clinic Name Normalized').size().reset_index(name='visit_count')

        # Sort by visit count descending
        visit_counts = visit_counts.sort_values('visit_count', ascending=False)

        # Convert to list of tuples
        result = [(row['Clinic Name Normalized'], row['visit_count'])
                  for _, row in visit_counts.iterrows()]

        logger.info(f"Extracted {len(result)} clinics with visit counts")
        if result:
            logger.info(f"Top clinic: {result[0][0]} with {result[0][1]} visits")

        return result

    except Exception as e:
        logger.error(f"Error extracting clinics with visit counts: {e}\n{traceback.format_exc()}")
        logger.warning("Falling back to empty list (will use alphabetical ordering)")
        return []

def generate_utilisation_report(file_path, exclude_polyclinics=False, exclude_hospitals=False):
    """
    Generate utilisation report from Excel file with clinic visit and amount data.

    Args:
        file_path: Path to the Excel file with utilization data
        exclude_polyclinics: Boolean to filter out polyclinics
        exclude_hospitals: Boolean to filter out government hospitals

    Returns:
        tuple: (report_filename, total_visits, total_amount, clinic_count)
    """
    try:
        # Read all sheets from the Excel file
        excel_file = pd.ExcelFile(file_path)
        all_data = []

        logger.info(f"Generating utilisation report from {os.path.basename(file_path)}")
        logger.info(f"Processing {len(excel_file.sheet_names)} sheets")

        # Combine all sheets into single DataFrame
        for sheet_name in excel_file.sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                logger.info(f"Sheet '{sheet_name}': {len(df)} rows")
                all_data.append(df)
            except Exception as sheet_error:
                logger.error(f"Error reading sheet '{sheet_name}': {sheet_error}")
                continue

        if not all_data:
            raise ValueError("No valid data found in Excel file")

        # Combine all sheets
        combined_df = pd.concat(all_data, ignore_index=True)
        logger.info(f"Combined data: {len(combined_df)} total rows")

        # Find clinic name column
        clinic_col = None
        for col in combined_df.columns:
            col_str = str(col).lower().strip()
            if 'clinic' in col_str and 'name' in col_str:
                clinic_col = col
                break

        if clinic_col is None:
            # Fallback: look for just 'clinic' or 'name'
            for col in combined_df.columns:
                col_str = str(col).lower().strip()
                if col_str in ['clinic name', 'clinic', 'name']:
                    clinic_col = col
                    break

        if clinic_col is None:
            raise ValueError("Could not find 'Clinic Name' column in Excel file")

        logger.info(f"Using clinic name column: '{clinic_col}'")

        # Find incurred amount column
        amount_col = None
        for col in combined_df.columns:
            col_str = str(col).lower().strip()
            if 'incurred' in col_str and ('amt' in col_str or 'amount' in col_str):
                amount_col = col
                break

        if amount_col is None:
            raise ValueError("Could not find 'INCURRED AMT. ($)' column in Excel file")

        logger.info(f"Using amount column: '{amount_col}'")

        # Clean data: remove rows with empty clinic names
        combined_df = combined_df[combined_df[clinic_col].notna()]
        combined_df = combined_df[combined_df[clinic_col].astype(str).str.strip() != '']

        # Normalize clinic names (strip whitespace)
        combined_df['Clinic Name Normalized'] = combined_df[clinic_col].astype(str).str.strip()

        # Convert amount to numeric (coerce errors to 0)
        combined_df['Amount Numeric'] = pd.to_numeric(combined_df[amount_col], errors='coerce').fillna(0)

        logger.info(f"After cleaning: {len(combined_df)} rows")

        # Apply filters if requested
        if exclude_polyclinics or exclude_hospitals:
            # Create lowercase version for filtering
            combined_df['Clinic Name Lower'] = combined_df['Clinic Name Normalized'].str.lower()

            before_filter = len(combined_df)

            if exclude_polyclinics:
                combined_df = combined_df[~combined_df['Clinic Name Lower'].str.contains('polyclinic', na=False)]
                logger.info(f"After polyclinic filter: {len(combined_df)} rows (removed {before_filter - len(combined_df)})")
                before_filter = len(combined_df)

            if exclude_hospitals:
                # Filter out government hospitals
                hospital_mask = combined_df['Clinic Name Lower'].apply(
                    lambda x: any(hosp in x for hosp in GOVERNMENT_HOSPITALS)
                )
                combined_df = combined_df[~hospital_mask]
                logger.info(f"After hospital filter: {len(combined_df)} rows (removed {before_filter - len(combined_df)})")

        if len(combined_df) == 0:
            raise ValueError("No data remaining after filtering")

        # Group by clinic name and aggregate
        utilisation_summary = combined_df.groupby('Clinic Name Normalized').agg({
            'Amount Numeric': 'sum',
            clinic_col: 'count'  # Count visits
        }).reset_index()

        # Rename columns
        utilisation_summary.columns = ['Clinic Name', 'Total Utilisation Amount ($)', 'Number of Visits']

        # Reorder columns to match expected format
        utilisation_summary = utilisation_summary[['Clinic Name', 'Number of Visits', 'Total Utilisation Amount ($)']]

        # Sort by Number of Visits (descending)
        utilisation_summary = utilisation_summary.sort_values('Number of Visits', ascending=False)

        # Reset index
        utilisation_summary = utilisation_summary.reset_index(drop=True)

        logger.info(f"Utilisation summary: {len(utilisation_summary)} clinics")

        # Calculate totals
        total_visits = int(utilisation_summary['Number of Visits'].sum())
        total_amount = float(utilisation_summary['Total Utilisation Amount ($)'].sum())
        clinic_count = len(utilisation_summary)

        logger.info(f"Totals: {total_visits} visits, ${total_amount:.2f}, {clinic_count} clinics")

        # Generate Excel file
        report_filename = f"utilisation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(PROCESSED_FOLDER, report_filename)

        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            utilisation_summary.to_excel(writer, sheet_name='Clinic Summary', index=False)

            # Get worksheet for formatting
            ws = writer.sheets['Clinic Summary']

            # Format currency column (C) with dollar sign
            for row in range(2, len(utilisation_summary) + 2):
                cell = ws[f'C{row}']
                cell.number_format = '$#,##0.00'

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width

            # Bold headers
            from openpyxl.styles import Font
            for cell in ws[1]:
                cell.font = Font(bold=True)

        logger.info(f"Utilisation report saved: {report_filename}")

        return report_filename, total_visits, total_amount, clinic_count

    except Exception as e:
        logger.error(f"Error generating utilisation report: {e}\n{traceback.format_exc()}")
        raise

@app.route('/match-clinics', methods=['POST'])
def match_clinics():
    """
    Compare clinic names from two Excel files and return:
    - Matched clinics (found in both files)
    - Unmatched in base (only in base file)
    - Unmatched in comparison (only in comparison file)
    Results are automatically downloaded as an Excel file with 3 sheets.
    """
    try:
        # Validate request
        if 'base_file' not in request.files or 'comparison_file' not in request.files:
            return jsonify({'error': 'Both base_file and comparison_file are required'}), 400

        base_file = request.files['base_file']
        comparison_file = request.files['comparison_file']

        if base_file.filename == '' or comparison_file.filename == '':
            return jsonify({'error': 'Both files must be selected'}), 400

        # Validate file extensions
        for file in [base_file, comparison_file]:
            if not file.filename.lower().endswith(('.xlsx', '.xls')):
                return jsonify({'error': f'Invalid file format for {file.filename}. Only Excel files are supported.'}), 400

        # Save uploaded files temporarily
        job_id = str(uuid.uuid4())
        base_path = os.path.join(UPLOAD_FOLDER, f"{job_id}_base_{base_file.filename}")
        comparison_path = os.path.join(UPLOAD_FOLDER, f"{job_id}_comparison_{comparison_file.filename}")

        base_file.save(base_path)
        comparison_file.save(comparison_path)

        # Get filter parameters
        exclude_polyclinics = request.form.get('exclude_polyclinics', 'false').lower() == 'true'
        exclude_hospitals = request.form.get('exclude_hospitals', 'false').lower() == 'true'
        generate_report = request.form.get('generate_report', 'false').lower() == 'true'
        top_n_filter = request.form.get('top_n_filter', None)  # 'top10', 'top20', or None

        logger.info(f"Matching clinics: Base='{base_file.filename}', Comparison='{comparison_file.filename}'")
        logger.info(f"Filters: exclude_polyclinics={exclude_polyclinics}, exclude_hospitals={exclude_hospitals}")
        logger.info(f"Generate utilisation report: {generate_report}")
        logger.info(f"Top N filter: {top_n_filter}")

        # Extract clinic names from both files
        # If top N filter is enabled, extract with visit counts; otherwise, just extract names
        if top_n_filter:
            base_clinics_with_counts = extract_clinics_with_visit_counts(base_path)
            logger.info(f"Extracted {len(base_clinics_with_counts)} clinics with visit counts from base file")
        else:
            base_clinics = extract_clinic_names_from_excel(base_path)
            logger.info(f"Extracted {len(base_clinics)} clinic names from base file")

        comparison_clinics = extract_clinic_names_from_excel(comparison_path)

        # Initialize filter counters
        base_polyclinics_filtered = 0
        base_hospitals_filtered = 0
        comparison_polyclinics_filtered = 0
        comparison_hospitals_filtered = 0

        # Initialize top N variables
        top_n_clinic_names = set()
        top_n_count = 0
        top_n_warning = None

        # Handle top N filtering if enabled
        if top_n_filter:
            # Determine the N value
            n_value = 10 if top_n_filter == 'top10' else 20

            # Apply filters to the list of (clinic, count) tuples if requested
            if exclude_polyclinics or exclude_hospitals:
                filtered_clinics_with_counts = []
                for clinic_name, visit_count in base_clinics_with_counts:
                    exclude_this = False

                    # Check polyclinic filter
                    if exclude_polyclinics and 'polyclinic' in clinic_name:
                        base_polyclinics_filtered += 1
                        exclude_this = True

                    # Check hospital filter
                    if exclude_hospitals and any(hosp in clinic_name for hosp in GOVERNMENT_HOSPITALS):
                        base_hospitals_filtered += 1
                        exclude_this = True

                    if not exclude_this:
                        filtered_clinics_with_counts.append((clinic_name, visit_count))

                base_clinics_with_counts = filtered_clinics_with_counts
                logger.info(f"Base: Filtered {base_polyclinics_filtered} polyclinics, {base_hospitals_filtered} hospitals from visit data")

            # Select top N clinics from filtered list
            top_n_list = base_clinics_with_counts[:n_value]
            top_n_clinic_names = {clinic[0] for clinic in top_n_list}
            top_n_count = len(top_n_clinic_names)

            logger.info(f"Selected top {n_value} clinics: got {top_n_count} clinics")

            # Check if we have fewer clinics than requested
            if top_n_count < n_value and top_n_count > 0:
                top_n_warning = f"Only {top_n_count} clinics available after filters (requested top {n_value})"
                logger.warning(top_n_warning)

            # Create base_clinics set from all clinics for regular matching
            base_clinics = {clinic[0] for clinic in base_clinics_with_counts}

        else:
            # Regular extraction (no top N)
            # Apply filters if requested
            if exclude_polyclinics or exclude_hospitals:
                base_clinics_filtered, base_polyclinics_filtered, base_hospitals_filtered = filter_excluded_clinics(
                    base_clinics, exclude_polyclinics, exclude_hospitals
                )
                base_clinics = base_clinics_filtered
                logger.info(f"Base: Filtered {base_polyclinics_filtered} polyclinics, {base_hospitals_filtered} hospitals")

        # Filter comparison clinics (always done the same way)
        if exclude_polyclinics or exclude_hospitals:
            comparison_clinics_filtered, comparison_polyclinics_filtered, comparison_hospitals_filtered = filter_excluded_clinics(
                comparison_clinics, exclude_polyclinics, exclude_hospitals
            )
            comparison_clinics = comparison_clinics_filtered
            logger.info(f"Comparison: Filtered {comparison_polyclinics_filtered} polyclinics, {comparison_hospitals_filtered} hospitals")

        # Log final counts
        logger.info(f"Final clinic counts: Base={len(base_clinics)}, Comparison={len(comparison_clinics)}")

        # Validate that we have clinics remaining after filtering
        if len(base_clinics) == 0 or len(comparison_clinics) == 0:
            return jsonify({
                'error': 'All clinics were filtered out. Please adjust filter settings.',
                'base_total': len(base_clinics) if not top_n_filter else len(base_clinics_with_counts),
                'comparison_total': len(comparison_clinics),
                'base_filtered_out': base_polyclinics_filtered + base_hospitals_filtered,
                'comparison_filtered_out': comparison_polyclinics_filtered + comparison_hospitals_filtered
            }), 400

        # Perform matching (case-insensitive) - always perform full matching
        matched = base_clinics.intersection(comparison_clinics)
        unmatched_base = base_clinics - comparison_clinics
        unmatched_comparison = comparison_clinics - base_clinics

        logger.info(f"Match results: {len(matched)} matched, {len(unmatched_base)} unmatched in base, {len(unmatched_comparison)} unmatched in comparison")

        # Calculate top N matching if enabled
        top_n_matched = set()
        top_n_unmatched = set()
        if top_n_filter and top_n_count > 0:
            top_n_matched = top_n_clinic_names.intersection(comparison_clinics)
            top_n_unmatched = top_n_clinic_names - comparison_clinics
            logger.info(f"Top {n_value} match results: {len(top_n_matched)} matched, {len(top_n_unmatched)} unmatched")

        # Create results Excel file with 3 sheets
        results_filename = f"clinic_match_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        results_path = os.path.join(PROCESSED_FOLDER, results_filename)

        with pd.ExcelWriter(results_path, engine='openpyxl') as writer:
            # Sheet 1: Matched clinics
            matched_df = pd.DataFrame({
                'Clinic Name': sorted(matched, key=str.lower)
            })
            matched_df.index = range(1, len(matched_df) + 1)
            matched_df.to_excel(writer, sheet_name='Matched', index=True, index_label='S/N')

            # Sheet 2: Unmatched in Base
            unmatched_base_df = pd.DataFrame({
                'Clinic Name': sorted(unmatched_base, key=str.lower)
            })
            unmatched_base_df.index = range(1, len(unmatched_base_df) + 1)
            unmatched_base_df.to_excel(writer, sheet_name='Unmatched in Base', index=True, index_label='S/N')

            # Sheet 3: Unmatched in Comparison
            unmatched_comparison_df = pd.DataFrame({
                'Clinic Name': sorted(unmatched_comparison, key=str.lower)
            })
            unmatched_comparison_df.index = range(1, len(unmatched_comparison_df) + 1)
            unmatched_comparison_df.to_excel(writer, sheet_name='Unmatched in Comparison', index=True, index_label='S/N')

        # Generate utilisation report if requested
        report_filename = None
        total_visits = 0
        total_amount = 0.0
        clinic_count = 0

        if generate_report:
            try:
                logger.info("Generating utilisation report from base file...")
                report_filename, total_visits, total_amount, clinic_count = generate_utilisation_report(
                    base_path, exclude_polyclinics, exclude_hospitals
                )
                logger.info(f"Utilisation report generated: {report_filename}")
            except Exception as report_error:
                logger.error(f"Failed to generate utilisation report: {report_error}")
                # Continue with matching results even if report generation fails

        # Clean up uploaded files
        try:
            os.remove(base_path)
            os.remove(comparison_path)
        except Exception as cleanup_error:
            logger.warning(f"Failed to clean up uploaded files: {cleanup_error}")

        # Return results with filter breakdown
        response_data = {
            'success': True,
            'matched_count': len(matched),
            'unmatched_base_count': len(unmatched_base),
            'unmatched_comparison_count': len(unmatched_comparison),
            'download_filename': results_filename,
            'base_total': len(base_clinics),
            'comparison_total': len(comparison_clinics),
            'base_polyclinics_filtered': base_polyclinics_filtered,
            'base_hospitals_filtered': base_hospitals_filtered,
            'comparison_polyclinics_filtered': comparison_polyclinics_filtered,
            'comparison_hospitals_filtered': comparison_hospitals_filtered,
            'base_total_after_filter': len(base_clinics),
            'comparison_total_after_filter': len(comparison_clinics)
        }

        # Add utilisation report data if generated
        if report_filename:
            response_data['utilisation_report_filename'] = report_filename
            response_data['total_visits'] = total_visits
            response_data['total_amount'] = total_amount
            response_data['clinic_count'] = clinic_count

        # Add top N matching data if enabled
        if top_n_filter:
            response_data['top_n_enabled'] = True
            response_data['top_n_filter_type'] = top_n_filter
            response_data['top_n_count'] = top_n_count
            response_data['top_n_matched_count'] = len(top_n_matched)
            response_data['top_n_unmatched_count'] = len(top_n_unmatched)
            if top_n_warning:
                response_data['top_n_warning'] = top_n_warning
        else:
            response_data['top_n_enabled'] = False

        return jsonify(response_data)

    except Exception as e:
        logger.error(f"Error in match_clinics: {e}\n{traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

@app.route('/download-match/<filename>', methods=['GET'])
def download_match_result(filename):
    """Download clinic matching results file"""
    try:
        file_path = os.path.join(PROCESSED_FOLDER, filename)

        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404

        # Verify it's a safe filename (no path traversal)
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'error': 'Invalid filename'}), 400

        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Error downloading match result: {e}")
        return jsonify({'error': str(e)}), 500

def startup_check():
    """Perform startup checks to ensure all dependencies are available"""
    try:
        logger.info("Starting Excel Transformer Backend...")
        logger.info(f"Python version: {sys.version}")
        logger.info(f"Flask environment: {os.environ.get('FLASK_ENV', 'development')}")
        logger.info(f"Concurrent processing: {'enabled' if CONCURRENT_SUPPORT else 'disabled (sequential fallback)'}")

        # Test basic imports
        import pandas as pd
        logger.info(f"Pandas version: {pd.__version__}")

        # Test geocoding service initialization (non-blocking)
        try:
            geocoding_service = GeocodingService()
            logger.info("Geocoding service initialized successfully")
        except Exception as e:
            logger.warning(f"Geocoding service initialization issue: {e}")

        logger.info("Startup checks completed successfully")
        return True

    except Exception as e:
        logger.error(f"Startup check failed: {e}")
        return False

# Perform startup check
startup_check()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(debug=debug, host='0.0.0.0', port=port)