"""
GP Panel Comparison Processor

Extracts clinic data from HSBC Fullerton GP Panel Excel files and compares
previous vs current month listings to identify added, removed, and updated clinics.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional
import os
import re
import logging

logger = logging.getLogger(__name__)

# Sheet configuration with column mappings
SHEET_CONFIG = {
    'gp_sgp': {
        'pattern': 'GP (SGP)',
        'label': 'GP (Singapore)',
        'header_row': 19,
        'data_start': 20,
        'columns': {
            'provider_code': 4,
            'clinic_name': 5,
            'region': 2,
            'area': 3,
            'address': 6,
            'tel': 7,
            'mon_fri': 8,
            'mon_fri_eve': 9,
            'sat': 10,
            'sun': 11,
            'ph': 12,
            'remarks': 13
        }
    },
    'gp_msia': {
        'pattern': 'GP (Msia)',
        'label': 'GP (Malaysia/JB)',
        'header_row': 18,
        'data_start': 19,
        'columns': {
            'provider_code': 4,
            'clinic_name': 5,
            'region': 3,
            'area': 2,
            'address': 6,
            'tel': 7,
            'mon_fri': 8,
            'mon_fri_eve': 9,
            'sat': 10,
            'sun': 11,
            'ph': 12,
            'remarks': 13
        }
    },
    'tcm': {
        'pattern': 'TCM',
        'label': 'TCM',
        'header_row': 18,
        'data_start': 19,
        'columns': {
            'provider_code': 4,
            'clinic_name': 5,
            'region': 2,
            'area': 3,
            'address': 6,
            'tel': 7,
            'mon_fri': 8,
            'mon_fri_eve': 9,
            'sat': 10,
            'sun': 11,
            'ph': 12,
            'remarks': 13
        }
    }
}


@dataclass
class PanelClinic:
    """Represents a clinic from the GP Panel listing."""
    provider_code: str
    clinic_name: str
    region: str
    area: str
    address: str
    tel: str
    operating_hours: Dict[str, str] = field(default_factory=dict)
    remarks: str = ''
    sheet_type: str = ''

    def get_unique_id(self) -> str:
        """Get unique identifier for comparison."""
        if self.provider_code and self.provider_code.upper().startswith('FHG'):
            return self.provider_code.upper().strip()
        return self.clinic_name.upper().strip()

    def to_dict(self) -> dict:
        """Convert to dictionary for JSON serialization."""
        return {
            'provider_code': self.provider_code,
            'clinic_name': self.clinic_name,
            'region': self.region,
            'area': self.area,
            'address': self.address,
            'tel': self.tel,
            'operating_hours': self.operating_hours,
            'remarks': self.remarks,
            'sheet_type': self.sheet_type
        }


@dataclass
class PanelComparison:
    """Holds comparison results for a panel type."""
    added: List[PanelClinic] = field(default_factory=list)
    removed: List[PanelClinic] = field(default_factory=list)
    updated: List[Tuple[PanelClinic, PanelClinic, List[str]]] = field(default_factory=list)

    def to_dict(self) -> dict:
        """Convert to dictionary for JSON serialization."""
        return {
            'added': [c.to_dict() for c in self.added],
            'removed': [c.to_dict() for c in self.removed],
            'updated': [
                {
                    'old': old.to_dict(),
                    'new': new.to_dict(),
                    'changes': changes
                }
                for old, new, changes in self.updated
            ]
        }


def _normalize_value(value) -> str:
    """Normalize a cell value for comparison."""
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        return str(value).strip()
    return str(value).strip().replace('\n', ' ').replace('\r', '')


def _find_sheet_by_pattern(workbook: openpyxl.Workbook, pattern: str) -> Optional[str]:
    """Find sheet name matching pattern (case-insensitive, handles trailing spaces)."""
    pattern_lower = pattern.lower()
    for sheet_name in workbook.sheetnames:
        if pattern_lower in sheet_name.lower():
            return sheet_name
    return None


def extract_panel_clinics(filepath: str) -> Dict[str, List[PanelClinic]]:
    """
    Extract clinics from all 3 sheets of a GP Panel Excel file.

    Args:
        filepath: Path to the GP Panel Excel file

    Returns:
        Dictionary mapping sheet_type to list of PanelClinic objects
    """
    result = {}

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        logger.error(f"Failed to open workbook: {e}")
        raise ValueError(f"Cannot open Excel file: {e}")

    for sheet_type, config in SHEET_CONFIG.items():
        sheet_name = _find_sheet_by_pattern(wb, config['pattern'])

        if not sheet_name:
            logger.warning(f"Sheet matching '{config['pattern']}' not found")
            result[sheet_type] = []
            continue

        ws = wb[sheet_name]
        clinics = []
        cols = config['columns']
        data_start = config['data_start']
        seen_ids = set()

        for row_num in range(data_start, ws.max_row + 1):
            provider_code = _normalize_value(ws.cell(row=row_num, column=cols['provider_code']).value)
            clinic_name = _normalize_value(ws.cell(row=row_num, column=cols['clinic_name']).value)

            if not clinic_name:
                continue

            clinic = PanelClinic(
                provider_code=provider_code,
                clinic_name=clinic_name,
                region=_normalize_value(ws.cell(row=row_num, column=cols['region']).value),
                area=_normalize_value(ws.cell(row=row_num, column=cols['area']).value),
                address=_normalize_value(ws.cell(row=row_num, column=cols['address']).value),
                tel=_normalize_value(ws.cell(row=row_num, column=cols['tel']).value),
                operating_hours={
                    'mon_fri': _normalize_value(ws.cell(row=row_num, column=cols['mon_fri']).value),
                    'mon_fri_eve': _normalize_value(ws.cell(row=row_num, column=cols['mon_fri_eve']).value),
                    'sat': _normalize_value(ws.cell(row=row_num, column=cols['sat']).value),
                    'sun': _normalize_value(ws.cell(row=row_num, column=cols['sun']).value),
                    'ph': _normalize_value(ws.cell(row=row_num, column=cols['ph']).value),
                },
                remarks=_normalize_value(ws.cell(row=row_num, column=cols['remarks']).value),
                sheet_type=sheet_type
            )

            uid = clinic.get_unique_id()
            if uid in seen_ids:
                logger.warning(f"Duplicate ID '{uid}' in {sheet_type}, keeping first occurrence")
                continue

            seen_ids.add(uid)
            clinics.append(clinic)

        result[sheet_type] = clinics
        logger.info(f"Extracted {len(clinics)} clinics from {sheet_name}")

    wb.close()
    return result


def _detect_changes(prev: PanelClinic, curr: PanelClinic) -> List[str]:
    """Detect specific field changes between two clinic records."""
    changes = []

    if prev.clinic_name.upper() != curr.clinic_name.upper():
        changes.append(f"NAME: {prev.clinic_name} → {curr.clinic_name}")

    if prev.address.upper() != curr.address.upper():
        changes.append("ADDRESS changed")

    if prev.tel != curr.tel:
        changes.append(f"TEL: {prev.tel} → {curr.tel}")

    for key in ['mon_fri', 'mon_fri_eve', 'sat', 'sun', 'ph']:
        prev_val = prev.operating_hours.get(key, '')
        curr_val = curr.operating_hours.get(key, '')
        if prev_val != curr_val:
            label = key.upper().replace('_', ' ')
            changes.append(f"{label}: {prev_val or 'N/A'} → {curr_val or 'N/A'}")

    if prev.remarks != curr.remarks:
        changes.append("REMARKS changed")

    return changes


def compare_panels(
    prev_clinics: Dict[str, List[PanelClinic]],
    curr_clinics: Dict[str, List[PanelClinic]]
) -> Dict[str, PanelComparison]:
    """
    Compare two panel extractions and return changes per sheet type.

    Args:
        prev_clinics: Previous month's extracted clinics
        curr_clinics: Current month's extracted clinics

    Returns:
        Dictionary mapping sheet_type to PanelComparison objects
    """
    result = {}

    for sheet_type in SHEET_CONFIG.keys():
        prev_list = prev_clinics.get(sheet_type, [])
        curr_list = curr_clinics.get(sheet_type, [])

        prev_map = {c.get_unique_id(): c for c in prev_list}
        curr_map = {c.get_unique_id(): c for c in curr_list}

        prev_ids = set(prev_map.keys())
        curr_ids = set(curr_map.keys())

        comparison = PanelComparison()

        for uid in curr_ids - prev_ids:
            comparison.added.append(curr_map[uid])

        for uid in prev_ids - curr_ids:
            comparison.removed.append(prev_map[uid])

        for uid in prev_ids & curr_ids:
            prev_clinic = prev_map[uid]
            curr_clinic = curr_map[uid]
            changes = _detect_changes(prev_clinic, curr_clinic)
            if changes:
                comparison.updated.append((prev_clinic, curr_clinic, changes))

        result[sheet_type] = comparison
        logger.info(
            f"{sheet_type}: +{len(comparison.added)} -{len(comparison.removed)} "
            f"~{len(comparison.updated)}"
        )

    return result


def generate_comparison_excel(
    comparison: Dict[str, PanelComparison],
    prev_clinics: Dict[str, List[PanelClinic]],
    curr_clinics: Dict[str, List[PanelClinic]],
    output_path: str
) -> str:
    """
    Generate Excel report with comparison results.

    Args:
        comparison: Comparison results per sheet type
        prev_clinics: Previous month's clinics for counts
        curr_clinics: Current month's clinics for counts
        output_path: Path for output Excel file

    Returns:
        Path to generated file
    """
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    added_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    removed_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    updated_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws_summary = wb.active
    ws_summary.title = 'Summary'

    summary_headers = ['Panel', 'Previous', 'Current', 'Added', 'Removed', 'Updated']
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    row = 2
    for sheet_type, config in SHEET_CONFIG.items():
        prev_count = len(prev_clinics.get(sheet_type, []))
        curr_count = len(curr_clinics.get(sheet_type, []))
        comp = comparison.get(sheet_type, PanelComparison())

        ws_summary.cell(row=row, column=1, value=config['label']).border = thin_border
        ws_summary.cell(row=row, column=2, value=prev_count).border = thin_border
        ws_summary.cell(row=row, column=3, value=curr_count).border = thin_border
        ws_summary.cell(row=row, column=4, value=len(comp.added)).border = thin_border
        ws_summary.cell(row=row, column=5, value=len(comp.removed)).border = thin_border
        ws_summary.cell(row=row, column=6, value=len(comp.updated)).border = thin_border

        ws_summary.cell(row=row, column=4).fill = added_fill if comp.added else None
        ws_summary.cell(row=row, column=5).fill = removed_fill if comp.removed else None
        ws_summary.cell(row=row, column=6).fill = updated_fill if comp.updated else None

        row += 1

    for col in range(1, 7):
        ws_summary.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    for sheet_type, config in SHEET_CONFIG.items():
        comp = comparison.get(sheet_type, PanelComparison())

        if not comp.added and not comp.removed and not comp.updated:
            continue

        ws = wb.create_sheet(title=config['label'][:31])

        change_headers = ['Change Type', 'Provider Code', 'Clinic Name', 'Area', 'Region',
                         'Address', 'Tel', 'Change Details']
        for col, header in enumerate(change_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row = 2

        for clinic in comp.added:
            ws.cell(row=row, column=1, value='ADDED').fill = added_fill
            ws.cell(row=row, column=2, value=clinic.provider_code)
            ws.cell(row=row, column=3, value=clinic.clinic_name)
            ws.cell(row=row, column=4, value=clinic.area)
            ws.cell(row=row, column=5, value=clinic.region)
            ws.cell(row=row, column=6, value=clinic.address)
            ws.cell(row=row, column=7, value=clinic.tel)
            ws.cell(row=row, column=8, value='')
            for c in range(1, 9):
                ws.cell(row=row, column=c).border = thin_border
                ws.cell(row=row, column=c).fill = added_fill
            row += 1

        for clinic in comp.removed:
            ws.cell(row=row, column=1, value='REMOVED').fill = removed_fill
            ws.cell(row=row, column=2, value=clinic.provider_code)
            ws.cell(row=row, column=3, value=clinic.clinic_name)
            ws.cell(row=row, column=4, value=clinic.area)
            ws.cell(row=row, column=5, value=clinic.region)
            ws.cell(row=row, column=6, value=clinic.address)
            ws.cell(row=row, column=7, value=clinic.tel)
            ws.cell(row=row, column=8, value='')
            for c in range(1, 9):
                ws.cell(row=row, column=c).border = thin_border
                ws.cell(row=row, column=c).fill = removed_fill
            row += 1

        for old, new, changes in comp.updated:
            ws.cell(row=row, column=1, value='UPDATED').fill = updated_fill
            ws.cell(row=row, column=2, value=new.provider_code)
            ws.cell(row=row, column=3, value=new.clinic_name)
            ws.cell(row=row, column=4, value=new.area)
            ws.cell(row=row, column=5, value=new.region)
            ws.cell(row=row, column=6, value=new.address)
            ws.cell(row=row, column=7, value=new.tel)
            ws.cell(row=row, column=8, value='; '.join(changes))
            for c in range(1, 9):
                ws.cell(row=row, column=c).border = thin_border
                ws.cell(row=row, column=c).fill = updated_fill
            row += 1

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 50

    wb.save(output_path)
    wb.close()

    return output_path


def generate_email_draft(comparison: Dict[str, PanelComparison]) -> str:
    """
    Generate email draft text announcing panel changes.

    Args:
        comparison: Comparison results per sheet type

    Returns:
        Formatted email text
    """
    lines = [
        "Hi All,",
        "",
        "Please refer to attached for the latest GP Panel listing.",
        "",
        "Updates are as follows :-",
        ""
    ]

    msia_comp = comparison.get('gp_msia', PanelComparison())
    lines.append("Msia GP :-")
    lines.append("Added :-")
    if msia_comp.added:
        for clinic in msia_comp.added:
            lines.append(f"- {clinic.clinic_name}")
    else:
        lines.append("None")
    lines.append("")
    lines.append("Removed :-")
    if msia_comp.removed:
        for clinic in msia_comp.removed:
            lines.append(f"- {clinic.clinic_name}")
    else:
        lines.append("None")
    lines.append("")

    sgp_comp = comparison.get('gp_sgp', PanelComparison())
    lines.append("Singapore GP :-")
    lines.append("Added :-")
    if sgp_comp.added:
        for clinic in sgp_comp.added:
            lines.append(f"- {clinic.clinic_name}")
    else:
        lines.append("None")
    lines.append("")
    lines.append("Removed :-")
    if sgp_comp.removed:
        for clinic in sgp_comp.removed:
            lines.append(f"- {clinic.clinic_name}")
    else:
        lines.append("None")
    lines.append("")

    tcm_comp = comparison.get('tcm', PanelComparison())
    if tcm_comp.added or tcm_comp.removed:
        lines.append("TCM :-")
        lines.append("Added :-")
        if tcm_comp.added:
            for clinic in tcm_comp.added:
                lines.append(f"- {clinic.clinic_name}")
        else:
            lines.append("None")
        lines.append("")
        lines.append("Removed :-")
        if tcm_comp.removed:
            for clinic in tcm_comp.removed:
                lines.append(f"- {clinic.clinic_name}")
        else:
            lines.append("None")
        lines.append("")

    lines.append("Thank you!")

    return '\n'.join(lines)
